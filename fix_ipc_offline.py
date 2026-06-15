#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Offline IPC Fix Tool for Batch Summary

This tool re-calculates td_ipc_avg using the CORRECT formula:
    IPC_avg = sum(instructions) / sum(cycles)

instead of the previous incorrect simple average formula.

Usage:
    python fix_ipc_offline.py --result-dir <batch_result_base_dir>
    python fix_ipc_offline.py --result-dir ./batch_results_20260615

Steps:
    1. Find all devkit_top_down.log files in subdirectories
    2. Re-parse using correct IPC formula
    3. Update analysis_report.xlsx in each qemu_monitor/ subdirectory
    4. Re-generate batch_summary.xlsx with corrected data
"""

import os
import sys
import argparse
import re
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Tuple
from collections import defaultdict

# Import the corrected parser from qemu_monitor package
try:
    from qemu_monitor.parsers import parse_devkit_top_down
    QEMU_MONITOR_AVAILABLE = True
except ImportError:
    print("Warning: qemu_monitor.parsers not available, using built-in parser")
    QEMU_MONITOR_AVAILABLE = False


def parse_devkit_top_down_offline(log_path: str) -> Dict:
    """Parse DevKit top-down log with CORRECT IPC formula

    This is the corrected version that uses:
        IPC_avg = sum(instructions) / sum(cycles)

    Args:
        log_path: Path to devkit_top_down.log

    Returns:
        dict with all metrics including corrected ipc_avg
    """
    if not os.path.exists(log_path):
        return {'error': f'File not found: {log_path}', 'report_count': 0}

    try:
        with open(log_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
    except Exception as e:
        return {'error': f'Failed to read file: {e}', 'report_count': 0}

    # Split into reports
    reports = re.split(r'TOP-DOWN Summary Report-\d+', content)

    result = {
        'cycles': [],
        'instructions': [],
        'ipc': [],
        'bad_speculation': [],
        'frontend_bound': [],
        'retiring': [],
        'backend_bound': [],
        'l3_bound': [],
        'mem_bound': [],
        'mem_latency_bound': [],
        'mem_bandwidth_bound': [],
        'timestamps': [],
        'report_count': 0
    }

    # Parse each report
    for report_idx, report in enumerate(reports[1:], start=1):  # Skip first empty split
        if not report.strip():
            continue

        result['report_count'] += 1

        # Extract timestamp
        time_match = re.search(r'Time:(\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2})', report)
        if time_match:
            result['timestamps'].append(time_match.group(1))
        else:
            result['timestamps'].append(f'Report-{report_idx}')

        # Parse Cycles, Instructions, IPC
        cycles_match = re.search(r'Cycles\s+([\d,]+)', report)
        inst_match = re.search(r'Instructions\s+([\d,]+)', report)
        ipc_match = re.search(r'IPC\s+([\d.]+)', report)

        result['cycles'].append(float(cycles_match.group(1).replace(',', '')) if cycles_match else 0)
        result['instructions'].append(float(inst_match.group(1).replace(',', '')) if inst_match else 0)
        result['ipc'].append(float(ipc_match.group(1)) if ipc_match else 0)

        # Parse top-down metrics
        bad_spec_match = re.search(r'Bad Speculation\s+([\d.]+)', report)
        frontend_match = re.search(r'Frontend Bound\s+([\d.]+)', report)
        retiring_match = re.search(r'Retiring\s+([\d.]+)', report)
        backend_match = re.search(r'Backend Bound\s+([\d.]+)', report)

        result['bad_speculation'].append(float(bad_spec_match.group(1)) if bad_spec_match else 0)
        result['frontend_bound'].append(float(frontend_match.group(1)) if frontend_match else 0)
        result['retiring'].append(float(retiring_match.group(1)) if retiring_match else 0)
        result['backend_bound'].append(float(backend_match.group(1)) if backend_match else 0)

        # Parse memory bound metrics
        l3_bound_match = re.search(r'L3 Bound\s+([\d.]+)', report)
        mem_bound_match = re.search(r'Mem Bound\s+([\d.]+)', report)
        lat_bound_match = re.search(r'Latency bound\s+([\d.]+)', report)
        bw_bound_match = re.search(r'Bandwidth bound\s+([\d.]+)', report)

        result['l3_bound'].append(float(l3_bound_match.group(1)) if l3_bound_match else 0)
        result['mem_bound'].append(float(mem_bound_match.group(1)) if mem_bound_match else 0)
        result['mem_latency_bound'].append(float(lat_bound_match.group(1)) if lat_bound_match else 0)
        result['mem_bandwidth_bound'].append(float(bw_bound_match.group(1)) if bw_bound_match else 0)

    if result['report_count'] == 0:
        return {'error': 'No valid reports found', 'report_count': 0}

    # Calculate averages with CORRECT IPC formula
    avg_result = {'report_count': result['report_count']}

    # Standard averages for non-IPC metrics
    for key in ['cycles', 'instructions', 'bad_speculation', 'frontend_bound',
                'retiring', 'backend_bound', 'l3_bound', 'mem_bound',
                'mem_latency_bound', 'mem_bandwidth_bound']:
        if result[key]:
            avg_result[f'{key}_avg'] = sum(result[key]) / len(result[key])
            avg_result[f'{key}_max'] = max(result[key])
            avg_result[f'{key}_min'] = min(result[key])
        else:
            avg_result[f'{key}_avg'] = 0.0
            avg_result[f'{key}_max'] = 0.0
            avg_result[f'{key}_min'] = 0.0

    # CORRECT IPC calculation: sum(instructions) / sum(cycles)
    total_instructions = sum(result['instructions']) if result['instructions'] else 0
    total_cycles = sum(result['cycles']) if result['cycles'] else 0
    avg_result['ipc_avg'] = total_instructions / total_cycles if total_cycles > 0 else 0.0
    avg_result['ipc_max'] = max(result['ipc']) if result['ipc'] else 0.0
    avg_result['ipc_min'] = min(result['ipc']) if result['ipc'] else 0.0

    # IMPORTANT: Keep original IPC array for comparison
    avg_result['ipc'] = result['ipc']  # Store raw IPC values for OLD formula calculation

    # Add timeline data
    avg_result['timestamps'] = result['timestamps']
    avg_result['timeline'] = {
        'timestamp': result['timestamps'],
        'ipc': result['ipc'],
        'bad_speculation': result['bad_speculation'],
        'frontend_bound': result['frontend_bound'],
        'retiring': result['retiring'],
        'backend_bound': result['backend_bound'],
    }

    return avg_result


def find_all_devkit_logs(result_base_dir: str) -> List[Dict]:
    """Find all devkit_top_down.log files in batch result directory

    Returns:
        List of dicts: {task_id, log_path, qemu_monitor_dir, task_dir}
    """
    log_files = []

    # Walk through all subdirectories
    for root, dirs, files in os.walk(result_base_dir):
        # Look for devkit_top_down.log in qemu_monitor/ subdirectory
        if 'devkit_top_down.log' in files:
            log_path = os.path.join(root, 'devkit_top_down.log')

            # Determine task_id from directory structure
            # Typically: batch_results_XXX/vmX_ratioY_activeZ/qemu_monitor/
            rel_path = os.path.relpath(root, result_base_dir)
            parts = rel_path.split(os.sep)

            # Extract task_id from directory name
            task_dir = None
            qemu_monitor_dir = root

            # Find task directory (parent of qemu_monitor)
            if 'qemu_monitor' in parts:
                idx = parts.index('qemu_monitor')
                if idx > 0:
                    task_dir = os.path.join(result_base_dir, *parts[:idx])
                    task_id = parts[idx - 1]  # e.g., "vm4_ratio0.8_active0.75"
                else:
                    task_id = "unknown"
            else:
                task_id = parts[-1] if parts else "unknown"
                task_dir = root

            log_files.append({
                'task_id': task_id,
                'log_path': log_path,
                'qemu_monitor_dir': qemu_monitor_dir,
                'task_dir': task_dir
            })

    return log_files


def update_analysis_report(qemu_monitor_dir: str, parsed_data: Dict) -> bool:
    """Update analysis_report.xlsx with corrected IPC value

    Args:
        qemu_monitor_dir: Directory containing analysis_report.xlsx
        parsed_data: Corrected parsed data from parse_devkit_top_down_offline

    Returns:
        True if update successful, False otherwise
    """
    excel_path = os.path.join(qemu_monitor_dir, 'analysis_report.xlsx')

    if not os.path.exists(excel_path):
        print(f"  Warning: analysis_report.xlsx not found in {qemu_monitor_dir}")
        return False

    if 'error' in parsed_data:
        print(f"  Error parsing data: {parsed_data['error']}")
        return False

    try:
        # Read all sheets from existing Excel file
        all_sheets = {}
        with pd.ExcelFile(excel_path) as xls:
            for sheet_name in xls.sheet_names:
                all_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)

        # Update DevKit_TopDown sheet with corrected IPC
        if 'DevKit_TopDown' in all_sheets:
            df_topdown = all_sheets['DevKit_TopDown']

            # Find and update IPC Avg row
            ipc_avg_idx = None
            ipc_max_idx = None
            ipc_min_idx = None

            for idx, row in df_topdown.iterrows():
                metric = str(row.get('Metric', '')).strip()
                if metric == 'IPC Avg':
                    ipc_avg_idx = idx
                elif metric == 'IPC Max':
                    ipc_max_idx = idx
                elif metric == 'IPC Min':
                    ipc_min_idx = idx

            # Update values
            if ipc_avg_idx is not None:
                old_ipc_avg = df_topdown.loc[ipc_avg_idx, 'Value']
                new_ipc_avg = parsed_data['ipc_avg']

                df_topdown.loc[ipc_avg_idx, 'Value'] = new_ipc_avg

                print(f"  IPC Avg: {old_ipc_avg:.4f} -> {new_ipc_avg:.4f} (corrected)")

            if ipc_max_idx is not None:
                df_topdown.loc[ipc_max_idx, 'Value'] = parsed_data['ipc_max']

            if ipc_min_idx is not None:
                df_topdown.loc[ipc_min_idx, 'Value'] = parsed_data['ipc_min']

            all_sheets['DevKit_TopDown'] = df_topdown

        # Also update TopDown_Timeline sheet if it exists
        if 'TopDown_Timeline' in all_sheets and 'timeline' in parsed_data:
            # Timeline IPC values are per-report, keep them unchanged
            # But we could add a note that avg is recalculated
            pass

        # Write back to Excel
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for sheet_name, df in all_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        return True

    except Exception as e:
        print(f"  Error updating Excel: {e}")
        return False


def extract_qemu_metrics_from_excel(result_dir: str) -> Dict:
    """Extract all QEMU metrics from analysis_report.xlsx (for batch summary)

    This is a copy of the function from batch_test_scheduler.py
    """
    qemu_dir = os.path.join(result_dir, "qemu_monitor")
    excel_path = os.path.join(qemu_dir, "analysis_report.xlsx")

    if not os.path.exists(excel_path):
        return {}

    metrics = {}

    try:
        # Summary sheet
        try:
            df_summary = pd.read_excel(excel_path, sheet_name="Summary")
            for idx, row in df_summary.iterrows():
                metric = str(row["Metric"]).strip() if pd.notna(row["Metric"]) else ""
                value = row["Value"]
                if metric == "VM Avg CPU":
                    metrics["avg_cpu_percent"] = float(value) if pd.notna(value) else 0
                elif metric == "VM Peak Total CPU":
                    metrics["max_cpu_percent"] = float(value) if pd.notna(value) else 0
        except Exception:
            pass

        # DevKit_TopDown sheet (13 metrics)
        try:
            df_topdown = pd.read_excel(excel_path, sheet_name="DevKit_TopDown")
            for idx, row in df_topdown.iterrows():
                metric = str(row["Metric"]).strip() if pd.notna(row["Metric"]) else ""
                value = row["Value"]
                key_map = {
                    "Cycles Avg": "td_cycles_avg",
                    "Instructions Avg": "td_instructions_avg",
                    "IPC Avg": "td_ipc_avg",
                    "IPC Max": "td_ipc_max",
                    "IPC Min": "td_ipc_min",
                    "Bad Speculation (%)": "td_bad_speculation",
                    "Frontend Bound (%)": "td_frontend_bound",
                    "Retiring (%)": "td_retiring",
                    "Backend Bound (%)": "td_backend_bound",
                    "L3 Bound (%)": "td_l3_bound",
                    "Mem Bound (%)": "td_mem_bound",
                    "Latency Bound (%)": "td_latency_bound",
                    "Bandwidth Bound (%)": "td_bandwidth_bound",
                }
                if metric in key_map:
                    metrics[key_map[metric]] = float(value) if pd.notna(value) else 0
        except Exception:
            pass

        # DevKit_Memory sheet
        try:
            df_mem = pd.read_excel(excel_path, sheet_name="DevKit_Memory")
            for idx, row in df_mem.iterrows():
                metric = str(row["Metric"]).strip() if pd.notna(row["Metric"]) else ""
                value = row["Value"]
                key_map = {
                    "L1D Miss (%)": "mem_l1d_miss",
                    "L1I Miss (%)": "mem_l1i_miss",
                    "L2D Miss (%)": "mem_l2d_miss",
                    "L2I Miss (%)": "mem_l2i_miss",
                    "DDR Read (MB/s)": "mem_ddr_read",
                    "DDR Write (MB/s)": "mem_ddr_write",
                }
                if metric in key_map:
                    metrics[key_map[metric]] = float(value) if pd.notna(value) else 0
        except Exception:
            pass

    except Exception as e:
        print(f"Warning: Failed to extract metrics from {excel_path}: {e}")

    return metrics


def regenerate_batch_summary(result_base_dir: str, output_path: str = None):
    """Regenerate batch_summary.xlsx with corrected IPC values

    Args:
        result_base_dir: Base directory containing all task result directories
        output_path: Output path for new batch_summary.xlsx (default: auto-generated)

    IMPORTANT: This function RE-PARSES devkit_top_down.log files to get correct IPC.
    It does NOT read from existing Excel files (which may have wrong IPC).
    """
    print("\n" + "=" * 70)
    print("Regenerating batch_summary.xlsx with corrected IPC values")
    print("IMPORTANT: Re-parsing log files, NOT reading from Excel")
    print("=" * 70)

    # Find all devkit_top_down.log files
    log_files = find_all_devkit_logs(result_base_dir)

    if not log_files:
        print("Error: No devkit_top_down.log files found")
        return

    print(f"Found {len(log_files)} devkit_top_down.log files")

    # Collect metrics from each task by RE-PARSING log files
    all_results = {'tasks': []}
    ipc_comparison_data = []

    for log_info in log_files:
        print(f"\nProcessing {log_info['task_id']}...")

        # Parse task_id to extract parameters
        # Format: vmX_ratioY_activeZ
        match = re.match(r'vm(\d+)_ratio([\d.]+)_active([\d.]+)', log_info['task_id'])
        if match:
            vm_count = int(match.group(1))
            ratio = float(match.group(2))
            active_percent = float(match.group(3))
        else:
            vm_count = 0
            ratio = 0.0
            active_percent = 0.0

        # === KEY FIX: Re-parse log file with correct formula ===
        log_path = log_info['log_path']

        if QEMU_MONITOR_AVAILABLE:
            parsed_data = parse_devkit_top_down(log_path)
        else:
            parsed_data = parse_devkit_top_down_offline(log_path)

        if 'error' in parsed_data:
            print(f"  Error parsing log: {parsed_data['error']}")
            qemu_metrics = {}
        else:
            # Build qemu_metrics dict from parsed data (CORRECT IPC!)
            qemu_metrics = {
                # Correct IPC values from log file
                'td_ipc_avg': parsed_data['ipc_avg'],  # CORRECTED!
                'td_ipc_max': parsed_data['ipc_max'],
                'td_ipc_min': parsed_data['ipc_min'],
                'td_cycles_avg': parsed_data['cycles_avg'],
                'td_instructions_avg': parsed_data['instructions_avg'],

                # Topdown metrics (averages)
                'td_bad_speculation': parsed_data['bad_speculation_avg'],
                'td_frontend_bound': parsed_data['frontend_bound_avg'],
                'td_retiring': parsed_data['retiring_avg'],
                'td_backend_bound': parsed_data['backend_bound_avg'],
                'td_l3_bound': parsed_data['l3_bound_avg'],
                'td_mem_bound': parsed_data['mem_bound_avg'],
                'td_latency_bound': parsed_data['mem_latency_bound_avg'],
                'td_bandwidth_bound': parsed_data['mem_bandwidth_bound_avg'],
            }

            # Extract other metrics from Excel (CPU, memory, etc.)
            other_metrics = extract_qemu_metrics_from_excel(log_info['task_dir'])
            qemu_metrics.update(other_metrics)

            # Calculate OLD IPC for comparison
            old_ipc_avg = sum(parsed_data['ipc']) / len(parsed_data['ipc']) if parsed_data['ipc'] else 0
            new_ipc_avg = parsed_data['ipc_avg']

            print(f"  OLD IPC avg (wrong): {old_ipc_avg:.4f}")
            print(f"  NEW IPC avg (correct): {new_ipc_avg:.4f}")
            print(f"  Difference: {abs(old_ipc_avg - new_ipc_avg):.4f}")
            print(f"  Reports: {parsed_data['report_count']}")

            ipc_comparison_data.append({
                'task_id': log_info['task_id'],
                'old_ipc': old_ipc_avg,
                'new_ipc': new_ipc_avg,
                'diff': abs(old_ipc_avg - new_ipc_avg),
                'reports': parsed_data['report_count']
            })

        # Build task result
        task_result = {
            'task_id': log_info['task_id'],
            'success': True,
            'parameters': {
                'vm_count': vm_count,
                'ratio': ratio,
                'active_percent': active_percent,
                'active_vm_count': int(vm_count * active_percent)
            },
            'browser_metrics': {},
            'qemu_metrics': qemu_metrics
        }

        all_results['tasks'].append(task_result)

    # Generate output path
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(result_base_dir, f"batch_summary_corrected_{timestamp}.xlsx")

    # Generate summary Excel
    generate_summary_excel(all_results, output_path)

    print(f"\nSuccess! Corrected batch_summary saved to:")
    print(f"  {output_path}")

    # === Print IPC Comparison Table ===
    if ipc_comparison_data:
        print("\n" + "=" * 80)
        print("IPC Calculation Comparison Table (OLD vs NEW formula)")
        print("=" * 80)
        print(f"{'Task ID':<30} {'OLD IPC':<12} {'NEW IPC':<12} {'Diff':<10} {'Reports':<8}")
        print("-" * 80)

        for item in ipc_comparison_data:
            print(f"{item['task_id']:<30} {item['old_ipc']:<12.4f} {item['new_ipc']:<12.4f} "
                  f"{item['diff']:<10.4f} {item['reports']:<8}")

        # Summary statistics
        avg_diff = sum(item['diff'] for item in ipc_comparison_data) / len(ipc_comparison_data)
        max_diff = max(item['diff'] for item in ipc_comparison_data)

        print("-" * 80)
        print(f"Average difference: {avg_diff:.4f}")
        print(f"Maximum difference: {max_diff:.4f}")
        print("=" * 80)

        # Warning if differences are small
        if max_diff < 0.01:
            print("\n⚠ Note: IPC differences are small (<0.01)")
            print("  This means each report's IPC values are similar.")
            print("  The formula difference is minimal for uniform data.")
        else:
            print("\n✓ IPC corrections are significant (>0.01)")
            print("  The new formula gives different results for your data.")

    print("\nKey: OLD formula = avg(IPC values), NEW formula = sum(instructions)/sum(cycles)")


def generate_summary_excel(results: Dict, output_path: str):
    """Generate batch_summary.xlsx with all metrics (copy from batch_test_scheduler.py)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Build DataFrame
        rows = []
        for task_data in results["tasks"]:
            params = task_data.get("parameters", {})
            qemu = task_data.get("qemu_metrics", {})

            row = {
                "test_id": task_data.get("task_id", ""),
                "vm_count": params.get("vm_count", 0),
                "ratio": params.get("ratio", 0),
                "active_percent": params.get("active_percent", 0),
                "active_vm_count": params.get("active_vm_count", 0),
                "success": task_data.get("success", False),

                # DevKit TopDown metrics (13)
                "td_cycles_avg": qemu.get("td_cycles_avg", 0),
                "td_instructions_avg": qemu.get("td_instructions_avg", 0),
                "td_ipc_avg": qemu.get("td_ipc_avg", 0),  # CORRECTED IPC!
                "td_ipc_max": qemu.get("td_ipc_max", 0),
                "td_ipc_min": qemu.get("td_ipc_min", 0),
                "td_bad_speculation": qemu.get("td_bad_speculation", 0),
                "td_frontend_bound": qemu.get("td_frontend_bound", 0),
                "td_retiring": qemu.get("td_retiring", 0),
                "td_backend_bound": qemu.get("td_backend_bound", 0),
                "td_l3_bound": qemu.get("td_l3_bound", 0),
                "td_mem_bound": qemu.get("td_mem_bound", 0),
                "td_latency_bound": qemu.get("td_latency_bound", 0),
                "td_bandwidth_bound": qemu.get("td_bandwidth_bound", 0),

                # DevKit Memory metrics
                "mem_l1d_miss": qemu.get("mem_l1d_miss", 0),
                "mem_l1i_miss": qemu.get("mem_l1i_miss", 0),
                "mem_l2d_miss": qemu.get("mem_l2d_miss", 0),
                "mem_l2i_miss": qemu.get("mem_l2i_miss", 0),
                "mem_ddr_read": qemu.get("mem_ddr_read", 0),
                "mem_ddr_write": qemu.get("mem_ddr_write", 0),

                # Summary metrics
                "avg_cpu_percent": qemu.get("avg_cpu_percent", 0),
                "max_cpu_percent": qemu.get("max_cpu_percent", 0),
            }
            rows.append(row)

        df = pd.DataFrame(rows)

        # Write to Excel with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Batch Summary"

        # Add header row
        headers = list(df.columns)
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for row_idx, row_data in enumerate(df.values, 2):
            ws.append(list(row_data))

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_idx in range(2, len(rows) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

        # Highlight td_ipc_avg column (corrected!)
        ipc_col_idx = headers.index("td_ipc_avg") + 1
        ipc_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        ws.cell(row=1, column=ipc_col_idx).fill = ipc_fill

        wb.save(output_path)

    except Exception as e:
        print(f"Error generating Excel: {e}")
        # Fallback: save as CSV
        csv_path = output_path.replace('.xlsx', '.csv')
        pd.DataFrame(rows).to_csv(csv_path, index=False)
        print(f"Saved as CSV instead: {csv_path}")


def main():
    parser = argparse.ArgumentParser(
        description='Offline IPC Fix Tool - Re-calculate td_ipc_avg with correct formula',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Fix IPC in a specific batch result directory
    python fix_ipc_offline.py --result-dir ./batch_results_20260615

    # Fix IPC and generate new batch_summary with custom name
    python fix_ipc_offline.py --result-dir ./batch_results --output ./corrected_summary.xlsx

IPC Formula:
    OLD (wrong): avg(IPC_1, IPC_2, ..., IPC_n)
    NEW (correct): sum(instructions) / sum(cycles)

This tool ALWAYS re-parses devkit_top_down.log files to get correct IPC.
        """
    )

    parser.add_argument('--result-dir', required=True,
                        help='Base directory containing batch test results (e.g., batch_results_20260615)')
    parser.add_argument('--output',
                        help='Output path for corrected batch_summary.xlsx (default: auto-generated)')
    parser.add_argument('--step-by-step', action='store_true',
                        help='Also update each analysis_report.xlsx (optional, for reference)')

    args = parser.parse_args()

    result_base_dir = args.result_dir

    if not os.path.exists(result_base_dir):
        print(f"Error: Result directory not found: {result_base_dir}")
        sys.exit(1)

    print("=" * 70)
    print("Offline IPC Fix Tool")
    print("=" * 70)
    print(f"Result directory: {result_base_dir}")
    print(f"Correct IPC formula: sum(instructions) / sum(cycles)")
    print("=" * 70)

    # Optional: Update individual analysis_report.xlsx files
    if args.step_by_step:
        print("\n[Step-by-step mode] Updating analysis_report.xlsx files...")
        log_files = find_all_devkit_logs(result_base_dir)

        if not log_files:
            print("Error: No devkit_top_down.log files found")
            sys.exit(1)

        updated_count = 0
        for log_info in log_files:
            print(f"\n[{log_info['task_id']}]")

            if QEMU_MONITOR_AVAILABLE:
                parsed_data = parse_devkit_top_down(log_info['log_path'])
            else:
                parsed_data = parse_devkit_top_down_offline(log_info['log_path'])

            if 'error' not in parsed_data:
                if update_analysis_report(log_info['qemu_monitor_dir'], parsed_data):
                    updated_count += 1

        print(f"\nUpdated {updated_count} analysis_report.xlsx files")

    # MAIN: Regenerate batch_summary with correct IPC (always from log files)
    regenerate_batch_summary(result_base_dir, args.output)


if __name__ == '__main__':
    main()