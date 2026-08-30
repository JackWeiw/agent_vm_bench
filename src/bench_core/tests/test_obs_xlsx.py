"""Phase 3: XlsxReportRenderer produces the expected sheets/columns."""
from __future__ import annotations

import pytest

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl is a core dep")
from openpyxl import load_workbook  # noqa: E402

from bench_core.config import KernelConfig  # noqa: E402
from bench_core.observability import ReplayObservability  # noqa: E402
from bench_core.obs_xlsx import XlsxReportRenderer  # noqa: E402
from bench_core.schemas import BenchSandbox, ReplayMetrics  # noqa: E402
from env_provider import SandboxInstance  # noqa: E402


def _seeded_observability(*, replay_mode="trajectory", with_retry=True) -> tuple[ReplayObservability, KernelConfig]:
    state = BenchSandbox.from_instance(SandboxInstance(id="x", index=0), "replay")
    m = ReplayMetrics()
    for i in range(3):
        m.add(
            latency=0.1 + i * 0.05,
            success=True,
            action_type="shell",
            resume_sec=0.05,
            pause_sec=0.05,
            slice_total_sec=1.0,
            running_slot_held_sec=0.8,
            interaction_total_sec=1.0,
            create_sec=1.0 + i * 0.5,
            kill_sec=0.5 + i * 0.1,
        )
    if with_retry:
        m.record_retry_event("retry_queued", operation="resume", time_lost_sec=0.05)
        m.append_retries_per_slice(1)
    state.replay_metrics = m
    cfg = KernelConfig(
        workflow_type="replay",
        replay_mode=replay_mode,
        total_count=2,
        replay_running_concurrency=1,
        test_duration=300,
    )
    obs = ReplayObservability(
        cfg,
        {0: state},
        admission_snapshot={
            "running": 1,
            "total": 2,
            "qps": 100.0,
            "running_slots": {
                "maximum": 1,
                "active": 0,
                "peak_active": 1,
                "granted": 3,
                "average_queue_wait_sec": 0.01,
                "waiting": 0,
            },
            "qps_limiter": {
                "qps": 100.0,
                "inflight_cap": 4,
                "in_flight": 0,
                "dispatched": 9,
                "average_wait_sec": 0.001,
                "max_wait_sec": 0.01,
                "dispatched_by_operation": {
                    "resume": 3,
                    "pause": 3,
                    "cleanup": 0,
                    "create": 3,
                    "command": 0,
                },
                "waiting": 0,
                "waiting_by_operation": {
                    "resume": 0,
                    "pause": 0,
                    "cleanup": 0,
                    "create": 0,
                    "command": 0,
                },
            },
        },
        wall_sec=10.0,
    )
    return obs, cfg


class TestXlsxReportRenderer:
    def test_produces_all_expected_sheets(self, tmp_path):
        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        wb = load_workbook(str(path))
        names = wb.sheetnames
        for sheet in (
            "Overview",
            "Per-step timings",
            "Lifecycle overhead",
            "Admission & QPS",
            "Throughput & overcommit",
            "Trajectory summary",
            "Retry impact",
        ):
            assert sheet in names, f"missing sheet {sheet}; got {names}"

    def test_admission_sheet_has_per_op_columns(self, tmp_path):
        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        wb = load_workbook(str(path))
        ws = wb["Admission & QPS"]
        headers = [c.value for c in ws[1]]
        assert "operation" in headers
        assert "dispatched" in headers

    def test_throughput_sheet_has_metrics(self, tmp_path):
        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        wb = load_workbook(str(path))
        ws = wb["Throughput & overcommit"]
        rows = [[c.value for c in r] for r in ws.iter_rows()]
        # has steps_per_sec / overcommit_ratio rows
        joined = " ".join(str(v) for row in rows for v in row)
        assert "steps_per_sec" in joined
        assert "overcommit_ratio" in joined

    def test_retry_impact_sheet_present_when_retries_exist(self, tmp_path):
        obs, _ = _seeded_observability(with_retry=True)
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        wb = load_workbook(str(path))
        ws = wb["Retry impact"]
        rows = [[c.value for c in r] for r in ws.iter_rows()]
        joined = " ".join(str(v) for row in rows for v in row)
        assert "retry_count" in joined

    def test_trajectory_sheet_omitted_content_when_not_trajectory(self, tmp_path):
        # lifecycle mode: Trajectory summary sheet exists but is empty/flagged
        obs, _ = _seeded_observability(replay_mode="lifecycle")
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        wb = load_workbook(str(path))
        assert "Trajectory summary" in wb.sheetnames
        ws = wb["Trajectory summary"]
        # no create_sec data in lifecycle mode -> header only
        assert ws.max_row == 1


def test_renderer_accepts_series_path_and_draws_per_step_linechart(tmp_path):
    from unittest.mock import MagicMock

    obs = MagicMock()
    obs.config.workflow_type = "replay"
    obs.config.replay_mode = "lifecycle"
    obs.config.total_count = 1
    obs.config.replay_running_concurrency = 1
    obs.config.test_duration = 10
    obs.wall_sec = 10.0
    obs.total_steps = 1
    obs.overcommit_ratio = 1.0
    obs.concurrency = 1
    obs.steps_per_sec = 0.1
    obs.effective_parallelism = 1.0
    obs.exec_wall_utilization = 1.0
    obs.retry_count = 0
    obs.retry_count_by_op = {}
    obs.time_lost_to_retry_sec = 0.0
    obs.retries_per_slice_p95 = 0.0
    obs.create_sec_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.kill_sec_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.slot_held_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.admission_snapshot = None
    # one sandbox with one step's durations
    m = MagicMock()
    m.latencies = [0.5]
    m.action_type_latencies = {}
    m.resume_secs = [0.1]
    m.pause_secs = [0.2]
    m.slice_total_secs = [0.8]
    m.running_slot_held_secs = [0.8]
    m.interaction_total_secs = [0.8]
    m.create_secs = []
    m.kill_secs = []
    m.success_count = 1
    m.failed_count = 0
    state = MagicMock()
    state.replay_metrics = m
    obs.states = {0: state}

    out = tmp_path / "obs.xlsx"
    XlsxReportRenderer(obs, series_path=None).render(out)
    wb = openpyxl.load_workbook(out)
    assert "Per-step timings" in wb.sheetnames
    # a LineChart exists on the sheet
    assert len(wb["Per-step timings"]._charts) >= 1


def test_per_step_linechart_references_all_data_rows(tmp_path):
    from unittest.mock import MagicMock
    from openpyxl import load_workbook

    obs = MagicMock()
    obs.config.workflow_type = "replay"
    obs.config.replay_mode = "lifecycle"
    obs.config.total_count = 1
    obs.config.replay_running_concurrency = 1
    obs.config.test_duration = 10
    obs.wall_sec = 10.0
    obs.total_steps = 3
    obs.overcommit_ratio = 1.0
    obs.concurrency = 1
    obs.steps_per_sec = 0.3
    obs.effective_parallelism = 1.0
    obs.exec_wall_utilization = 1.0
    obs.retry_count = 0
    obs.retry_count_by_op = {}
    obs.time_lost_to_retry_sec = 0.0
    obs.retries_per_slice_p95 = 0.0
    obs.create_sec_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.kill_sec_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.slot_held_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.admission_snapshot = None
    m = MagicMock()
    m.latencies = [0.1, 0.2, 0.3]  # THREE steps
    m.action_type_latencies = {}
    m.resume_secs = [0.1, 0.2, 0.3]
    m.pause_secs = [0.1, 0.2, 0.3]
    m.slice_total_secs = [0.5, 0.6, 0.7]
    m.running_slot_held_secs = [0.4, 0.5, 0.6]
    m.interaction_total_secs = [0.5, 0.6, 0.7]
    m.create_secs = []
    m.kill_secs = []
    m.success_count = 3
    m.failed_count = 0
    state = MagicMock()
    state.replay_metrics = m
    obs.states = {0: state}

    out = tmp_path / "obs.xlsx"
    XlsxReportRenderer(obs, series_path=None).render(out)
    wb = load_workbook(out)
    ws = wb["Per-step timings"]
    # the per-step detail block has 3 data rows; the latency series must
    # reference exactly 3 value cells (not 2 -- the off-by-one bug drops one).
    charts = ws._charts
    assert charts, "expected a LineChart on Per-step timings"
    # find the chart whose series references the latency_ms column (B)
    found = False
    for ch in charts:
        for s in ch.series:
            ref = getattr(s.val, "numRef", None)
            f = ref.f if ref is not None else None
            if f and "$B$" in f:
                # count rows in the range, e.g. '...!$B$6:$B$8' -> 3 rows
                import re

                m_ = re.search(r"\$B\$(\d+):\$B\$(\d+)", f)
                assert m_, f"unexpected val ref format: {f}"
                lo, hi = int(m_.group(1)), int(m_.group(2))
                assert hi - lo + 1 == 3, f"latency series references {hi-lo+1} rows, expected 3 (off-by-one? ref={f})"
                found = True
    assert found, "no LineChart series referenced the latency_ms (B) column"


def test_concurrency_states_sheet_from_series(tmp_path):
    from unittest.mock import MagicMock

    from bench_core.lifecycle_series import LifecycleSeriesWriter

    # write two step events spanning ~2 seconds
    sp = tmp_path / "s.jsonl"
    w = LifecycleSeriesWriter(sp)
    w.write(
        {
            "event": "step",
            "sandbox_index": 0,
            "step_index": 0,
            "resume_start": 0.0,
            "resume_end": 0.5,
            "exec_start": 0.5,
            "exec_end": 1.5,
            "pause_start": 1.5,
            "pause_end": 2.0,
        }
    )
    w.close()

    obs = MagicMock()  # minimal; the concurrency sheet reads only series_path
    r = XlsxReportRenderer(obs, series_path=sp)
    out = tmp_path / "o.xlsx"
    # stub the other sheets by calling just the new method
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    r._sheet_concurrency_states(wb)
    wb.save(out)
    wb2 = openpyxl.load_workbook(out)
    ws = wb2["Concurrency states"]
    # header + >=1 data row, one chart
    assert ws.max_row >= 2
    assert len(ws._charts) >= 1
