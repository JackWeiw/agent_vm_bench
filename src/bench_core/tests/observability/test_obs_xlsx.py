"""Phase 3: XlsxReportRenderer produces the expected sheets/columns."""
from __future__ import annotations

import pytest

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl is a core dep")
from openpyxl import load_workbook  # noqa: E402

from bench_core.config import KernelConfig  # noqa: E402
from bench_core.observability.replay_obs import ReplayObservability  # noqa: E402
from bench_core.observability.obs_xlsx import XlsxReportRenderer  # noqa: E402
from bench_core.schemas import BenchSandbox, ReplayMetrics  # noqa: E402
from env_provider import SandboxInstance  # noqa: E402


def _seeded_observability(*, replay_mode="trajectory", with_retry=True) -> tuple[ReplayObservability, KernelConfig]:
    state = BenchSandbox.from_instance(SandboxInstance(id="x", index=0), "replay")
    m = ReplayMetrics()
    for i in range(3):
        m.add(
            latency=0.1 + i * 0.05,
            success=True,
            action_type="shell",
            resume_sec=0.05,
            pause_sec=0.05,
            slice_total_sec=1.0,
            running_slot_held_sec=0.8,
            interaction_total_sec=1.0,
            create_sec=1.0 + i * 0.5,
            kill_sec=0.5 + i * 0.1,
        )
    if with_retry:
        m.record_retry_event("retry_queued", operation="resume", time_lost_sec=0.05)
        m.append_retries_per_slice(1)
    state.replay_metrics = m
    cfg = KernelConfig(
        workflow_type="replay",
        replay_mode=replay_mode,
        total_count=2,
        replay_running_concurrency=1,
        test_duration=300,
    )
    obs = ReplayObservability(
        cfg,
        {0: state},
        admission_snapshot={
            "running": 1,
            "total": 2,
            "qps": 100.0,
            "running_slots": {
                "maximum": 1,
                "active": 0,
                "peak_active": 1,
                "granted": 3,
                "average_queue_wait_sec": 0.01,
                "waiting": 0,
            },
            "qps_limiter": {
                "qps": 100.0,
                "inflight_cap": 4,
                "in_flight": 0,
                "dispatched": 9,
                "average_wait_sec": 0.001,
                "max_wait_sec": 0.01,
                "dispatched_by_operation": {
                    "resume": 3,
                    "pause": 3,
                    "cleanup": 0,
                    "create": 3,
                    "command": 0,
                },
                "waiting": 0,
                "waiting_by_operation": {
                    "resume": 0,
                    "pause": 0,
                    "cleanup": 0,
                    "create": 0,
                    "command": 0,
                },
            },
        },
        wall_sec=10.0,
    )
    return obs, cfg


class TestXlsxReportRenderer:
    def test_produces_all_expected_sheets(self, tmp_path):
        # Admission & QPS / Throughput & overcommit / Retry impact were folded
        # into Overview; the workbook now carries 8 base sheets and no standalone
        # scalar tabs. (Host-merged sheets are conditional, asserted separately.)
        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        names = load_workbook(str(path)).sheetnames
        for sheet in (
            "Overview",
            "Per-step timings",
            "Lifecycle overhead",
            "Trajectory summary",
            "Step detail",
            "Concurrency states",
            "Gantt",
            "Snapshot sizes",
        ):
            assert sheet in names, f"missing sheet {sheet}; got {names}"
        for removed in ("Admission & QPS", "Throughput & overcommit", "Retry impact"):
            assert removed not in names, f"{removed} was folded into Overview; got {names}"

    def test_overview_consolidates_admission_throughput_retry(self, tmp_path):
        # The former Admission & QPS / Throughput & overcommit / Retry impact
        # sheets' content now lives in Overview as grouped sections, so the
        # whole run reads on one sheet instead of four tiny scalar tabs.
        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        wb = load_workbook(str(path))
        ws = wb["Overview"]
        joined = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        # Throughput & overcommit scalars
        assert "steps_per_sec" in joined
        assert "effective_parallelism" in joined
        # Admission & QPS scalars + the per-operation dispatched/waiting table
        assert "qps" in joined
        assert "operation" in joined and "dispatched" in joined
        # Retry impact scalars
        assert "retry_count" in joined
        assert "time_lost_to_retry_sec" in joined

    def test_overview_metric_name_column_is_colored(self, tmp_path):
        # The data-name column (col A) carries a solid fill so metric labels
        # stand out from their values; section banners delimit the groups.
        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        wb = load_workbook(str(path))
        ws = wb["Overview"]
        name_cell = None
        banner_cell = None
        for row in ws.iter_rows():
            if row[0].value == "workflow_type":
                name_cell = row[0]
            if row[0].value == "Run":
                banner_cell = row[0]
        assert name_cell is not None, "workflow_type metric row not found in Overview"
        assert name_cell.fill.fill_type == "solid", "metric-name column must be color-filled"
        assert name_cell.font.bold, "metric-name column must be bold"
        assert banner_cell is not None, "Run section banner not found"
        assert banner_cell.fill.fill_type == "solid", "section banner must be color-filled"

    def test_overview_admission_has_per_op_table(self, tmp_path):
        # The Admission per-operation dispatched/waiting table renders as a
        # sub-table inside Overview (operation/dispatched/waiting header).
        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        ws = load_workbook(str(path))["Overview"]
        header_rows = [[c.value for c in r] for r in ws.iter_rows()]
        assert ["operation", "dispatched", "waiting"] in header_rows

    def test_overview_throughput_carries_values(self, tmp_path):
        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        ws = load_workbook(str(path))["Overview"]
        kv = {
            r[0].value: r[1].value
            for r in ws.iter_rows(min_col=1, max_col=2)
            if r[0].value is not None and r[1].value is not None
        }
        assert kv.get("steps_per_sec") == obs.steps_per_sec
        assert kv.get("concurrency") == obs.concurrency

    def test_overview_retry_carries_values(self, tmp_path):
        obs, _ = _seeded_observability(with_retry=True)
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        ws = load_workbook(str(path))["Overview"]
        kv = {
            r[0].value: r[1].value
            for r in ws.iter_rows(min_col=1, max_col=2)
            if r[0].value is not None and r[1].value is not None
        }
        assert kv.get("retry_count") == obs.retry_count
        assert kv.get("time_lost_to_retry_sec") == obs.time_lost_to_retry_sec
        # per-operation retry_queued:<op> rows are present
        assert any(k.startswith("retry_queued:") for k in kv), kv

    def test_trajectory_sheet_omitted_content_when_not_trajectory(self, tmp_path):
        # lifecycle mode: Trajectory summary sheet exists but is empty/flagged
        obs, _ = _seeded_observability(replay_mode="lifecycle")
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs).render(str(path))
        wb = load_workbook(str(path))
        assert "Trajectory summary" in wb.sheetnames
        ws = wb["Trajectory summary"]
        # no create_sec data in lifecycle mode -> header only
        assert ws.max_row == 1

    def test_host_sheets_ingested_in_single_render(self, tmp_path):
        from openpyxl import Workbook as _Wb

        host = tmp_path / "analysis_report.xlsx"
        src = _Wb()
        ws = src.active
        ws.title = "VM_Stats"
        ws.cell(row=1, column=1, value="vm")
        ws.cell(row=2, column=1, value="fc-1")
        n2 = src.create_sheet("NUMA_Overview")
        n2.cell(row=1, column=1, value="node0")
        src.save(host)

        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs, host_xlsx=host).render(str(path))
        wb = load_workbook(str(path))
        assert "VM_Stats" in wb.sheetnames
        assert wb["VM_Stats"]["A2"].value == "fc-1"
        assert "NUMA_Overview" in wb.sheetnames

    def test_host_merge_skipped_without_host_xlsx(self, tmp_path):
        obs, _ = _seeded_observability()
        path = tmp_path / "obs.xlsx"
        XlsxReportRenderer(obs, host_xlsx=None).render(str(path))
        wb = load_workbook(str(path))
        assert "VM_Stats" not in wb.sheetnames


def test_renderer_accepts_series_path_and_draws_per_step_linechart(tmp_path):
    from unittest.mock import MagicMock

    obs = MagicMock()
    obs.config.workflow_type = "replay"
    obs.config.replay_mode = "lifecycle"
    obs.config.total_count = 1
    obs.config.replay_running_concurrency = 1
    obs.config.test_duration = 10
    obs.wall_sec = 10.0
    obs.total_steps = 1
    obs.overcommit_ratio = 1.0
    obs.concurrency = 1
    obs.steps_per_sec = 0.1
    obs.effective_parallelism = 1.0
    obs.exec_wall_utilization = 1.0
    obs.retry_count = 0
    obs.retry_count_by_op = {}
    obs.time_lost_to_retry_sec = 0.0
    obs.retries_per_slice_p95 = 0.0
    obs.create_sec_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.kill_sec_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.slot_held_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.admission_snapshot = None
    # one sandbox with one step's durations
    m = MagicMock()
    m.latencies = [0.5]
    m.action_type_latencies = {}
    m.resume_secs = [0.1]
    m.pause_secs = [0.2]
    m.slice_total_secs = [0.8]
    m.running_slot_held_secs = [0.8]
    m.interaction_total_secs = [0.8]
    m.create_secs = []
    m.kill_secs = []
    m.success_count = 1
    m.failed_count = 0
    state = MagicMock()
    state.replay_metrics = m
    obs.states = {0: state}

    out = tmp_path / "obs.xlsx"
    XlsxReportRenderer(obs, series_path=None).render(out)
    wb = openpyxl.load_workbook(out)
    assert "Per-step timings" in wb.sheetnames
    # a LineChart exists on the sheet
    assert len(wb["Per-step timings"]._charts) >= 1


def test_per_step_linechart_references_all_data_rows(tmp_path):
    from unittest.mock import MagicMock
    from openpyxl import load_workbook

    obs = MagicMock()
    obs.config.workflow_type = "replay"
    obs.config.replay_mode = "lifecycle"
    obs.config.total_count = 1
    obs.config.replay_running_concurrency = 1
    obs.config.test_duration = 10
    obs.wall_sec = 10.0
    obs.total_steps = 3
    obs.overcommit_ratio = 1.0
    obs.concurrency = 1
    obs.steps_per_sec = 0.3
    obs.effective_parallelism = 1.0
    obs.exec_wall_utilization = 1.0
    obs.retry_count = 0
    obs.retry_count_by_op = {}
    obs.time_lost_to_retry_sec = 0.0
    obs.retries_per_slice_p95 = 0.0
    obs.create_sec_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.kill_sec_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.slot_held_stats = {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0}
    obs.admission_snapshot = None
    m = MagicMock()
    m.latencies = [0.1, 0.2, 0.3]  # THREE steps
    m.action_type_latencies = {}
    m.resume_secs = [0.1, 0.2, 0.3]
    m.pause_secs = [0.1, 0.2, 0.3]
    m.slice_total_secs = [0.5, 0.6, 0.7]
    m.running_slot_held_secs = [0.4, 0.5, 0.6]
    m.interaction_total_secs = [0.5, 0.6, 0.7]
    m.create_secs = []
    m.kill_secs = []
    m.success_count = 3
    m.failed_count = 0
    state = MagicMock()
    state.replay_metrics = m
    obs.states = {0: state}

    out = tmp_path / "obs.xlsx"
    XlsxReportRenderer(obs, series_path=None).render(out)
    wb = load_workbook(out)
    ws = wb["Per-step timings"]
    # the per-step detail block has 3 data rows; the latency series must
    # reference exactly 3 value cells (not 2 -- the off-by-one bug drops one).
    charts = ws._charts
    assert charts, "expected a LineChart on Per-step timings"
    # find the chart whose series references the latency_ms column (B)
    found = False
    for ch in charts:
        for s in ch.series:
            ref = getattr(s.val, "numRef", None)
            f = ref.f if ref is not None else None
            if f and "$B$" in f:
                # count rows in the range, e.g. '...!$B$6:$B$8' -> 3 rows
                import re

                m_ = re.search(r"\$B\$(\d+):\$B\$(\d+)", f)
                assert m_, f"unexpected val ref format: {f}"
                lo, hi = int(m_.group(1)), int(m_.group(2))
                assert hi - lo + 1 == 3, f"latency series references {hi-lo+1} rows, expected 3 (off-by-one? ref={f})"
                found = True
    assert found, "no LineChart series referenced the latency_ms (B) column"


def test_concurrency_states_sheet_from_series(tmp_path):
    from unittest.mock import MagicMock

    from bench_core.observability.lifecycle_series import LifecycleSeriesWriter

    # write two step events spanning ~2 seconds
    sp = tmp_path / "s.jsonl"
    w = LifecycleSeriesWriter(sp)
    w.write(
        {
            "event": "step",
            "sandbox_index": 0,
            "step_index": 0,
            "resume_start": 0.0,
            "resume_end": 0.5,
            "exec_start": 0.5,
            "exec_end": 1.5,
            "pause_start": 1.5,
            "pause_end": 2.0,
        }
    )
    w.close()

    obs = MagicMock()  # minimal; the concurrency sheet reads only series_path
    r = XlsxReportRenderer(obs, series_path=sp)
    out = tmp_path / "o.xlsx"
    # stub the other sheets by calling just the new method
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    r._sheet_concurrency_states(wb)
    wb.save(out)
    wb2 = openpyxl.load_workbook(out)
    ws = wb2["Concurrency states"]
    # header + >=1 data row, one chart
    assert ws.max_row >= 2
    assert len(ws._charts) >= 1


def test_gantt_sheet_embeds_png(tmp_path):
    pytest.importorskip("matplotlib")
    from unittest.mock import MagicMock

    from bench_core.observability.lifecycle_series import LifecycleSeriesWriter

    sp = tmp_path / "s.jsonl"
    w = LifecycleSeriesWriter(sp)
    w.write(
        {
            "event": "step",
            "sandbox_index": 0,
            "step_index": 0,
            "resume_start": 0.0,
            "resume_end": 0.5,
            "exec_start": 0.5,
            "exec_end": 1.5,
            "pause_start": 1.5,
            "pause_end": 2.0,
        }
    )
    w.close()

    r = XlsxReportRenderer(MagicMock(), series_path=sp)
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    r._sheet_gantt(wb, out_png=tmp_path / "gantt.png")
    wb.save(tmp_path / "o.xlsx")
    assert (tmp_path / "gantt.png").exists()


def test_snapshot_sizes_sheet(tmp_path):
    from unittest.mock import MagicMock

    from bench_core.observability.lifecycle_series import LifecycleSeriesWriter

    sp = tmp_path / "s.jsonl"
    w = LifecycleSeriesWriter(sp)
    w.write(
        {
            "event": "snapshot_size",
            "sandbox_index": 0,
            "sandbox_id": "a",
            "pause_seq": 1,
            "logical_bytes": 2 * 1024 * 1024,
            "disk_bytes": 1024 * 1024,
            "inherited_bytes": 0,
            "cumulative_bytes": 1024 * 1024,
            "generations": 1,
            "files": 1,
        }
    )
    w.close()

    r = XlsxReportRenderer(MagicMock(), series_path=sp)
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    r._sheet_snapshot_sizes(wb)
    wb.save(tmp_path / "o.xlsx")
    ws = openpyxl.load_workbook(tmp_path / "o.xlsx")["Snapshot sizes"]
    assert ws.max_row >= 2
    assert len(ws._charts) >= 1


def test_step_detail_sheet_breaks_down_per_trajectory(tmp_path):
    """Step detail: one row per step event, broken down by trajectory_id,
    sorted by (trajectory, sandbox, step), with a frozen header + autofilter."""
    from unittest.mock import MagicMock

    from bench_core.observability.lifecycle_series import LifecycleSeriesWriter

    sp = tmp_path / "s.jsonl"
    w = LifecycleSeriesWriter(sp)
    # Two trajectories, two sandboxes. Written out of order to exercise the sort.
    w.write(
        {
            "event": "step",
            "sandbox_index": 1,
            "trajectory_id": "traj-b",
            "round_id": 0,
            "step_index": 0,
            "action_type": "shell",
            "slice_failed": False,
            "resume_sec": 0.1,
            "resume_queue_wait_sec": 0.02,
            "resume_api_sec": 0.05,
            "resume_ready_wait_sec": 0.03,
            "exec_sec": 0.5,
            "pause_sec": 0.2,
            "pause_queue_wait_sec": 0.04,
            "pause_api_sec": 0.16,
            "slice_total_sec": 0.8,
            "interaction_total_sec": 0.8,
            "slot_contention_wait_sec": 0.0,
            "running_slot_held_sec": 0.7,
            "exit_code": 0,
            "timed_out": False,
        }
    )
    w.write(
        {
            "event": "step",
            "sandbox_index": 0,
            "trajectory_id": "traj-a",
            "round_id": 0,
            "step_index": 1,
            "action_type": "edit",
            "slice_failed": False,
            "resume_sec": 0.1,
            "resume_queue_wait_sec": 0.02,
            "resume_api_sec": 0.05,
            "resume_ready_wait_sec": 0.03,
            "exec_sec": 0.4,
            "pause_sec": 0.2,
            "pause_queue_wait_sec": 0.04,
            "pause_api_sec": 0.16,
            "slice_total_sec": 0.7,
            "interaction_total_sec": 0.7,
            "slot_contention_wait_sec": 0.0,
            "running_slot_held_sec": 0.6,
            "exit_code": 0,
            "timed_out": False,
        }
    )
    # a failed step (slice_failed=True) is still emitted as a row
    w.write(
        {
            "event": "step",
            "sandbox_index": 0,
            "trajectory_id": "traj-a",
            "round_id": 0,
            "step_index": 0,
            "action_type": "shell",
            "slice_failed": True,
            "resume_sec": 0.0,
            "resume_queue_wait_sec": 0.0,
            "resume_api_sec": 0.0,
            "resume_ready_wait_sec": 0.0,
            "exec_sec": 0.0,
            "pause_sec": 0.0,
            "pause_queue_wait_sec": 0.0,
            "pause_api_sec": 0.0,
            "slice_total_sec": 0.0,
            "interaction_total_sec": 0.0,
            "slot_contention_wait_sec": 0.0,
            "running_slot_held_sec": 0.0,
            "exit_code": 1,
            "timed_out": True,
        }
    )
    # a non-step event is ignored
    w.write({"event": "snapshot_size", "sandbox_index": 0})
    w.close()

    r = XlsxReportRenderer(MagicMock(), series_path=sp)
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    r._sheet_step_detail(wb)
    out = tmp_path / "o.xlsx"
    wb.save(out)
    ws = openpyxl.load_workbook(out)["Step detail"]

    headers = [c.value for c in ws[1]]
    assert "trajectory_id" in headers
    assert "exec_sec" in headers
    assert "slice_failed" in headers
    # sub-decomposition columns (already emitted by the runner into the series)
    # are surfaced so the sum invariants are inspectable in the sheet itself.
    assert "resume_queue_wait_sec" in headers
    assert "resume_api_sec" in headers
    assert "resume_ready_wait_sec" in headers
    assert "pause_queue_wait_sec" in headers
    assert "pause_api_sec" in headers
    assert "running_slot_held_sec" in headers
    # header + 3 data rows (the snapshot_size event is dropped)
    assert ws.max_row == 4
    # sorted: traj-a step0, traj-a step1, traj-b step0
    traj_col = headers.index("trajectory_id") + 1
    step_col = headers.index("step_index") + 1
    assert ws.cell(2, traj_col).value == "traj-a"
    assert ws.cell(2, step_col).value == 0
    assert ws.cell(3, traj_col).value == "traj-a"
    assert ws.cell(3, step_col).value == 1
    assert ws.cell(4, traj_col).value == "traj-b"
    # the failed step row carries slice_failed=True and timed_out=True
    failed_col = headers.index("slice_failed") + 1
    timed_col = headers.index("timed_out") + 1
    assert ws.cell(2, failed_col).value is True
    assert ws.cell(2, timed_col).value is True
    # sum invariant on the success row: resume_sec == queue + api + ready_wait
    # (rounded to 3 dp), pause_sec == queue + api. Guards against column-order
    # drift silently breaking the parent/child relationship.
    rsm = ws.cell(3, headers.index("resume_sec") + 1).value
    rq = ws.cell(3, headers.index("resume_queue_wait_sec") + 1).value
    ra = ws.cell(3, headers.index("resume_api_sec") + 1).value
    rr = ws.cell(3, headers.index("resume_ready_wait_sec") + 1).value
    assert rsm is not None and round(rq + ra + rr, 3) == rsm
    psm = ws.cell(3, headers.index("pause_sec") + 1).value
    pq = ws.cell(3, headers.index("pause_queue_wait_sec") + 1).value
    pa = ws.cell(3, headers.index("pause_api_sec") + 1).value
    assert psm is not None and round(pq + pa, 3) == psm
    # frozen header + autofilter on the data range
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_step_detail_sheet_empty_without_series(tmp_path):
    """No series file -> Step detail sheet exists with a header only (no crash)."""
    from unittest.mock import MagicMock

    from openpyxl import Workbook

    r = XlsxReportRenderer(MagicMock(), series_path=None)
    wb = Workbook()
    wb.remove(wb.active)
    r._sheet_step_detail(wb)
    wb.save(tmp_path / "o.xlsx")
    ws = openpyxl.load_workbook(tmp_path / "o.xlsx")["Step detail"]
    assert ws.max_row == 1  # header only
    assert ws.cell(1, 1).value == "trajectory_id"


def test_trajectory_summary_attributes_cost_per_instance(tmp_path):
    """Trajectory summary: per-instance SUMS (not percentiles) so each
    trajectory's wall-clock decomposes into pause/resume/exec/wait totals --
    "where did this trajectory's time go", which percentiles (available in the
    Step detail / Lifecycle overhead sheets) do not show."""
    from unittest.mock import MagicMock

    from bench_core.observability.lifecycle_series import LifecycleSeriesWriter

    sp = tmp_path / "s.jsonl"
    w = LifecycleSeriesWriter(sp)
    # Two trajectories, 3 steps each, distinct resume/pause/wait so sums are
    # attributable per instance. slice_total = resume + exec + pause exactly.
    for tid, resume, pause, slot_wait in (("traj-a", 0.10, 0.20, 0.01), ("traj-b", 0.50, 0.60, 0.05)):
        for i in range(3):
            w.write(
                {
                    "event": "step",
                    "sandbox_index": i,
                    "trajectory_id": tid,
                    "step_index": i,
                    "exec_sec": 0.4,
                    "resume_sec": resume,
                    "pause_sec": pause,
                    "slot_contention_wait_sec": slot_wait,
                    "resume_queue_wait_sec": 0.02,
                    "pause_queue_wait_sec": 0.03,
                    "running_slot_held_sec": 0.9,
                    "slice_total_sec": round(resume + 0.4 + pause, 3),
                    "interaction_total_sec": round(resume + 0.4 + pause + 0.05, 3),
                }
            )
    w.close()

    obs = MagicMock()
    obs.config.replay_mode = "lifecycle"
    r = XlsxReportRenderer(obs, series_path=sp)
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    r._sheet_trajectory_summary(wb)
    wb.save(tmp_path / "o.xlsx")
    ws = openpyxl.load_workbook(tmp_path / "o.xlsx")["Trajectory summary"]

    headers = [c.value for c in ws[1]]
    assert "trajectory_id" in headers
    assert "n_steps" in headers
    # sum-based cost-attribution columns (no percentiles on this sheet)
    assert "slice_total_sum_s" in headers
    assert "exec_sum_s" in headers
    assert "resume_sum_s" in headers
    assert "pause_sum_s" in headers
    assert "slot_wait_sum_s" in headers
    assert "resume_queue_wait_sum_s" in headers
    assert "pause_queue_wait_sum_s" in headers
    assert "running_slot_held_sum_s" in headers
    assert "avg_slice_s" in headers
    # no percentile columns remain (they live in Step detail / Lifecycle overhead)
    assert not any(h.endswith("_p50_s") for h in headers)
    # one row per trajectory (a, b), header + 2 rows
    assert ws.max_row == 3
    traj_col = headers.index("trajectory_id") + 1
    assert ws.cell(2, traj_col).value == "traj-a"
    assert ws.cell(3, traj_col).value == "traj-b"
    # traj-a: resume 3x0.10=0.30, pause 3x0.20=0.60, exec 3x0.4=1.2,
    # slice_total 3x0.70=2.10, slot_wait 3x0.01=0.03
    rs = headers.index("resume_sum_s") + 1
    ps = headers.index("pause_sum_s") + 1
    es = headers.index("exec_sum_s") + 1
    ss = headers.index("slice_total_sum_s") + 1
    sw = headers.index("slot_wait_sum_s") + 1
    assert ws.cell(2, rs).value == 0.3
    assert ws.cell(2, ps).value == 0.6
    assert ws.cell(2, es).value == 1.2
    assert ws.cell(2, ss).value == 2.1
    assert ws.cell(2, sw).value == 0.03
    # traj-b resume (1.50) > traj-a resume (0.30) -- per-instance separation
    assert ws.cell(3, rs).value == 1.5
    # sum invariant: slice_total_sum == resume_sum + exec_sum + pause_sum
    assert round(ws.cell(2, rs).value + ws.cell(2, es).value + ws.cell(2, ps).value, 3) == ws.cell(2, ss).value
    # avg_slice = slice_total_sum / n_steps (traj-a: 2.10 / 3 = 0.70)
    n_col = headers.index("n_steps") + 1
    avg_col = headers.index("avg_slice_s") + 1
    assert ws.cell(2, n_col).value == 3
    assert ws.cell(2, avg_col).value == 0.7


def test_trajectory_summary_has_stacked_cost_bar_chart(tmp_path):
    """The per-trajectory cost-attribution sums get a stacked bar chart so the
    "where did each trajectory's wall-clock go" table reads at a glance: exec /
    resume / pause stacked per trajectory (bar height == slice_total_sum)."""
    from unittest.mock import MagicMock

    from bench_core.observability.lifecycle_series import LifecycleSeriesWriter

    sp = tmp_path / "s.jsonl"
    w = LifecycleSeriesWriter(sp)
    for tid, resume, pause in (("traj-a", 0.10, 0.20), ("traj-b", 0.50, 0.60)):
        for i in range(3):
            w.write(
                {
                    "event": "step",
                    "sandbox_index": i,
                    "trajectory_id": tid,
                    "step_index": i,
                    "exec_sec": 0.4,
                    "resume_sec": resume,
                    "pause_sec": pause,
                    "slice_total_sec": round(resume + 0.4 + pause, 3),
                    "interaction_total_sec": round(resume + 0.4 + pause + 0.05, 3),
                }
            )
    w.close()

    obs = MagicMock()
    obs.config.replay_mode = "lifecycle"
    r = XlsxReportRenderer(obs, series_path=sp)
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    r._sheet_trajectory_summary(wb)
    out = tmp_path / "o.xlsx"
    wb.save(out)
    ws = openpyxl.load_workbook(out)["Trajectory summary"]
    charts = ws._charts
    assert charts, "expected a stacked bar chart on Trajectory summary"
    ch = charts[0]
    # three stacked cost-decomposition series: exec / resume / pause
    assert len(ch.series) == 3, f"expected 3 cost series; got {len(ch.series)}"


def test_trajectory_summary_chart_legend_uses_header_names(tmp_path):
    """The stacked cost bar chart's legend must take titles from the HEADER row
    (exec_sum_s / resume_sum_s / pause_sum_s), not from the first data row
    (which would surface the first trajectory's numeric sums as legend labels).

    Regression guard: header_row was derived as ws.max_row+1 on a fresh empty
    sheet, but openpyxl reports max_row==1 for an empty sheet, so the +1 aimed
    the chart's title Reference at row 2 (first data row) instead of row 1.
    """
    import re
    from unittest.mock import MagicMock

    from bench_core.observability.lifecycle_series import LifecycleSeriesWriter

    sp = tmp_path / "s.jsonl"
    w = LifecycleSeriesWriter(sp)
    for tid, resume, pause in (("traj-a", 0.10, 0.20), ("traj-b", 0.50, 0.60)):
        for i in range(3):
            w.write(
                {
                    "event": "step",
                    "sandbox_index": i,
                    "trajectory_id": tid,
                    "step_index": i,
                    "exec_sec": 0.4,
                    "resume_sec": resume,
                    "pause_sec": pause,
                    "slice_total_sec": round(resume + 0.4 + pause, 3),
                    "interaction_total_sec": round(resume + 0.4 + pause + 0.05, 3),
                }
            )
    w.close()

    obs = MagicMock()
    obs.config.replay_mode = "lifecycle"
    r = XlsxReportRenderer(obs, series_path=sp)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    r._sheet_trajectory_summary(wb)
    out = tmp_path / "o.xlsx"
    wb.save(out)
    ws = load_workbook(out)["Trajectory summary"]
    ch = ws._charts[0]

    # Each series title is a StrRef into a header cell (titles_from_data=True).
    # Resolve the referenced cell and read its value -- must be a column name,
    # not a numeric sum from the first trajectory's data row.
    labels = []
    for s in ch.series:
        f = s.tx.strRef.f if (s.tx is not None and s.tx.strRef is not None) else None
        m = re.search(r"!([A-Z]+)(\d+)$", f or "")
        assert m, f"unexpected series title reference: {f!r}"
        labels.append(ws[f"{m.group(1)}{m.group(2)}"].value)
    assert labels == [
        "exec_sum_s",
        "resume_sum_s",
        "pause_sum_s",
    ], f"chart legend should be header column names, got {labels}"


def _traj_series_file(sp, events):
    """Write a list of step-event dicts to a lifecycle series JSONL file."""
    from bench_core.observability.lifecycle_series import LifecycleSeriesWriter

    w = LifecycleSeriesWriter(sp)
    for ev in events:
        w.write(ev)
    w.close()


def _step_ev(
    tid,
    *,
    step_index=0,
    sandbox_index=0,
    exec_sec=0.4,
    resume_sec=0.1,
    pause_sec=0.2,
    slice_failed=False,
    timed_out=False,
    exit_code=0,
    slot_contention_wait_sec=0.0,
):
    s = round(resume_sec + exec_sec + pause_sec, 3)
    return {
        "event": "step",
        "sandbox_index": sandbox_index,
        "trajectory_id": tid,
        "step_index": step_index,
        "action_type": "shell",
        "resume_sec": resume_sec,
        "exec_sec": exec_sec,
        "pause_sec": pause_sec,
        "slice_total_sec": s,
        "interaction_total_sec": round(s + 0.05, 3),
        "slot_contention_wait_sec": slot_contention_wait_sec,
        "resume_queue_wait_sec": 0.0,
        "pause_queue_wait_sec": 0.0,
        "running_slot_held_sec": 0.0,
        "slice_failed": slice_failed,
        "timed_out": timed_out,
        "exit_code": exit_code,
    }


def test_trajectory_summary_has_failure_and_success_columns(tmp_path):
    """n_failed / n_timeout / success_rate columns let the user spot which
    trajectory fails or times out most -- the reference's Per-task summary has
    ok/failed/timeout; bench-core now matches + adds the rate."""
    from unittest.mock import MagicMock

    sp = tmp_path / "s.jsonl"
    _traj_series_file(
        sp,
        [
            # traj-a: 2 success + 1 failed slice
            _step_ev("traj-a", step_index=0),
            _step_ev("traj-a", step_index=1),
            _step_ev(
                "traj-a", step_index=2, slice_failed=True, exit_code=1, exec_sec=0.0, resume_sec=0.0, pause_sec=0.0
            ),
            # traj-b: 1 success + 1 timeout (exec completed but hit the timeout)
            _step_ev("traj-b", step_index=0),
            _step_ev("traj-b", step_index=1, timed_out=True, exit_code=124),
        ],
    )

    obs = MagicMock()
    obs.config.replay_mode = "lifecycle"
    r = XlsxReportRenderer(obs, series_path=sp)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    r._sheet_trajectory_summary(wb)
    out = tmp_path / "o.xlsx"
    wb.save(out)
    ws = load_workbook(out)["Trajectory summary"]

    headers = [c.value for c in ws[1]]
    assert "n_failed" in headers
    assert "n_timeout" in headers
    assert "success_rate" in headers

    traj_col = headers.index("trajectory_id") + 1
    nf = headers.index("n_failed") + 1
    nt = headers.index("n_timeout") + 1
    sr = headers.index("success_rate") + 1
    # traj-a row (row 2): 1 failed, 0 timeout, success_rate 2/3
    assert ws.cell(2, traj_col).value == "traj-a"
    assert ws.cell(2, nf).value == 1
    assert ws.cell(2, nt).value == 0
    assert round(ws.cell(2, sr).value, 6) == round(2 / 3, 6)
    # traj-b row (row 3): 0 failed, 1 timeout, success_rate 1.0 (timeout != slice_failed)
    assert ws.cell(3, traj_col).value == "traj-b"
    assert ws.cell(3, nf).value == 0
    assert ws.cell(3, nt).value == 1
    assert ws.cell(3, sr).value == 1.0


def test_trajectory_summary_has_data_bars_and_failure_color_scale(tmp_path):
    """Conditional formatting: data bars on slice_total_sum_s + slot_wait_sum_s
    (longer bar = slower / more queueing) and a red color scale on n_failed --
    the at-a-glance outlier highlighting the reference's per-trial table lacks."""
    from unittest.mock import MagicMock

    sp = tmp_path / "s.jsonl"
    _traj_series_file(
        sp,
        [
            _step_ev("traj-a", step_index=0, slot_contention_wait_sec=0.01),
            _step_ev("traj-a", step_index=1, slot_contention_wait_sec=0.01),
            _step_ev("traj-b", step_index=0, exec_sec=2.0, slot_contention_wait_sec=0.5),
            _step_ev("traj-b", step_index=1, exec_sec=2.0, slot_contention_wait_sec=0.5),
        ],
    )

    obs = MagicMock()
    obs.config.replay_mode = "lifecycle"
    r = XlsxReportRenderer(obs, series_path=sp)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    r._sheet_trajectory_summary(wb)
    out = tmp_path / "o.xlsx"
    wb.save(out)
    ws = load_workbook(out)["Trajectory summary"]

    headers = [c.value for c in ws[1]]
    from openpyxl.utils import get_column_letter

    slice_col = get_column_letter(headers.index("slice_total_sum_s") + 1)
    slot_col = get_column_letter(headers.index("slot_wait_sum_s") + 1)
    nf_col = get_column_letter(headers.index("n_failed") + 1)

    rule_cols = []  # (type, covered-column-letter)
    for cf in ws.conditional_formatting:
        for rule in cf.rules:
            for rng in str(cf.sqref).split():
                col = "".join(ch for ch in rng.split(":")[0] if ch.isalpha())
                rule_cols.append((rule.type, col))
    types = {t for t, _ in rule_cols}
    cols_by_type = {t: {c for _, c in rule_cols if _[0] == t} for t in types}
    # simpler: gather per-type column sets
    databar_cols = {c for t, c in rule_cols if t == "dataBar"}
    colorscale_cols = {c for t, c in rule_cols if t == "colorScale"}
    assert slice_col in databar_cols, f"expected dataBar on {slice_col}; got {databar_cols}"
    assert slot_col in databar_cols, f"expected dataBar on {slot_col}; got {databar_cols}"
    assert nf_col in colorscale_cols, f"expected colorScale on {nf_col}; got {colorscale_cols}"


def test_trajectory_summary_chart_anchor_clear_of_table(tmp_path):
    """Adding the failure cluster widened the table to 15 columns; the chart
    anchor must be COMPUTED from the header count (not the old hardcoded "N2")
    so the floating chart never sits on top of a data column. Regression guard."""
    import re
    from unittest.mock import MagicMock
    from openpyxl.utils import column_index_from_string

    sp = tmp_path / "s.jsonl"
    _traj_series_file(sp, [_step_ev("traj-a"), _step_ev("traj-b")])

    obs = MagicMock()
    obs.config.replay_mode = "lifecycle"
    r = XlsxReportRenderer(obs, series_path=sp)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    r._sheet_trajectory_summary(wb)
    ws = wb["Trajectory summary"]
    charts = ws._charts
    assert charts, "expected the cost chart"
    n_headers = ws.max_column  # 15
    # Read the anchor robustly (string in-memory; OneCellAnchor after reload).
    anchor = charts[0].anchor
    if isinstance(anchor, str):
        m = re.match(r"([A-Z]+)", anchor)
        anchor_col = column_index_from_string(m.group(1))
    else:
        marker = getattr(anchor, "from_", None) or getattr(anchor, "_from", None)
        anchor_col = marker.col + 1  # openpyxl markers are 0-based
    assert anchor_col > n_headers, (
        f"chart anchor col {anchor_col} must sit past the {n_headers}-col table; " f"anchor={anchor!r}"
    )


def test_trajectory_summary_cf_absent_without_series(tmp_path):
    """No series file -> no conditional-formatting rules (a DataBarRule on an
    empty/inverted range would corrupt the sheet). max_row stays 1."""
    obs, _ = _seeded_observability(replay_mode="lifecycle")
    path = tmp_path / "obs.xlsx"
    XlsxReportRenderer(obs).render(str(path))
    ws = load_workbook(str(path))["Trajectory summary"]
    assert ws.max_row == 1
    rules = [r for cf in ws.conditional_formatting for r in cf.rules]
    assert rules == [], f"no CF rules expected without a series; got {rules}"
