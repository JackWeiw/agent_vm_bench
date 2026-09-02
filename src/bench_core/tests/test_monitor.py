from __future__ import annotations

import subprocess
from pathlib import Path

import pytest

from bench_core.config import KernelConfig
from bench_core.monitor import MonitorConfig, MonitorController


def test_monitor_config_defaults_when_absent():
    cfg = KernelConfig.from_raw({"sandbox": {"total_count": 1}})
    assert isinstance(cfg.monitor, MonitorConfig)
    assert cfg.monitor.enabled == "auto"
    assert cfg.monitor.vmm == "auto"
    assert cfg.monitor.capture == "auto"
    assert cfg.monitor.interval == 3
    assert cfg.monitor.merge_report is True
    assert cfg.monitor.report_timeout == 300
    assert cfg.monitor.log_dir is None
    assert cfg.monitor.disks == "all"


def test_monitor_config_from_raw_overrides():
    raw = {
        "sandbox": {"total_count": 1},
        "monitor": {
            "enabled": "true",
            "vmm": "qemu",
            "interval": 5,
            "capture": "false",
            "numa": "0,1",
            "disks": "sda,nvme0n1",
            "stress_file": "/tmp/lock",
            "log_dir": "out/vm",
            "merge_report": False,
            "report_timeout": 120,
        },
    }
    cfg = KernelConfig.from_raw(raw)
    assert cfg.monitor.enabled == "true"
    assert cfg.monitor.vmm == "qemu"
    assert cfg.monitor.interval == 5
    assert cfg.monitor.capture == "false"
    assert cfg.monitor.numa == "0,1"
    assert cfg.monitor.disks == "sda,nvme0n1"
    assert cfg.monitor.stress_file == "/tmp/lock"
    assert cfg.monitor.log_dir == "out/vm"
    assert cfg.monitor.merge_report is False
    assert cfg.monitor.report_timeout == 120


class _StubProvider:
    """Minimal provider stub for MonitorController (no SDK)."""

    def __init__(self, name="stub", vmm_type=None, test_duration=300):
        self.name = name
        self.vmm_type = vmm_type
        self.test_duration = test_duration


def _cfg(**over):
    raw = {"sandbox": {"total_count": 1}, "report": {"output_dir": "out"}}
    if over:
        raw["monitor"] = over
    return KernelConfig.from_raw(raw)


def test_start_skips_when_no_vmm(caplog):
    cfg = _cfg()
    prov = _StubProvider(vmm_type=None)
    mc = MonitorController(cfg, prov)
    mc.start()
    assert mc._started is False
    assert any("disabled" in r.message for r in caplog.records)


def test_start_skips_when_binary_missing(monkeypatch, caplog):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: None)
    prov = _StubProvider(vmm_type="firecracker")
    mc = MonitorController(_cfg(), prov)
    mc.start()
    assert mc._started is False
    assert any("binary not found" in r.message for r in caplog.records)


def test_command_construction(monkeypatch, tmp_path):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    prov = _StubProvider(vmm_type="firecracker")
    mc = MonitorController(_cfg(stress_file=str(tmp_path / "lock")), prov)
    cmd = mc._cmd
    assert "--vmm" in cmd and "firecracker" in cmd
    assert "--stress-file" in cmd
    assert "--auto-skip" in cmd and "--enable-capture" in cmd  # capture=auto
    assert "-i" in cmd and "3" in cmd
    assert "--numa" in cmd
    assert "--disks" in cmd and "all" in cmd  # default
    assert "-t" in cmd  # hard upper-bound timer


def test_command_disks_override(monkeypatch, tmp_path):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    prov = _StubProvider(vmm_type="firecracker")
    mc = MonitorController(_cfg(disks="sda,nvme0n1", stress_file=str(tmp_path / "lock")), prov)
    cmd = mc._cmd
    # the custom disk list reaches the vm-monitor CLI verbatim
    i = cmd.index("--disks")
    assert cmd[i + 1] == "sda,nvme0n1"


def test_command_capture_false_omits_flags(monkeypatch, tmp_path):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    prov = _StubProvider(vmm_type="firecracker")
    mc = MonitorController(_cfg(capture="false", stress_file=str(tmp_path / "lock")), prov)
    assert "--enable-capture" not in mc._cmd
    assert "--auto-skip" not in mc._cmd


def test_start_removes_stale_lock(monkeypatch, tmp_path, caplog):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    monkeypatch.setattr("bench_core.monitor.subprocess.Popen", lambda *a, **kw: _FakeProc())
    lock = tmp_path / "lock"
    lock.touch()
    prov = _StubProvider(vmm_type="firecracker")
    mc = MonitorController(_cfg(stress_file=str(lock), log_dir=str(tmp_path)), prov)
    mc.start()
    assert mc._started is True
    assert not lock.exists()
    assert any("stale lock" in r.message.lower() for r in caplog.records)


def test_start_degrades_when_log_dir_unwritable(monkeypatch, tmp_path, caplog):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    # Create a regular file; using it as a parent dir makes mkdir() fail (FileExistsError is OSError).
    blocker = tmp_path / "blocker"
    blocker.write_text("not a dir")
    bad_log_dir = str(blocker / "subdir")
    prov = _StubProvider(vmm_type="firecracker")
    mc = MonitorController(_cfg(stress_file=str(tmp_path / "lock"), log_dir=bad_log_dir), prov)
    mc.start()
    assert mc._started is False
    assert any("disabled" in r.message for r in caplog.records)


class _FakeProc:
    def __init__(self):
        self.returncode = None
        self.terminated = False

    def poll(self):
        return self.returncode

    def terminate(self):
        self.terminated = True
        self.returncode = -15

    def wait(self, timeout=None):
        return self.returncode

    def kill(self):
        self.returncode = -9


def test_begin_end_stress_lifecycle(monkeypatch, tmp_path):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    monkeypatch.setattr("bench_core.monitor.subprocess.Popen", lambda *a, **kw: _FakeProc())
    lock = tmp_path / "lock"
    mc = MonitorController(_cfg(stress_file=str(lock), log_dir=str(tmp_path)), _StubProvider(vmm_type="firecracker"))
    mc.start()
    assert mc.stress_window is None  # not begun yet

    mc.begin_stress()
    assert lock.exists()
    assert mc._begin_ts is not None

    mc.end_stress()
    assert not lock.exists()
    assert mc._end_ts is not None
    assert mc.stress_window is not None and mc.stress_window >= 0.0


def test_end_stress_idempotent_when_no_lock(monkeypatch, tmp_path):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    monkeypatch.setattr("bench_core.monitor.subprocess.Popen", lambda *a, **kw: _FakeProc())
    lock = tmp_path / "lock"
    mc = MonitorController(_cfg(stress_file=str(lock), log_dir=str(tmp_path)), _StubProvider(vmm_type="firecracker"))
    mc.start()
    mc.begin_stress()
    lock.unlink()  # simulate external removal
    mc.end_stress()  # must not raise


def test_begin_end_noop_when_not_started():
    mc = MonitorController(_cfg(), _StubProvider(vmm_type=None))
    mc.begin_stress()
    mc.end_stress()
    assert mc.stress_window is None


def test_stop_collects_report_and_closes_handles(monkeypatch, tmp_path):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    monkeypatch.setattr("bench_core.monitor.subprocess.Popen", lambda *a, **kw: _FakeProc())
    mc = MonitorController(
        _cfg(stress_file=str(tmp_path / "lock"), log_dir=str(tmp_path), report_timeout=2),
        _StubProvider(vmm_type="firecracker"),
    )
    mc.start()
    # pre-create the report so stop() finds it immediately
    (tmp_path / "analysis_report.xlsx").write_text("x")
    artifacts = mc.stop()
    assert mc.report_xlsx == tmp_path / "analysis_report.xlsx"
    assert artifacts and artifacts[0] == mc.report_xlsx
    assert mc._stdout_fh is None and mc._stderr_fh is None  # closed


def test_stop_kills_overdue_process(monkeypatch, tmp_path):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    proc = _FakeProc()
    monkeypatch.setattr("bench_core.monitor.subprocess.Popen", lambda *a, **kw: proc)
    mc = MonitorController(
        _cfg(stress_file=str(tmp_path / "lock"), log_dir=str(tmp_path), report_timeout=1),
        _StubProvider(vmm_type="firecracker"),
    )
    mc.start()
    # no xlsx ever appears; proc never exits on its own (poll() stays None)
    mc.stop()
    assert proc.terminated is True  # terminate() called


def test_stop_noop_when_not_started():
    mc = MonitorController(_cfg(), _StubProvider(vmm_type=None))
    assert mc.stop() == []


def test_stop_handles_dead_subprocess(monkeypatch, tmp_path, caplog):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    proc = _FakeProc()
    proc.returncode = 1  # already dead, no report
    monkeypatch.setattr("bench_core.monitor.subprocess.Popen", lambda *a, **kw: proc)
    mc = MonitorController(
        _cfg(stress_file=str(tmp_path / "lock"), log_dir=str(tmp_path), report_timeout=2),
        _StubProvider(vmm_type="firecracker"),
    )
    mc.start()
    mc.stop()
    assert mc.report_xlsx is None
    assert any("without report" in r.message for r in caplog.records)


def _make_src_xlsx(path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "VM_Stats"
    ws.cell(row=1, column=1, value="vm")
    ws.cell(row=2, column=1, value="fc-1")
    wb.create_sheet("NUMA_Overview")
    wb["NUMA_Overview"].cell(row=1, column=1, value="node0")
    wb.save(path)


def test_merge_source_returns_report_when_enabled(monkeypatch, tmp_path):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    src = tmp_path / "analysis_report.xlsx"
    _make_src_xlsx(src)
    mc = MonitorController(
        _cfg(stress_file=str(tmp_path / "lock"), log_dir=str(tmp_path)), _StubProvider(vmm_type="firecracker")
    )
    mc.start()
    mc.report_xlsx = src  # pretend stop() found it
    assert mc.merge_source() == src


def test_merge_source_none_when_no_report(monkeypatch, tmp_path):
    mc = MonitorController(_cfg(), _StubProvider(vmm_type="firecracker"))
    # report_xlsx is None -> None, no raise
    assert mc.merge_source() is None


def test_merge_source_none_when_disabled(monkeypatch, tmp_path):
    monkeypatch.setattr("bench_core.monitor.shutil.which", lambda _: "/fake/vm-monitor")
    src = tmp_path / "analysis_report.xlsx"
    _make_src_xlsx(src)
    mc = MonitorController(
        _cfg(merge_report=False, stress_file=str(tmp_path / "lock"), log_dir=str(tmp_path)),
        _StubProvider(vmm_type="firecracker"),
    )
    mc.report_xlsx = src
    assert mc.merge_source() is None


def test_merge_source_none_when_missing(monkeypatch, tmp_path, caplog):
    mc = MonitorController(_cfg(), _StubProvider(vmm_type="firecracker"))
    mc.report_xlsx = tmp_path / "does-not-exist.xlsx"
    # must not raise; warns and returns None
    assert mc.merge_source() is None
    assert any("report missing" in r.message.lower() for r in caplog.records)
