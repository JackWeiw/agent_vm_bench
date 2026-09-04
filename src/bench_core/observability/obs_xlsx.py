"""Phase 3: xlsx observability renderer over the shared ReplayObservability model.

One data layer (ReplayObservability), two renderers (text ReportFormatter +
this xlsx). Emits a multi-sheet workbook; openpyxl is a core dep. Semantic
metrics (throughput / retry / trajectory percentiles) are owned by the model;
this renderer only flattens raw per-sandbox series into percentile rows for
the step / lifecycle / trajectory tables (the model exposes percentile dicts,
not the raw lists + counts those tables need). A render failure (e.g. openpyxl
missing on a minimal install) is caught by the caller (run_benchmark) which
falls back to the text report + a warning; this module raises normally.
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from bench_core.observability.trajectory_summary import SEG_KEYS, trajectory_summaries
from bench_core.utils import calc_percentiles

if TYPE_CHECKING:
    from bench_core.observability.replay_obs import ReplayObservability

# Host-level vm_monitor sheets copied into the obs workbook during the single
# render pass (cell values only). Kept here -- the renderer owns the merge -- so
# there is no second openpyxl load/save round-trip on the obs file (a round-trip
# would drop every chart and PNG embedded by this renderer).
_MERGE_SHEETS = ("VM_Stats", "NUMA_Overview", "DevKit_TopDown")


def _write_table(ws, headers: list[str], rows: list[list]) -> int:
    """Write a header row (bold) + data rows to a worksheet; return the row the
    header landed on.

    Bold the row just appended (tracked via ``ws.max_row`` after the append),
    not always row 1, so a sheet with multiple sub-tables (e.g. Trajectory
    summary) styles each header correctly. Callers that need to anchor a chart
    Reference on the header (titles_from_data) MUST use the returned row rather
    than ``ws.max_row + 1`` computed before the write: openpyxl reports
    ``max_row == 1`` for a fresh empty sheet, so a pre-write ``+1`` lands one row
    past the header and points chart title refs at the first data row.
    """
    bold = Font(bold=True)
    ws.append(headers)
    header_row = ws.max_row
    for c in ws[header_row]:
        c.font = bold
    for row in rows:
        ws.append(row)
    return header_row


def _add_line_chart(
    ws,
    title: str,
    y_title: str,
    cat_col: int,
    data_cols: list[int],
    n_rows: int,
    anchor: str,
    *,
    header_row: int = 1,
    first_data_row: int = 2,
) -> None:
    """Add a LineChart with one series per data column; categories from cat_col.

    n_rows is the number of DATA rows (excluding the header). The header lives
    at ``header_row``; data occupies ``header_row+1 .. header_row+n_rows``.
    """
    if n_rows <= 0:
        return
    ch = LineChart()
    ch.title = title
    ch.y_axis.title = y_title
    ch.x_axis.title = ws.cell(header_row, cat_col).value
    ch.height = 8
    ch.width = 16
    for col in data_cols:
        ref = Reference(ws, min_col=col, min_row=header_row, max_row=header_row + n_rows)
        ch.add_data(ref, titles_from_data=True)
    cats = Reference(ws, min_col=cat_col, min_row=first_data_row, max_row=header_row + n_rows)
    ch.set_categories(cats)
    ws.add_chart(ch, anchor)


def _add_trajectory_cost_chart(ws, headers: list[str], header_row: int, n_traj: int) -> None:
    """Stacked bar of per-trajectory exec / resume / pause sums (bar height ==
    slice_total_sum) so the cost-attribution table's "where did each
    trajectory's wall-clock go" reads at a glance.

    Categories = trajectory_id (col 1); the three series take their titles from
    the table header (exec_sum_s / resume_sum_s / pause_sum_s) -- located by
    HEADER NAME so a column reorder never silently breaks the chart. The anchor
    is COMPUTED from the header count (1-col gap past the table) so adding
    columns never lands the floating chart on top of data: the old hardcoded
    "N2" overlapped data once the failure cluster widened the table to 15 cols
    (col N became a data column)."""
    if n_traj <= 0:
        return
    ch = BarChart()
    ch.type = "col"
    ch.grouping = "stacked"
    ch.overlap = 100
    ch.title = "Per-trajectory cost decomposition (s)"
    ch.y_axis.title = "seconds"
    ch.x_axis.title = "trajectory_id"
    ch.height = 8
    ch.width = 16
    for name in ("exec_sum_s", "resume_sum_s", "pause_sum_s"):
        col = headers.index(name) + 1
        ref = Reference(ws, min_col=col, min_row=header_row, max_row=header_row + n_traj)
        ch.add_data(ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + n_traj)
    ch.set_categories(cats)
    anchor_col = get_column_letter(len(headers) + 2)
    ws.add_chart(ch, f"{anchor_col}2")


def _pcts_row(label: str, values: list[float]) -> list:
    """One row: [label, n, min, max, avg, p50, p95, p99] for a latency list."""
    s = calc_percentiles(values)
    return [label, len(values), s["min"], s["max"], s["avg"], s["p50"], s["p95"], s["p99"]]


def _round_or_none(v, ndigits: int = 3):
    """Round a series float to ndigits, preserving None/missing as None."""
    if v is None:
        return None
    try:
        return round(float(v), ndigits)
    except (TypeError, ValueError):
        return None


# Overview dashboard styling. The metric-name column (col A) gets a light fill
# so the label stands out from its value, and section banners (Run / Throughput
# / Admission & QPS / Retry impact) delimit the areas that used to be separate
# scalar sheets. The three standalone sheets are gone; Overview is the single
# summary -- duplicating their scalars across two places would be the
# redundancy the lean-docs rule warns against.
_OVERVIEW_NAME_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
_OVERVIEW_BANNER_FILL = PatternFill(fill_type="solid", fgColor="2E5496")
_OVERVIEW_BANNER_FONT = Font(bold=True, color="FFFFFF")
_OVERVIEW_NAME_FONT = Font(bold=True)


def _overview_banner(ws, title: str) -> None:
    """A filled section-divider row spanning the first 3 columns."""
    ws.append([title])
    row = ws.max_row
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.fill = _OVERVIEW_BANNER_FILL
        cell.font = _OVERVIEW_BANNER_FONT


def _overview_kv(ws, name, value) -> None:
    """A metric row: name (filled + bold col A) | value (col B)."""
    ws.append([name, value])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.fill = _OVERVIEW_NAME_FILL
    cell.font = _OVERVIEW_NAME_FONT


def _overview_subtable(ws, headers: list[str], rows: list[list]) -> None:
    """A small multi-column sub-table (e.g. the per-operation dispatched/waiting
    breakdown) inside the Overview dashboard. Header bolded; col A of each data
    row filled so the data-name column stays visually consistent across the
    sheet."""
    bold = Font(bold=True)
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = bold
    for row in rows:
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).fill = _OVERVIEW_NAME_FILL


class XlsxReportRenderer:
    """Render a ReplayObservability model to a multi-sheet xlsx workbook."""

    def __init__(
        self,
        observability: ReplayObservability,
        series_path: Path | None = None,
        host_xlsx: Path | None = None,
    ) -> None:
        self.obs = observability
        self.series_path = series_path
        self.host_xlsx = host_xlsx

    def render(self, path: str | Path) -> None:
        """Write the workbook to ``path`` (overwrites). Creates all 8 base sheets.

        Overview is the consolidated summary dashboard (Run / Throughput /
        Admission & QPS / Retry impact sections); the former standalone scalar
        sheets were folded into it.
        """
        wb = Workbook()
        # openpyxl seeds one default sheet; remove it after building named sheets.
        wb.remove(wb.active)
        self._sheet_overview(wb)
        self._sheet_per_step_timings(wb)
        self._sheet_lifecycle_overhead(wb)
        self._sheet_trajectory_summary(wb)
        self._sheet_step_detail(wb)
        self._sheet_concurrency_states(wb)
        gantt_png = Path(path).parent / f"{Path(path).stem.split('_obs')[0]}_gantt.png"
        self._sheet_gantt(wb, out_png=gantt_png)
        self._sheet_snapshot_sizes(wb)
        # Merge host-level sheets last, in the same write pass -- avoids a second
        # load/save round-trip that would drop the charts/PNG embedded above.
        self._sheet_host_merged(wb)
        wb.save(path)

    # --- sheets ---

    def _sheet_overview(self, wb: Workbook) -> None:
        """The single summary dashboard.

        Consolidates the former Admission & QPS / Throughput & overcommit /
        Retry impact scalar sheets into grouped, color-coded sections here, so
        the whole run reads on one sheet. The metric-name column is filled +
        bolded to make labels prominent against their values; section banners
        delimit the areas.
        """
        ws = wb.create_sheet("Overview")
        obs = self.obs
        cfg = obs.config
        total_success = sum(s.replay_metrics.success_count for s in obs.states.values())
        total_failed = sum(s.replay_metrics.failed_count for s in obs.states.values())
        na = "n/a"

        # --- Run (identity + outcome) ---
        _overview_banner(ws, "Run")
        for name, value in (
            ("workflow_type", cfg.workflow_type),
            ("replay_mode", cfg.replay_mode),
            ("total_count", cfg.total_count),
            ("running_concurrency", cfg.replay_running_concurrency),
            ("test_duration", cfg.test_duration),
            ("wall_sec", obs.wall_sec),
            ("total_steps", obs.total_steps),
            ("success", total_success),
            ("failed", total_failed),
            ("overcommit_ratio", obs.overcommit_ratio),
        ):
            _overview_kv(ws, name, value)

        # --- Throughput & overcommit ---
        ws.append([])
        _overview_banner(ws, "Throughput & overcommit")
        for name, value in (
            ("steps_per_sec", obs.steps_per_sec if obs.steps_per_sec is not None else na),
            ("effective_parallelism", obs.effective_parallelism if obs.effective_parallelism is not None else na),
            ("exec_wall_utilization", obs.exec_wall_utilization if obs.exec_wall_utilization is not None else na),
            ("concurrency", obs.concurrency),
        ):
            _overview_kv(ws, name, value)

        # --- Admission & QPS (running-slot lease + QPS rate limiter) ---
        snap = obs.admission_snapshot or {}
        rs = snap.get("running_slots") or {}
        ql = snap.get("qps_limiter") or {}
        if rs or ql:
            ws.append([])
            _overview_banner(ws, "Admission & QPS")
            admission_rows: list[tuple] = []
            if rs:
                for k in ("maximum", "active", "peak_active", "granted", "average_queue_wait_sec", "waiting"):
                    admission_rows.append((f"running_slots.{k}", rs.get(k)))
            if ql:
                for k in ("qps", "inflight_cap", "in_flight", "dispatched", "average_wait_sec", "max_wait_sec"):
                    admission_rows.append((k, ql.get(k)))
            for name, value in admission_rows:
                _overview_kv(ws, name, value)
            # Per-operation dispatched/waiting sub-table (the lifecycle calls
            # the limiter sees, broken down by resume/pause/create/...).
            dispatched_by_op = (ql.get("dispatched_by_operation") or {}) if ql else {}
            waiting_by_op = (ql.get("waiting_by_operation") or {}) if ql else {}
            if dispatched_by_op or waiting_by_op:
                ws.append([])
                _overview_subtable(
                    ws,
                    ["operation", "dispatched", "waiting"],
                    [
                        [op, dispatched_by_op.get(op, 0), waiting_by_op.get(op, 0)]
                        for op in sorted(set(dispatched_by_op) | set(waiting_by_op))
                    ],
                )

        # --- Retry impact ---
        ws.append([])
        _overview_banner(ws, "Retry impact")
        retry_rows: list[tuple] = [
            ("retry_count", obs.retry_count),
            ("time_lost_to_retry_sec", obs.time_lost_to_retry_sec),
            ("retries_per_slice_p95", obs.retries_per_slice_p95),
        ]
        for op, n in obs.retry_count_by_op.items():
            retry_rows.append((f"retry_queued:{op}", n))
        for name, value in retry_rows:
            _overview_kv(ws, name, value)

        # Wide metric names (running_slots.average_queue_wait_sec, retry_queued:*)
        # need a generous col A to read without truncation.
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 16

    def _sheet_per_step_timings(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Per-step timings")
        all_lat: list[float] = []
        for s in self.obs.states.values():
            all_lat.extend(s.replay_metrics.latencies)
        rows = [_pcts_row("latency", all_lat)]
        # per action_type
        by_action: dict[str, list[float]] = {}
        for s in self.obs.states.values():
            for act, vals in s.replay_metrics.action_type_latencies.items():
                by_action.setdefault(act, []).extend(vals)
        for act in sorted(by_action):
            rows.append(_pcts_row(act, by_action[act]))
        _write_table(ws, ["bucket", "n", "min", "max", "avg", "p50", "p95", "p99"], rows)
        # Per-step detail rows (latency per step, concatenated across sandboxes).
        step_rows = [[i + 1, round(v * 1000, 1)] for i, v in enumerate(all_lat)]
        if step_rows:
            ws.append([])
            ws.append(["step_index", "latency_ms"])
            hdr = ws.max_row
            for c in ws[hdr]:
                c.font = Font(bold=True)
            for r in step_rows:
                ws.append(r)
            _add_line_chart(
                ws,
                "Per-step latency",
                "ms",
                ws.cell(hdr, 1).column,
                [ws.cell(hdr, 2).column],
                len(step_rows),
                f"A{ws.max_row + 2}",
                header_row=hdr,
                first_data_row=hdr + 1,
            )

    def _sheet_lifecycle_overhead(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Lifecycle overhead")
        obs = self.obs
        lists = {
            "resume": [],
            "pause": [],
            "slice_total": [],
            "slot_held": [],
            "interaction": [],
        }
        for s in obs.states.values():
            m = s.replay_metrics
            lists["resume"].extend(m.resume_secs)
            lists["pause"].extend(m.pause_secs)
            lists["slice_total"].extend(m.slice_total_secs)
            lists["slot_held"].extend(m.running_slot_held_secs)
            lists["interaction"].extend(m.interaction_total_secs)
        rows = [_pcts_row(label, vals) for label, vals in lists.items()]
        _write_table(ws, ["segment", "n", "min", "max", "avg", "p50", "p95", "p99"], rows)
        # Per-step detail + line chart (resume/pause/slice ms over step index).
        n = len(lists["resume"])
        if n:
            ws.append([])
            ws.append(["step_index", "resume_ms", "pause_ms", "slice_ms"])
            hdr = ws.max_row
            for c in ws[hdr]:
                c.font = Font(bold=True)
            for i in range(n):
                ws.append(
                    [
                        i + 1,
                        round(lists["resume"][i] * 1000, 1),
                        round(lists["pause"][i] * 1000, 1),
                        round(lists["slice_total"][i] * 1000, 1),
                    ]
                )
            _add_line_chart(
                ws,
                "Per-step lifecycle overhead",
                "ms",
                1,
                [2, 3, 4],
                n,
                f"A{ws.max_row + 2}",
                header_row=hdr,
                first_data_row=hdr + 1,
            )

    def _sheet_trajectory_summary(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Trajectory summary")
        obs = self.obs
        # Per-trajectory cumulative cost attribution (seconds). Sums, not
        # percentiles: a per-instance summary answers "where did this
        # trajectory's wall-clock go" (pause vs resume vs exec vs waits), and
        # per-step distributions already live in the Step detail sheet (filter
        # by trajectory_id) and pooled in the Lifecycle overhead sheet. n_steps
        # counts every recorded step event (success + slice_failed); a failed
        # slice contributes 0 to the sums (it did no work) but still counts as
        # an attempted step, so avg_slice reflects the per-attempt cost.
        wrote_profile = False
        if self.series_path is not None and Path(self.series_path).exists():
            from bench_core.observability.lifecycle_series import load_events

            # Union-keyed (step + trajectory_create/kill/failed) so a trajectory
            # that failed at create (zero step events) still appears -- the worst
            # trajectories are the ones a comparison most needs to surface.
            summaries = trajectory_summaries(load_events(Path(self.series_path)))
            if summaries:
                headers = [
                    "trajectory_id",
                    "n_steps",
                    "n_failed",
                    "n_timeout",
                    "success_rate",
                    "slice_total_sum_s",
                    "exec_sum_s",
                    "resume_sum_s",
                    "pause_sum_s",
                    "interaction_total_sum_s",
                    "slot_wait_sum_s",
                    "resume_queue_wait_sum_s",
                    "pause_queue_wait_sum_s",
                    "running_slot_held_sum_s",
                    "avg_slice_s",
                ]
                sr_idx = headers.index("success_rate")
                rows = []
                for s in summaries:
                    sums = s["sums"]
                    rows.append(
                        [
                            s["trajectory_id"],
                            s["n_steps"],
                            s["n_failed"],
                            s["n_timeout"],
                            s["success_rate"],  # None when 0 steps attempted
                            *(round(sums[k], 3) for k in SEG_KEYS),
                            round(s["avg_slice"], 3),
                        ]
                    )
                header_row = _write_table(ws, headers, rows)
                wrote_profile = True
                _add_trajectory_cost_chart(ws, headers, header_row, len(rows))
                # At-a-glance outlier highlighting the reference's per-trial
                # table lacks: data bars on the two cost drivers (total slice +
                # slot queueing -- longer bar = slower / more queueing) and a red
                # color scale on the failure count. Only when trajectories
                # exist, so a series-less sheet stays empty (max_row==1, no CF
                # on an empty range). success_rate=None (0 steps) gets a hover
                # comment so the blank cell is not mistaken for "0 = all failed".
                first, last = header_row + 1, header_row + len(rows)
                for ri, row in enumerate(rows, start=first):
                    if row[sr_idx] is None:
                        ws.cell(ri, sr_idx + 1).comment = Comment(
                            "no steps attempted (created/killed with 0 steps); " "0.0 = all steps failed",
                            "bench-core",
                        )

                def _col_range(col: int) -> str:
                    letter = get_column_letter(col)
                    return f"{letter}{first}:{letter}{last}"

                ws.conditional_formatting.add(
                    _col_range(headers.index("slice_total_sum_s") + 1),
                    DataBarRule(start_type="min", end_type="max", color="638EC6"),
                )
                ws.conditional_formatting.add(
                    _col_range(headers.index("slot_wait_sum_s") + 1),
                    DataBarRule(start_type="min", end_type="max", color="638EC6"),
                )
                ws.conditional_formatting.add(
                    _col_range(headers.index("n_failed") + 1),
                    ColorScaleRule(
                        start_type="num",
                        start_value=0,
                        start_color="FFFFFF",
                        end_type="max",
                        end_color="FF0000",
                    ),
                )
        # Trajectory-mode-only: ephemeral create/kill lifecycle (not in any other
        # sheet). slot_held is intentionally omitted -- it is already pooled in
        # the Lifecycle overhead sheet.
        if getattr(obs.config, "replay_mode", None) == "trajectory" and any(
            s.replay_metrics.create_secs for s in obs.states.values()
        ):
            if wrote_profile:
                ws.append([])
            rows = [_pcts_row("create_sec", [v for s in obs.states.values() for v in s.replay_metrics.create_secs])]
            rows.append(_pcts_row("kill_sec", [v for s in obs.states.values() for v in s.replay_metrics.kill_secs]))
            _write_table(ws, ["segment", "n", "min", "max", "avg", "p50", "p95", "p99"], rows)

    def _sheet_step_detail(self, wb: Workbook) -> None:
        """Per-trajectory per-step raw rows -- the detail behind the percentile tables.

        The Per-step timings sheet pools every step (no trajectory_id); the
        Trajectory summary sheet shows per-trajectory percentiles only. This sheet
        is the missing middle: one row per recorded step event (success and
        slice_failed), broken down by trajectory, so each trajectory's per-step
        resume/exec/pause/slice timings are inspectable directly. Sourced from the
        series JSONL; empty when there is no series file (e.g. a minimal install).

        Duration columns are seconds (matching the reference step-detail.csv), not
        the milliseconds used by the Per-step/Lifecycle chart sheets. The
        sub-segments sit next to their parent total so the sum invariants are
        visually verifiable: ``resume_sec == resume_queue_wait_sec +
        resume_api_sec + resume_ready_wait_sec`` and ``pause_sec ==
        pause_queue_wait_sec + pause_api_sec``.
        """
        ws = wb.create_sheet("Step detail")
        headers = [
            "trajectory_id",
            "sandbox_index",
            "round_id",
            "step_index",
            "action_type",
            "slice_failed",
            "resume_sec",
            "resume_queue_wait_sec",
            "resume_api_sec",
            "resume_ready_wait_sec",
            "exec_sec",
            "pause_sec",
            "pause_queue_wait_sec",
            "pause_api_sec",
            "slice_total_sec",
            "interaction_total_sec",
            "slot_contention_wait_sec",
            "running_slot_held_sec",
            "exit_code",
            "timed_out",
        ]
        if self.series_path is None or not Path(self.series_path).exists():
            _write_table(ws, headers, [])
            return
        from bench_core.observability.lifecycle_series import load_events

        rows: list[list] = []
        for ev in load_events(Path(self.series_path)):
            if ev.get("event") != "step":
                continue
            rows.append(
                [
                    ev.get("trajectory_id") or "",
                    ev.get("sandbox_index"),
                    ev.get("round_id"),
                    ev.get("step_index"),
                    ev.get("action_type") or "",
                    bool(ev.get("slice_failed")),
                    _round_or_none(ev.get("resume_sec")),
                    _round_or_none(ev.get("resume_queue_wait_sec")),
                    _round_or_none(ev.get("resume_api_sec")),
                    _round_or_none(ev.get("resume_ready_wait_sec")),
                    _round_or_none(ev.get("exec_sec")),
                    _round_or_none(ev.get("pause_sec")),
                    _round_or_none(ev.get("pause_queue_wait_sec")),
                    _round_or_none(ev.get("pause_api_sec")),
                    _round_or_none(ev.get("slice_total_sec")),
                    _round_or_none(ev.get("interaction_total_sec")),
                    _round_or_none(ev.get("slot_contention_wait_sec")),
                    _round_or_none(ev.get("running_slot_held_sec")),
                    ev.get("exit_code"),
                    bool(ev.get("timed_out")),
                ]
            )
        # Sort by trajectory, then sandbox, then step -- so each trajectory's
        # steps read top-to-bottom in execution order.
        rows.sort(key=lambda r: (str(r[0]), r[1] if r[1] is not None else 0, r[3] if r[3] is not None else 0))
        _write_table(ws, headers, rows)
        if rows:
            # Freeze the header + enable autofilter so the user can pivot by
            # trajectory / action_type / exit_code without re-sorting in Excel.
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

    def _sheet_concurrency_states(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Concurrency states")
        headers = ["second", "pausing", "paused", "resuming", "exec", "active"]
        if self.series_path is None or not Path(self.series_path).exists():
            _write_table(ws, headers, [])
            return
        from bench_core.observability.lifecycle_reconstruct import reconstruct_concurrency
        from bench_core.observability.lifecycle_series import load_events

        bins = reconstruct_concurrency(load_events(Path(self.series_path)))
        rows = [[b["second"], b["pausing"], b["paused"], b["resuming"], b["exec"], b["active"]] for b in bins]
        _write_table(ws, headers, rows)
        if rows:
            _add_line_chart(ws, "Tasks per state (dominant each second)", "count", 1, [2, 3, 4, 5, 6], len(rows), "H2")

    def _sheet_gantt(self, wb: Workbook, *, out_png: Path | None = None) -> None:
        ws = wb.create_sheet("Gantt")
        ws.append(["Gantt (per-sandbox phase timeline)"])
        if self.series_path is None or not Path(self.series_path).exists():
            ws.append(["no series data"])
            return
        try:
            import matplotlib

            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
        except ImportError:
            ws.append(["matplotlib not installed; install with 'pip install matplotlib'"])
            return

        from bench_core.observability.lifecycle_reconstruct import gantt_segments
        from bench_core.observability.lifecycle_series import load_events

        rows = gantt_segments(load_events(Path(self.series_path)))
        if not rows:
            ws.append(["no step events"])
            return
        color = {"pausing": "#d62728", "paused": "#ffd8a8", "resuming": "#9467bd", "exec": "#2ca02c"}
        n = len(rows)
        # Cap total height -- without a cap a 768-sandbox fleet yields a ~424 in
        # figure (explodes memory / renders unreadable). 40 in fits ~70 sandboxes
        # at 0.55 in/row; larger fleets scale the per-row height down proportionally.
        fig_h = min(max(8, n * 0.55 + 2), 40)
        fig, ax = plt.subplots(figsize=(28, fig_h))
        t0 = min(a for _, segs in rows for a, _, _ in segs)
        for yi, (name, segs) in enumerate(rows):
            for a, b, ph in segs:
                ax.barh(yi, b - a, left=a - t0, height=0.75, color=color[ph], edgecolor="none", zorder=3)
        ax.set_yticks(range(n))
        ax.set_yticklabels([r[0] for r in rows])
        # Cramped fleets: shrink the per-row label so 768 rows stay legible.
        if n > 60:
            ax.tick_params(axis="y", labelsize=4)
        ax.set_xlabel("Relative time (s, from earliest event)")
        ax.invert_yaxis()
        ax.grid(True, axis="x", linestyle=":", alpha=0.4)
        ax.legend(
            [plt.Rectangle((0, 0), 1, 1, fc=c) for c in color.values()],
            list(color.keys()),
            loc="upper right",
        )
        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, dpi=130, format="png")
        png_bytes = buf.getvalue()
        plt.close(fig)
        if out_png is not None:
            out_png.write_bytes(png_bytes)
        from openpyxl.drawing.image import Image as XLImage

        ws.add_image(XLImage(io.BytesIO(png_bytes)), "A3")

    def _sheet_snapshot_sizes(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Snapshot sizes")
        headers = [
            "pause_seq",
            "sandbox_index",
            "sandbox_id",
            "logical_mb",
            "disk_mb",
            "inherited_mb",
            "cumulative_mb",
            "generations",
            "files",
        ]
        if self.series_path is None or not Path(self.series_path).exists():
            _write_table(ws, headers, [])
            return
        from bench_core.observability.lifecycle_reconstruct import snapshot_rows
        from bench_core.observability.lifecycle_series import load_events

        rows = snapshot_rows(load_events(Path(self.series_path)))
        out = [
            [
                r["pause_seq"],
                r["sandbox_index"],
                r["sandbox_id"],
                r["logical_mb"],
                r["disk_mb"],
                r["inherited_mb"],
                r["cumulative_mb"],
                r["generations"],
                r["files"],
            ]
            for r in rows
        ]
        _write_table(ws, headers, out)
        if out:
            _add_line_chart(ws, "Snapshot size per pause", "MiB", 1, [4, 5, 7], len(out), "K2")

    def _sheet_host_merged(self, wb: Workbook) -> None:
        """Copy host-level vm_monitor sheets (cell values) into the workbook.

        Done in the single render pass so the charts/PNG embedded above survive
        (a post-save ``load_workbook``+``save`` round-trip would drop them).
        Cell values only -- vm_monitor's own charts stay in its source file.
        """
        if self.host_xlsx is None or not Path(self.host_xlsx).exists():
            return
        try:
            from openpyxl import load_workbook
        except ImportError:  # pragma: no cover - openpyxl is a core dep
            return
        try:
            src = load_workbook(self.host_xlsx, read_only=True)
        except Exception:
            # Missing/corrupt host report must never break the obs workbook.
            return
        try:
            for name in _MERGE_SHEETS:
                if name not in src.sheetnames or name in wb.sheetnames:
                    continue
                src_ws = src[name]
                dst_ws = wb.create_sheet(name)
                for row in src_ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
        finally:
            src.close()
