#!/usr/bin/env python3
"""Strictly verify the enhanced NYC taxi workbook benchmark output."""

from __future__ import annotations

import argparse
import csv
import gc
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_SHEETS = [
    "Raw_Sample",
    "Hourly_Summary",
    "Daily_Summary",
    "Zone_Summary",
    "Payment_Summary",
    "Fare_Distance_Bands",
    "Reconciliation",
]
FORMULA_CELLS = ("B5", "B6", "B7", "B8", "B9", "B13", "B14", "B15", "B16")
ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}
EXECUTIVE_CSV_ROWS = (
    ("Total Trips", "B5"),
    ("Total Fare Revenue", "B6"),
    ("Overall Average Fare", "B7"),
    ("Removal Rate", "B8"),
    ("Reconciliation Status", "B9"),
    ("Selected Scenario", "B12"),
    ("Selected Fare Multiplier", "B13"),
    ("Selected Trip Multiplier", "B14"),
    ("Projected Trips", "B15"),
    ("Projected Fare Revenue", "B16"),
)
RECONCILIATION_CSV_ROWS = tuple(range(2, 7)) + tuple(range(10, 21)) + (24, 25)


def parse_args() -> argparse.Namespace:
    """Parse verifier paths."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("recalc_json", type=Path)
    parser.add_argument("report_json", type=Path)
    parser.add_argument("executive_csv", type=Path)
    parser.add_argument("reconciliation_csv", type=Path)
    return parser.parse_args()


def add_check(
    checks: dict[str, Any],
    failures: list[str],
    name: str,
    ok: bool,
    detail: Any,
) -> None:
    """Record one explicit verifier check."""
    checks[name] = {"ok": bool(ok), "detail": detail}
    if not ok:
        failures.append(name)


def color_endswith(cell: Any, rgb: str) -> bool:
    """Return whether a cell uses the expected RGB font color."""
    color = cell.font.color
    if color is None or color.type != "rgb" or color.rgb is None:
        return False
    return str(color.rgb).upper().endswith(rgb.upper())


def count_formulas(workbook: Any) -> int:
    """Count formulas outside the large raw-data sheet."""
    total = 0
    for sheet in workbook.worksheets:
        if sheet.title == "Raw_Sample":
            continue
        for row in sheet.iter_rows():
            total += sum(
                isinstance(cell.value, str) and cell.value.startswith("=")
                for cell in row
            )
    return total


def verify_formula_workbook(
    workbook_path: Path,
    checks: dict[str, Any],
    failures: list[str],
) -> None:
    """Verify workbook structure, formulas, formatting, and Excel features."""
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        expected_order = ["Executive_Summary", *BASE_SHEETS]
        add_check(
            checks,
            failures,
            "sheet_order",
            workbook.sheetnames[: len(expected_order)] == expected_order,
            workbook.sheetnames,
        )
        add_check(
            checks,
            failures,
            "raw_sample_preserved",
            workbook["Raw_Sample"].max_row >= 100_001
            and workbook["Raw_Sample"].max_column >= 25,
            {
                "rows": workbook["Raw_Sample"].max_row,
                "columns": workbook["Raw_Sample"].max_column,
            },
        )
        executive = workbook["Executive_Summary"]
        formulas = {cell: executive[cell].value for cell in FORMULA_CELLS}
        add_check(
            checks,
            failures,
            "required_formulas",
            all(isinstance(value, str) and value.startswith("=") for value in formulas.values()),
            formulas,
        )
        cross_sheet = {
            cell: executive[cell].value
            for cell in ("B5", "B6", "B7", "B8")
        }
        add_check(
            checks,
            failures,
            "cross_sheet_formulas",
            all("!" in str(value) for value in cross_sheet.values()),
            cross_sheet,
        )
        total_formulas = count_formulas(workbook)
        add_check(
            checks,
            failures,
            "formula_count",
            total_formulas >= 45,
            total_formulas,
        )
        add_check(
            checks,
            failures,
            "scenario_default",
            executive["B12"].value == "Base",
            executive["B12"].value,
        )
        validations = list(executive.data_validations.dataValidation)
        add_check(
            checks,
            failures,
            "scenario_data_validation",
            any("B12" in str(validation.sqref) for validation in validations),
            [str(validation.sqref) for validation in validations],
        )
        add_check(
            checks,
            failures,
            "executive_charts",
            len(executive._charts) >= 2,
            len(executive._charts),
        )
        total_charts = sum(len(sheet._charts) for sheet in workbook.worksheets)
        add_check(
            checks,
            failures,
            "base_charts_preserved",
            total_charts >= 4
            and len(workbook["Hourly_Summary"]._charts) >= 1
            and len(workbook["Payment_Summary"]._charts) >= 1,
            total_charts,
        )
        add_check(
            checks,
            failures,
            "conditional_formatting",
            len(executive.conditional_formatting) >= 2,
            len(executive.conditional_formatting),
        )
        comment_cells = [
            cell.coordinate
            for row in executive.iter_rows()
            for cell in row
            if cell.comment is not None
        ]
        add_check(
            checks,
            failures,
            "source_comments",
            len(comment_cells) >= 2,
            comment_cells,
        )
        add_check(
            checks,
            failures,
            "freeze_panes",
            executive.freeze_panes is not None,
            str(executive.freeze_panes),
        )
        add_check(
            checks,
            failures,
            "font_color_convention",
            color_endswith(executive["B12"], "0000FF")
            and color_endswith(executive["B5"], "008000")
            and color_endswith(executive["B13"], "000000"),
            {
                "B12": str(executive["B12"].font.color),
                "B5": str(executive["B5"].font.color),
                "B13": str(executive["B13"].font.color),
            },
        )
        add_check(
            checks,
            failures,
            "no_external_links",
            not workbook._external_links,
            len(workbook._external_links),
        )
    finally:
        workbook.close()
        del workbook
        gc.collect()


def verify_cached_values(
    workbook_path: Path,
    checks: dict[str, Any],
    failures: list[str],
) -> None:
    """Verify LibreOffice-produced cached formula values."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        executive = workbook["Executive_Summary"]
        values = {cell: executive[cell].value for cell in FORMULA_CELLS}
        numeric_cells = ("B5", "B6", "B7", "B8", "B13", "B14", "B15", "B16")
        numeric_ok = all(
            isinstance(values[cell], (int, float)) and math.isfinite(float(values[cell]))
            for cell in numeric_cells
        )
        add_check(checks, failures, "cached_numeric_values", numeric_ok, values)
        add_check(
            checks,
            failures,
            "scenario_cached_values",
            numeric_ok
            and math.isclose(float(values["B13"]), 1.0, rel_tol=1e-6)
            and math.isclose(float(values["B14"]), 1.0, rel_tol=1e-6)
            and math.isclose(float(values["B15"]), float(values["B5"]), rel_tol=1e-6)
            and math.isclose(float(values["B16"]), float(values["B6"]), rel_tol=1e-6),
            values,
        )
        add_check(
            checks,
            failures,
            "reconciliation_status",
            str(values["B9"]).upper() == "PASS",
            values["B9"],
        )
        error_cells: list[str] = []
        for sheet_name in ["Executive_Summary", *BASE_SHEETS[1:]]:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value in ERROR_VALUES:
                        error_cells.append(f"{sheet_name}!{cell.coordinate}")
        add_check(
            checks,
            failures,
            "zero_cached_formula_errors",
            not error_cells,
            error_cells,
        )
    finally:
        workbook.close()


def verify_recalc_report(
    recalc_path: Path,
    checks: dict[str, Any],
    failures: list[str],
) -> None:
    """Verify the bundled xlsx recalc script result."""
    try:
        recalc = json.loads(recalc_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        add_check(checks, failures, "recalc_report", False, str(exc))
        return
    ok = (
        recalc.get("status") == "success"
        and recalc.get("total_errors") == 0
        and int(recalc.get("total_formulas") or 0) >= 45
    )
    add_check(checks, failures, "recalc_report", ok, recalc)


def _csv_matches(actual: str, expected: Any) -> bool:
    if isinstance(expected, (int, float)) and not isinstance(expected, bool):
        try:
            return math.isclose(float(actual), float(expected), rel_tol=1e-9, abs_tol=1e-6)
        except ValueError:
            return False
    return actual == str(expected)


def _read_metric_csv(path: Path) -> tuple[list[dict[str, str]], str | None]:
    try:
        with path.open(encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames != ["Metric", "Value"]:
                return [], f"unexpected header: {reader.fieldnames}"
            return list(reader), None
    except OSError as exc:
        return [], str(exc)


def verify_csv_deliverables(
    workbook_path: Path,
    executive_csv: Path,
    reconciliation_csv: Path,
    checks: dict[str, Any],
    failures: list[str],
) -> None:
    """Verify both CSV files were derived from the final cached workbook."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        executive = workbook["Executive_Summary"]
        reconciliation = workbook["Reconciliation"]
        expected_executive = [(label, executive[cell].value) for label, cell in EXECUTIVE_CSV_ROWS]
        expected_reconciliation = [
            (reconciliation.cell(row=row, column=1).value, reconciliation.cell(row=row, column=2).value)
            for row in RECONCILIATION_CSV_ROWS
        ]
    finally:
        workbook.close()

    for name, path, expected in (
        ("executive_csv", executive_csv, expected_executive),
        ("reconciliation_csv", reconciliation_csv, expected_reconciliation),
    ):
        rows, error = _read_metric_csv(path)
        actual = [(row.get("Metric", ""), row.get("Value", "")) for row in rows]
        ok = (
            error is None
            and len(actual) == len(expected)
            and all(
                actual_label == str(expected_label) and _csv_matches(actual_value, expected_value)
                for (actual_label, actual_value), (expected_label, expected_value) in zip(actual, expected)
            )
        )
        add_check(
            checks,
            failures,
            name,
            ok,
            {"path": str(path), "rows": len(actual), "expected_rows": len(expected), "error": error},
        )


def main() -> int:
    """Run all strict workbook checks and write a machine-readable report."""
    args = parse_args()
    checks: dict[str, Any] = {}
    failures: list[str] = []
    add_check(
        checks,
        failures,
        "workbook_file",
        args.workbook.is_file() and args.workbook.stat().st_size >= 10_000_000,
        args.workbook.stat().st_size if args.workbook.exists() else 0,
    )
    if not failures:
        verify_formula_workbook(args.workbook, checks, failures)
        verify_cached_values(args.workbook, checks, failures)
    verify_recalc_report(args.recalc_json, checks, failures)
    verify_csv_deliverables(
        args.workbook,
        args.executive_csv,
        args.reconciliation_csv,
        checks,
        failures,
    )
    report = {
        "status": "success" if not failures else "failed",
        "checks": checks,
        "failures": failures,
    }
    args.report_json.write_text(
        json.dumps(report, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0 if not failures else 1


if __name__ == "__main__":
    sys.exit(main())
