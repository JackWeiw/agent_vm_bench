#!/usr/bin/env python3
"""Derive the XLSX execution summary from the produced files."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("verification", type=Path)
    parser.add_argument("recalculation", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("executive_csv", type=Path)
    parser.add_argument("reconciliation_csv", type=Path)
    return parser.parse_args()


def read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"expected JSON object: {path}")
    return value


def chart_name(chart: Any) -> str:
    return type(chart).__name__


def csv_metadata(path: Path) -> dict[str, Any]:
    with path.open("rb") as handle:
        digest = hashlib.sha256(handle.read()).hexdigest()
    with path.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))
    return {"path": str(path), "rows": max(0, len(rows) - 1), "sha256": digest}


def collect_features(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        executive = workbook["Executive_Summary"]
        formula_cells = ("B5", "B6", "B7", "B8", "B9", "B13", "B14", "B15", "B16")
        formulas = {
            cell: executive[cell].value
            for cell in formula_cells
            if isinstance(executive[cell].value, str) and executive[cell].value.startswith("=")
        }
        comments = [
            f"Executive_Summary!{cell.coordinate}"
            for row in executive.iter_rows()
            for cell in row
            if cell.comment is not None
        ]
        charts = {
            sheet.title: [chart_name(chart) for chart in sheet._charts]
            for sheet in workbook.worksheets
            if sheet._charts
        }
        raw = workbook["Raw_Sample"]
        return {
            "sheet_order": workbook.sheetnames,
            "executive_summary_sheet": "Executive_Summary" in workbook.sheetnames,
            "formulas": formulas,
            "scenario_default": executive["B12"].value,
            "scenario_dropdown_ranges": [
                str(validation.sqref) for validation in executive.data_validations.dataValidation
            ],
            "charts": charts,
            "chart_count": sum(len(items) for items in charts.values()),
            "conditional_formatting_rules": len(executive.conditional_formatting),
            "freeze_panes": str(executive.freeze_panes) if executive.freeze_panes else None,
            "comment_cells": comments,
            "raw_sample_dimensions": {"rows": raw.max_row, "columns": raw.max_column},
            "external_link_count": len(workbook._external_links),
        }
    finally:
        workbook.close()


def atomic_write_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(value, handle, indent=2, ensure_ascii=False, sort_keys=True)
            handle.write("\n")
        os.replace(temporary_name, path)
    finally:
        Path(temporary_name).unlink(missing_ok=True)


def main() -> int:
    args = parse_args()
    verification = read_json(args.verification)
    recalculation = read_json(args.recalculation)
    summary = {
        "schema_version": "xlsx-execution-summary-v2",
        "workbook": args.workbook.name,
        "features": collect_features(args.workbook),
        "verifier": {
            "status": verification.get("status"),
            "failures": verification.get("failures", []),
        },
        "recalculation": {
            "status": recalculation.get("status"),
            "total_errors": recalculation.get("total_errors"),
            "total_formulas": recalculation.get("total_formulas"),
            "error": recalculation.get("error"),
        },
        "files": {
            "workbook": str(args.workbook),
            "verification": str(args.verification),
            "recalculation": str(args.recalculation),
            "executive_csv": csv_metadata(args.executive_csv),
            "reconciliation_csv": csv_metadata(args.reconciliation_csv),
        },
    }
    atomic_write_json(args.output, summary)
    print(json.dumps(summary, indent=2, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
