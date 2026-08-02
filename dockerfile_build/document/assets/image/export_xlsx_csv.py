#!/usr/bin/env python3
"""Export verified CSV deliverables from a recalculated XLSX workbook."""

from __future__ import annotations

import argparse
import csv
import os
import tempfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


EXECUTIVE_ROWS = (
    ("Total Trips", "B5"),
    ("Total Fare Revenue", "B6"),
    ("Overall Average Fare", "B7"),
    ("Removal Rate", "B8"),
    ("Reconciliation Status", "B9"),
    ("Selected Scenario", "B12"),
    ("Selected Fare Multiplier", "B13"),
    ("Selected Trip Multiplier", "B14"),
    ("Projected Trips", "B15"),
    ("Projected Fare Revenue", "B16"),
)
RECONCILIATION_ROWS = tuple(range(2, 7)) + tuple(range(10, 21)) + (24, 25)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("executive_csv", type=Path)
    parser.add_argument("reconciliation_csv", type=Path)
    return parser.parse_args()


def atomic_write_csv(path: Path, header: tuple[str, str], rows: Iterable[tuple[object, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(header)
            writer.writerows(rows)
        os.replace(temporary_name, path)
    finally:
        Path(temporary_name).unlink(missing_ok=True)


def main() -> int:
    args = parse_args()
    workbook = load_workbook(args.workbook, data_only=True, read_only=True)
    try:
        executive = workbook["Executive_Summary"]
        reconciliation = workbook["Reconciliation"]
        executive_rows = [(label, executive[cell].value) for label, cell in EXECUTIVE_ROWS]
        reconciliation_rows = [
            (reconciliation.cell(row=row, column=1).value, reconciliation.cell(row=row, column=2).value)
            for row in RECONCILIATION_ROWS
        ]
        missing = [label for label, value in executive_rows + reconciliation_rows if value is None]
        if missing:
            raise ValueError(f"cannot export missing cached values: {missing}")
    finally:
        workbook.close()

    atomic_write_csv(args.executive_csv, ("Metric", "Value"), executive_rows)
    atomic_write_csv(args.reconciliation_csv, ("Metric", "Value"), reconciliation_rows)
    print(f"exported {len(executive_rows)} executive rows to {args.executive_csv}")
    print(f"exported {len(reconciliation_rows)} reconciliation rows to {args.reconciliation_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
