import argparse
import json
import math
import os
import sys
import time

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(line_buffering=True)

parser = argparse.ArgumentParser(description="Build the stable TC4 XLSX template")
parser.add_argument("--input-dir", required=True)
parser.add_argument("--output-dir", required=True)
args = parser.parse_args()

OUT = args.output_dir
INPUT = args.input_dir
os.makedirs(OUT, exist_ok=True)

t0 = time.time()
print("Reading parquet...", flush=True)
raw = pd.read_parquet(f"{INPUT}/yellow_tripdata_2024-01.parquet")
total_input = len(raw)
print(f"Shape: {raw.shape} ({total_input:,} rows)", flush=True)

print("Reading zone lookup...", flush=True)
zones = pd.read_csv(f"{INPUT}/taxi_zone_lookup.csv")
zone_map = zones.set_index("LocationID")[["Borough", "Zone", "service_zone"]].to_dict(
    "index"
)

print("Cleaning data...", flush=True)
df = raw.copy()
df["tpep_pickup_datetime"] = pd.to_datetime(df["tpep_pickup_datetime"], errors="coerce")
df["tpep_dropoff_datetime"] = pd.to_datetime(
    df["tpep_dropoff_datetime"], errors="coerce"
)

mask = df["tpep_pickup_datetime"].notna() & df["tpep_dropoff_datetime"].notna()
mask &= df["tpep_pickup_datetime"] >= pd.Timestamp("2024-01-01")
mask &= df["tpep_pickup_datetime"] < pd.Timestamp("2024-02-01")
mask &= df["fare_amount"] >= 0
mask &= df["total_amount"] >= 0
mask &= df["tip_amount"] >= 0
mask &= df["tolls_amount"] >= 0
mask &= df["trip_distance"].between(0, 100)
mask &= df["passenger_count"].fillna(1).between(0, 6)
mask &= df["payment_type"].isin([1, 2, 3, 4])
mask &= df["VendorID"].isin([1, 2])
mask &= df["tpep_dropoff_datetime"] >= df["tpep_pickup_datetime"]
mask &= df["RatecodeID"].fillna(1).between(1, 6)
mask &= df["PULocationID"].isin(zone_map)
mask &= df["DOLocationID"].isin(zone_map)

df = df[mask].copy()
n1 = len(df)
print(f"After initial filter: {n1:,} ({total_input - n1} removed)", flush=True)

df["pickup_date"] = df["tpep_pickup_datetime"].dt.date
df["pickup_hour"] = df["tpep_pickup_datetime"].dt.hour
df["pickup_weekday"] = df["tpep_pickup_datetime"].dt.dayofweek
df["trip_duration_min"] = (
    df["tpep_dropoff_datetime"] - df["tpep_pickup_datetime"]
).dt.total_seconds() / 60.0
df = df[df["trip_duration_min"].between(0, 360)].copy()
n2 = len(df)
print(f"After duration filter: {n2:,} ({n1 - n2} removed)", flush=True)


def _db(m):
    if m <= 10:
        return "0-10 min"
    elif m <= 20:
        return "10-20 min"
    elif m <= 30:
        return "20-30 min"
    elif m <= 45:
        return "30-45 min"
    elif m <= 60:
        return "45-60 min"
    else:
        return "60+ min"


df["duration_bin"] = df["trip_duration_min"].map(_db)

borough_dict = zones.set_index("LocationID")["Borough"].to_dict()
zone_dict = zones.set_index("LocationID")["Zone"].to_dict()
svc_dict = zones.set_index("LocationID")["service_zone"].to_dict()
df["PU_Borough"] = df["PULocationID"].map(borough_dict).fillna("Unknown")
df["PU_Zone"] = df["PULocationID"].map(zone_dict).fillna("Unknown")
df["PU_ServiceZone"] = df["PULocationID"].map(svc_dict).fillna("Unknown")
df["DO_Borough"] = df["DOLocationID"].map(borough_dict).fillna("Unknown")
df["DO_Zone"] = df["DOLocationID"].map(zone_dict).fillna("Unknown")
df["DO_ServiceZone"] = df["DOLocationID"].map(svc_dict).fillna("Unknown")
pay_map = {1: "Credit Card", 2: "Cash", 3: "No Charge", 4: "Dispute", 5: "Unknown"}
df["payment_label"] = df["payment_type"].map(pay_map)

print("Aggregating hourly...", flush=True)
hourly = (
    df.groupby("pickup_hour")
    .agg(
        trips=("VendorID", "count"),
        tf=("fare_amount", "sum"),
        tt=("tip_amount", "sum"),
        ttl=("tolls_amount", "sum"),
        ta=("total_amount", "sum"),
        af=("fare_amount", "mean"),
        at=("tip_amount", "mean"),
        ad=("trip_distance", "mean"),
        adur=("trip_duration_min", "mean"),
    )
    .reset_index()
)
hourly.columns = [
    "Hour",
    "Trips",
    "Total_Fare",
    "Total_Tip",
    "Total_Tolls",
    "Total_Amount",
    "Avg_Fare",
    "Avg_Tip",
    "Avg_Distance",
    "Avg_Duration",
]
hourly = hourly.sort_values("Hour")

print("Aggregating daily...", flush=True)
day_map = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}
daily = (
    df.groupby("pickup_date")
    .agg(
        trips=("VendorID", "count"),
        tf=("fare_amount", "sum"),
        tt=("tip_amount", "sum"),
        ttl=("tolls_amount", "sum"),
        te=("extra", "sum"),
        tm=("mta_tax", "sum"),
        tc=("congestion_surcharge", "sum"),
        ta=("total_amount", "sum"),
        af=("fare_amount", "mean"),
        at=("tip_amount", "mean"),
        ad=("trip_distance", "mean"),
    )
    .reset_index()
)
daily.columns = [
    "Date",
    "Trips",
    "Total_Fare",
    "Total_Tip",
    "Total_Tolls",
    "Total_Extra",
    "Total_MTA_Tax",
    "Total_Congestion_Surcharge",
    "Total_Amount",
    "Avg_Fare",
    "Avg_Tip",
    "Avg_Distance",
]
daily = daily.sort_values("Date")
daily["Weekday"] = daily["Date"].apply(lambda d: day_map[d.weekday()])
daily = daily[
    [
        "Date",
        "Weekday",
        "Trips",
        "Total_Fare",
        "Total_Tip",
        "Total_Tolls",
        "Total_Extra",
        "Total_MTA_Tax",
        "Total_Congestion_Surcharge",
        "Total_Amount",
        "Avg_Fare",
        "Avg_Tip",
        "Avg_Distance",
    ]
]

print("Aggregating zone...", flush=True)
zone_sum = (
    df.groupby("PU_Borough")
    .agg(
        trips=("VendorID", "count"),
        tf=("fare_amount", "sum"),
        tt=("tip_amount", "sum"),
        ta=("total_amount", "sum"),
        af=("fare_amount", "mean"),
        at=("tip_amount", "mean"),
        ad=("trip_distance", "mean"),
    )
    .reset_index()
)
zone_sum.columns = [
    "Borough",
    "Trips",
    "Total_Fare",
    "Total_Tip",
    "Total_Amount",
    "Avg_Fare",
    "Avg_Tip",
    "Avg_Distance",
]
zone_sum = zone_sum.sort_values("Trips", ascending=False)

print("Aggregating payment...", flush=True)
pay_sum = (
    df.groupby("payment_label")
    .agg(
        trips=("VendorID", "count"),
        tf=("fare_amount", "sum"),
        tt=("tip_amount", "sum"),
        ta=("total_amount", "sum"),
        af=("fare_amount", "mean"),
        at=("tip_amount", "mean"),
    )
    .reset_index()
)
pay_sum.columns = [
    "Payment_Type",
    "Trips",
    "Total_Fare",
    "Total_Tip",
    "Total_Amount",
    "Avg_Fare",
    "Avg_Tip",
]
pay_sum = pay_sum.sort_values("Trips", ascending=False)

print("Binning fare/distance...", flush=True)
df["fare_bin"] = pd.cut(
    df["fare_amount"],
    bins=[0, 5, 10, 15, 20, 30, 50, 100, 5000],
    labels=[
        "$0-5",
        "$5-10",
        "$10-15",
        "$15-20",
        "$20-30",
        "$30-50",
        "$50-100",
        "$100+",
    ],
    right=False,
)
df["dist_bin"] = pd.cut(
    df["trip_distance"],
    bins=[0, 0.5, 1, 2, 3, 5, 10, 30, 100],
    labels=[
        "0-0.5mi",
        "0.5-1mi",
        "1-2mi",
        "2-3mi",
        "3-5mi",
        "5-10mi",
        "10-30mi",
        "30-100mi",
    ],
    right=False,
)
band = (
    df.groupby(["fare_bin", "dist_bin"], observed=False)
    .agg(trips=("VendorID", "count"), ta=("total_amount", "sum"))
    .reset_index()
)
band.columns = ["Fare_Range", "Distance_Range", "Trips", "Total_Amount"]
band["Pct_Trips"] = band["Trips"] / band["Trips"].sum() * 100
band = band.sort_values(["Fare_Range", "Distance_Range"])

print("Sampling 100K rows...", flush=True)
sample = df.sample(n=min(100000, len(df)), random_state=42).copy()
sample = sample.sort_values("tpep_pickup_datetime").reset_index(drop=True)
sample_cols = [
    "VendorID",
    "tpep_pickup_datetime",
    "tpep_dropoff_datetime",
    "passenger_count",
    "trip_distance",
    "RatecodeID",
    "PULocationID",
    "PU_Borough",
    "PU_Zone",
    "DOLocationID",
    "DO_Borough",
    "DO_Zone",
    "payment_type",
    "payment_label",
    "fare_amount",
    "extra",
    "mta_tax",
    "tip_amount",
    "tolls_amount",
    "improvement_surcharge",
    "congestion_surcharge",
    "Airport_fee",
    "total_amount",
    "trip_duration_min",
    "duration_bin",
]
sample = sample[sample_cols]

# ===== Build Excel =====
print("Building workbook...", flush=True)
wb = Workbook()
hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
hdr_fill = PatternFill("solid", fgColor="2F5496")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(size=10, name="Arial")
num_font = Font(size=10, name="Arial")
pct_fmt = "0.0%"
usd_fmt = "$#,##0.00"
int_fmt = "#,##0"
int0_fmt = "#,##0"
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
blue_fill = PatternFill("solid", fgColor="D6E4F0")


def wh(ws, row, hdrs):
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border


def wr(ws, row, vals, fmts=None):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = data_font
        c.border = thin_border
        if fmts and ci <= len(fmts) and fmts[ci - 1]:
            c.number_format = fmts[ci - 1]


def aw(ws):
    for col in ws.columns:
        mx = 0
        cl = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[cl].width = min(mx + 3, 40)


# Sheet 1: Raw_Sample
ws1 = wb.active
ws1.title = "Raw_Sample"
sh = [
    "VendorID",
    "Pickup_DateTime",
    "Dropoff_DateTime",
    "Passengers",
    "Trip_Distance",
    "RatecodeID",
    "PU_LocationID",
    "PU_Borough",
    "PU_Zone",
    "DO_LocationID",
    "DO_Borough",
    "DO_Zone",
    "Payment_Type",
    "Payment_Label",
    "Fare_Amount",
    "Extra",
    "MTA_Tax",
    "Tip_Amount",
    "Tolls_Amount",
    "Improvement_Surcharge",
    "Congestion_Surcharge",
    "Airport_Fee",
    "Total_Amount",
    "Duration_Min",
    "Duration_Bin",
]
wh(ws1, 1, sh)
sf = (
    [int0_fmt]
    + [None] * 2
    + [
        int0_fmt,
        "0.00",
        int0_fmt,
        int0_fmt,
        None,
        None,
        int0_fmt,
        None,
        None,
        int0_fmt,
        None,
        usd_fmt,
        usd_fmt,
        usd_fmt,
        usd_fmt,
        usd_fmt,
        usd_fmt,
        usd_fmt,
        usd_fmt,
        usd_fmt,
        "0.00",
        None,
    ]
)
for ri, row_data in enumerate(sample.itertuples(index=False), 2):
    vals = []
    for v in row_data:
        if isinstance(v, pd.Timestamp):
            vals.append(v.strftime("%Y-%m-%d %H:%M:%S"))
        elif isinstance(v, float) and math.isnan(v):
            vals.append(None)
        else:
            vals.append(v)
    wr(ws1, ri, vals, sf)
ws1.auto_filter.ref = f"A1:{get_column_letter(len(sh))}{len(sample) + 1}"
ws1.freeze_panes = "A2"
aw(ws1)

# Sheet 2: Hourly_Summary
ws2 = wb.create_sheet("Hourly_Summary")
h2 = [
    "Hour",
    "Trips",
    "Total_Fare",
    "Total_Tip",
    "Total_Tolls",
    "Total_Amount",
    "Avg_Fare",
    "Avg_Tip",
    "Avg_Distance",
    "Avg_Duration",
]
wh(ws2, 1, h2)
f2 = [
    int0_fmt,
    int0_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    "0.00",
    "0.0",
]
for ri, row_data in enumerate(hourly.itertuples(index=False), 2):
    wr(ws2, ri, row_data, f2)
kr = 2 + len(hourly) + 2
ws2.cell(row=kr, column=1, value="KPI").font = Font(bold=True, size=11, name="Arial")
ws2.cell(row=kr, column=1).fill = blue_fill
kps = [
    ("Total Trips", f"=SUM(B2:B{kr - 1})", int_fmt),
    ("Avg Hourly Trips", f"=AVERAGE(B2:B{kr - 1})", "0"),
    ("Peak Hour Trips", f"=MAX(B2:B{kr - 1})", int_fmt),
    ("Total Fare Revenue", f"=SUM(C2:C{kr - 1})", usd_fmt),
    ("Total Amount", f"=SUM(F2:F{kr - 1})", usd_fmt),
    ("Overall Avg Fare", f"=AVERAGE(G2:G{kr - 1})", usd_fmt),
    ("Overall Avg Tip", f"=AVERAGE(H2:H{kr - 1})", usd_fmt),
    ("Peak Avg Duration", f"=MAX(J2:J{kr - 1})", "0.0"),
]
for i, (lb, fm, nf) in enumerate(kps):
    r = kr + 1 + i
    ws2.cell(row=r, column=1, value=lb).font = Font(bold=True, size=10, name="Arial")
    ws2.cell(row=r, column=2, value=fm).number_format = nf
    ws2.cell(row=r, column=2).font = Font(size=10, name="Arial", color="0000FF")
ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{1 + len(hourly)}"
ws2.freeze_panes = "A2"
aw(ws2)

# Chart 1: Hourly Trip Bar
c1 = BarChart()
c1.type = "col"
c1.title = "Hourly Trip Distribution - Jan 2024"
c1.y_axis.title = "Trips"
c1.x_axis.title = "Hour of Day"
c1.style = 10
d1 = Reference(ws2, min_col=2, min_row=1, max_row=1 + len(hourly))
cat1 = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(hourly))
c1.add_data(d1, titles_from_data=True)
c1.set_categories(cat1)
c1.series[0].graphicalProperties.solidFill = "4472C4"
c1.legend = None
c1.width = 20
c1.height = 12
ws2.add_chart(c1, f"A{kr}")

# Sheet 3: Daily_Summary
ws3 = wb.create_sheet("Daily_Summary")
h3 = [
    "Date",
    "Weekday",
    "Trips",
    "Total_Fare",
    "Total_Tip",
    "Total_Tolls",
    "Total_Extra",
    "Total_MTA_Tax",
    "Total_Congestion_Surcharge",
    "Total_Amount",
    "Avg_Fare",
    "Avg_Tip",
    "Avg_Distance",
]
wh(ws3, 1, h3)
f3 = [
    None,
    None,
    int_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    usd_fmt,
    "0.00",
]
de = 1 + len(daily)
for ri, row_data in enumerate(daily.itertuples(index=False), 2):
    vals = list(row_data)
    if hasattr(vals[0], "strftime"):
        vals[0] = vals[0].strftime("%Y-%m-%d")
    wr(ws3, ri, vals, f3)
kr3 = de + 2
ws3.cell(row=kr3, column=1, value="KPI").font = Font(bold=True, size=11, name="Arial")
ws3.cell(row=kr3, column=1).fill = blue_fill
k3l = [
    ("Total Trips (Month)", f"=SUM(C2:C{de})", int_fmt),
    ("Avg Daily Trips", f"=AVERAGE(C2:C{de})", "0"),
    ("Max Daily Trips", f"=MAX(C2:C{de})", int_fmt),
    ("Total Fare Revenue", f"=SUM(D2:D{de})", usd_fmt),
    ("Total Amount (Month)", f"=SUM(J2:J{de})", usd_fmt),
    ("Avg Daily Fare Rev", f"=AVERAGE(D2:D{de})", usd_fmt),
    ("Avg Daily Tip", f"=AVERAGE(E2:E{de})", usd_fmt),
]
for i, (lb, fm, nf) in enumerate(k3l):
    r = kr3 + 1 + i
    ws3.cell(row=r, column=1, value=lb).font = Font(bold=True, size=10, name="Arial")
    ws3.cell(row=r, column=2, value=fm).number_format = nf
    ws3.cell(row=r, column=2).font = Font(size=10, name="Arial", color="0000FF")
ws3.auto_filter.ref = f"A1:{get_column_letter(len(h3))}{de}"
ws3.freeze_panes = "A2"
aw(ws3)

# Sheet 4: Zone_Summary
ws4 = wb.create_sheet("Zone_Summary")
h4 = [
    "Borough",
    "Trips",
    "Total_Fare",
    "Total_Tip",
    "Total_Amount",
    "Avg_Fare",
    "Avg_Tip",
    "Avg_Distance",
    "Pct_Trips",
]
wh(ws4, 1, h4)
ze = 1 + len(zone_sum)
f4 = [None, int_fmt, usd_fmt, usd_fmt, usd_fmt, usd_fmt, usd_fmt, "0.00", pct_fmt]
ztt = zone_sum["Trips"].sum()
for ri, row_data in enumerate(zone_sum.itertuples(index=False), 2):
    vals = list(row_data)
    vals.append(vals[1] / ztt)
    wr(ws4, ri, vals, f4)
ztr = ze + 1
ws4.cell(row=ztr, column=1, value="Total").font = Font(bold=True, size=10, name="Arial")
for ci, lm in [
    (2, f"=SUM(B2:B{ze})"),
    (3, f"=SUM(C2:C{ze})"),
    (4, f"=SUM(D2:D{ze})"),
    (5, f"=SUM(E2:E{ze})"),
]:
    ws4.cell(row=ztr, column=ci, value=lm).number_format = (
        usd_fmt if ci > 2 else int_fmt
    )
ws4.cell(row=ztr, column=9, value=1).number_format = pct_fmt
ws4.auto_filter.ref = f"A1:{get_column_letter(len(h4))}{ze}"
ws4.freeze_panes = "A2"
aw(ws4)

# Sheet 5: Payment_Summary
ws5 = wb.create_sheet("Payment_Summary")
h5 = [
    "Payment_Type",
    "Trips",
    "Total_Fare",
    "Total_Tip",
    "Total_Amount",
    "Avg_Fare",
    "Avg_Tip",
    "Pct_Trips",
]
wh(ws5, 1, h5)
pe = 1 + len(pay_sum)
f5 = [None, int_fmt, usd_fmt, usd_fmt, usd_fmt, usd_fmt, usd_fmt, pct_fmt]
ptt = pay_sum["Trips"].sum()
for ri, row_data in enumerate(pay_sum.itertuples(index=False), 2):
    vals = list(row_data)
    vals.append(vals[1] / ptt)
    wr(ws5, ri, vals, f5)
ptr = pe + 1
ws5.cell(row=ptr, column=1, value="Total").font = Font(bold=True, size=10, name="Arial")
for ci, lm in [
    (2, f"=SUM(B2:B{pe})"),
    (3, f"=SUM(C2:C{pe})"),
    (4, f"=SUM(D2:D{pe})"),
    (5, f"=SUM(E2:E{pe})"),
]:
    ws5.cell(row=ptr, column=ci, value=lm).number_format = (
        usd_fmt if ci > 2 else int_fmt
    )
ws5.cell(row=ptr, column=8, value=1).number_format = pct_fmt
ws5.auto_filter.ref = f"A1:{get_column_letter(len(h5))}{pe}"
ws5.freeze_panes = "A2"
aw(ws5)

# Chart 2: Payment Pie
c2 = PieChart()
c2.title = "Payment Type Distribution - Jan 2024"
c2.style = 10
d2 = Reference(ws5, min_col=2, min_row=1, max_row=1 + len(pay_sum))
cat2 = Reference(ws5, min_col=1, min_row=2, max_row=1 + len(pay_sum))
c2.add_data(d2, titles_from_data=True)
c2.set_categories(cat2)
c2.dataLabels = DataLabelList()
c2.dataLabels.showPercent = True
c2.dataLabels.showCatName = True
c2.width = 16
c2.height = 12
ws5.add_chart(c2, "A9")

# Sheet 6: Fare_Distance_Bands
ws6 = wb.create_sheet("Fare_Distance_Bands")
h6 = ["Fare_Range", "Distance_Range", "Trips", "Total_Amount", "Pct_Trips"]
wh(ws6, 1, h6)
be = 1 + len(band)
f6 = [None, None, int_fmt, usd_fmt, pct_fmt]
for ri, row_data in enumerate(band.itertuples(index=False), 2):
    vals = list(row_data)
    if isinstance(vals[0], float) and math.isnan(vals[0]):
        vals[0] = "Unknown"
    if isinstance(vals[1], float) and math.isnan(vals[1]):
        vals[1] = "Unknown"
    wr(ws6, ri, vals, f6)
ws6.auto_filter.ref = f"A1:{get_column_letter(len(h6))}{be}"
ws6.freeze_panes = "A2"
aw(ws6)

# Sheet 7: Reconciliation
ws7 = wb.create_sheet("Reconciliation")
ws7.cell(row=1, column=1, value="Data Quality Summary").font = Font(
    bold=True, size=13, name="Arial", color="2F5496"
)
dq = [
    ("Total Raw Records", total_input, int_fmt),
    ("Records After Cleaning", n2, int_fmt),
    ("Records Removed", total_input - n2, int_fmt),
    ("Removal Rate", (total_input - n2) / total_input, pct_fmt),
    ("Sample Size", len(sample), int_fmt),
]
for i, (lb, vl, nf) in enumerate(dq):
    r = 2 + i
    ws7.cell(row=r, column=1, value=lb).font = Font(bold=True, size=10, name="Arial")
    ws7.cell(row=r, column=2, value=vl).number_format = nf
sr = 2 + len(dq) + 2
ws7.cell(row=sr, column=1, value="Cross-Sheet Reconciliation (Formulas)").font = Font(
    bold=True, size=13, name="Arial", color="2F5496"
)
# Reference formula row numbers
hr_kpi = kr  # sheet2 kpi header row (e.g. row 28)
dr_kpi = kr3  # sheet3 kpi header row (e.g. row 34)
# KPI data rows: Total Trips at hr_kpi+1, Total Fare at hr_kpi+4 (Hourly)
#                Total Trips at dr_kpi+1, Total Fare at dr_kpi+4 (Daily)
ri1 = sr + 1
rits = [
    ("Hourly_Summary Total Trips", f"='Hourly_Summary'!B{hr_kpi + 1}", int_fmt),
    ("Daily_Summary Total Trips", f"='Daily_Summary'!B{dr_kpi + 1}", int_fmt),
    ("Zone_Summary Total Trips", f"='Zone_Summary'!B{ztr}", int_fmt),
    ("Payment_Summary Total Trips", f"='Payment_Summary'!B{ptr}", int_fmt),
    ("Hourly_Summary Total Fare", f"='Hourly_Summary'!B{hr_kpi + 4}", usd_fmt),
    ("Daily_Summary Total Fare", f"='Daily_Summary'!B{dr_kpi + 4}", usd_fmt),
    ("Zone_Summary Total Fare", f"='Zone_Summary'!C{ztr}", usd_fmt),
    ("Payment_Summary Total Fare", f"='Payment_Summary'!C{ptr}", usd_fmt),
    ("Trips Diff (Hourly - Daily)", f"=B{ri1}-B{ri1 + 1}", int_fmt),
    ("Trips Diff (Zone - Payment)", f"=B{ri1 + 2}-B{ri1 + 3}", int_fmt),
    ("Fare Diff (Hourly - Daily)", f"=B{ri1 + 4}-B{ri1 + 5}", usd_fmt),
]
for i, (lb, fm, nf) in enumerate(rits):
    r = ri1 + i
    ws7.cell(row=r, column=1, value=lb).font = Font(bold=True, size=10, name="Arial")
    ws7.cell(row=r, column=2, value=fm).number_format = nf
    ws7.cell(row=r, column=2).font = Font(size=10, name="Arial", color="006400")
sr2 = ri1 + len(rits) + 2
ws7.cell(row=sr2, column=1, value="Verification").font = Font(
    bold=True, size=13, name="Arial", color="2F5496"
)
vr1 = sr2 + 1
vits = [
    (
        "All Trip Totals Match",
        f'=IF(AND(B{ri1}=B{ri1 + 1},B{ri1 + 1}=B{ri1 + 2},B{ri1 + 2}=B{ri1 + 3}),"YES - Consistent","MISMATCH - Check Formulas")',
        None,
    ),
    (
        "All Fare Totals Match",
        f'=IF(AND(B{ri1 + 4}=B{ri1 + 5},B{ri1 + 5}=B{ri1 + 6},B{ri1 + 6}=B{ri1 + 7}),"YES - Consistent","MISMATCH - Check Formulas")',
        None,
    ),
]
for i, (lb, fm, nf) in enumerate(vits):
    r = vr1 + i
    ws7.cell(row=r, column=1, value=lb).font = Font(bold=True, size=10, name="Arial")
    ws7.cell(row=r, column=2, value=fm)
    ws7.cell(row=r, column=2).font = num_font
ws7.column_dimensions["A"].width = 50
ws7.column_dimensions["B"].width = 30

xlsx = f"{OUT}/monthly_operations_template.xlsx"
wb.save(xlsx)
print(f"Saved: {xlsx}", flush=True)

# CSV summaries
print("Writing CSVs...", flush=True)
pd.DataFrame(
    {
        "Metric": [
            "Total Raw Records",
            "After Cleaning",
            "Removed",
            "Removal Rate",
            "Total Trips",
            "Total Fare",
            "Total Tip",
            "Total Tolls",
            "Total Extra",
            "Total MTA Tax",
            "Total Congestion Surcharge",
            "Total Amount",
            "Avg Fare",
            "Avg Tip",
            "Avg Trip Distance",
            "Peak Hour",
            "Peak Hour Trips",
            "Credit Card %",
            "Cash %",
            "Manhattan Trips",
            "Brooklyn Trips",
            "Queens Trips",
            "Bronx Trips",
            "Staten Island Trips",
        ],
        "Value": [
            total_input,
            n2,
            total_input - n2,
            f"{(total_input - n2) / total_input * 100:.1f}%",
            f"{n2:,}",
            f"${df['fare_amount'].sum():,.2f}",
            f"${df['tip_amount'].sum():,.2f}",
            f"${df['tolls_amount'].sum():,.2f}",
            f"${df['extra'].sum():,.2f}",
            f"${df['mta_tax'].sum():,.2f}",
            f"${df['congestion_surcharge'].sum():,.2f}",
            f"${df['total_amount'].sum():,.2f}",
            f"${df['fare_amount'].mean():.2f}",
            f"${df['tip_amount'].mean():.2f}",
            f"{df['trip_distance'].mean():.2f} mi",
            f"{hourly.loc[hourly['Trips'].idxmax(), 'Hour']}:00",
            f"{hourly['Trips'].max():,}",
            f"{(pay_sum.loc[pay_sum['Payment_Type'] == 'Credit Card', 'Trips'].iloc[0] / n2 * 100):.1f}%",
            f"{(pay_sum.loc[pay_sum['Payment_Type'] == 'Cash', 'Trips'].iloc[0] / n2 * 100):.1f}%",
            f"{zone_sum.loc[zone_sum['Borough'] == 'Manhattan', 'Trips'].iloc[0]:,}",
            f"{zone_sum.loc[zone_sum['Borough'] == 'Brooklyn', 'Trips'].iloc[0] if 'Brooklyn' in zone_sum['Borough'].values else 0:,}",
            f"{zone_sum.loc[zone_sum['Borough'] == 'Queens', 'Trips'].iloc[0] if 'Queens' in zone_sum['Borough'].values else 0:,}",
            f"{zone_sum.loc[zone_sum['Borough'] == 'Bronx', 'Trips'].iloc[0] if 'Bronx' in zone_sum['Borough'].values else 0:,}",
            f"{zone_sum.loc[zone_sum['Borough'] == 'Staten Island', 'Trips'].iloc[0] if 'Staten Island' in zone_sum['Borough'].values else 0:,}",
        ],
    }
).to_csv(f"{OUT}/prepared_monthly_operations_summary.csv", index=False)

pd.DataFrame(
    {
        "Source": [
            "Raw Data (parquet)",
            "After Cleaning",
            "Hourly Sum",
            "Daily Sum",
            "Zone Sum",
            "Payment Sum",
        ],
        "Trips": [
            total_input,
            n2,
            hourly["Trips"].sum(),
            daily["Trips"].sum(),
            zone_sum["Trips"].sum(),
            pay_sum["Trips"].sum(),
        ],
        "Total_Fare": [
            raw["fare_amount"].sum(),
            df["fare_amount"].sum(),
            hourly["Total_Fare"].sum(),
            daily["Total_Fare"].sum(),
            zone_sum["Total_Fare"].sum(),
            pay_sum["Total_Fare"].sum(),
        ],
        "Total_Amount": [
            raw["total_amount"].sum(),
            df["total_amount"].sum(),
            hourly["Total_Amount"].sum(),
            daily["Total_Amount"].sum(),
            zone_sum["Total_Amount"].sum(),
            pay_sum["Total_Amount"].sum(),
        ],
    }
).to_csv(f"{OUT}/prepared_reconciliation_summary.csv", index=False)

manifest = {
    "status": "template_built",
    "source_rows": total_input,
    "clean_rows": n2,
    "sample_rows": len(sample),
    "expected_sheets": [
        "Raw_Sample",
        "Hourly_Summary",
        "Daily_Summary",
        "Zone_Summary",
        "Payment_Summary",
        "Fare_Distance_Bands",
        "Reconciliation",
    ],
    "reconciliation": {
        "trip_status_cell": f"B{vr1}",
        "fare_status_cell": f"B{vr1 + 1}",
        "success_value": "YES - Consistent",
        "trip_total_cells": [f"B{ri1 + offset}" for offset in range(4)],
        "fare_total_cells": [f"B{ri1 + offset}" for offset in range(4, 8)],
    },
    "expected_clean_total_fare": round(float(df["fare_amount"].sum()), 2),
    "expected_clean_total_amount": round(float(df["total_amount"].sum()), 2),
    "minimum_formula_count": 1,
}
with open(f"{OUT}/template_manifest.json", "w", encoding="utf-8") as handle:
    json.dump(manifest, handle, ensure_ascii=False, indent=2)
    handle.write("\n")

t1 = time.time()
print(f"Done in {t1 - t0:.1f}s", flush=True)
print(f"Trips: raw={total_input:,} clean={n2:,}", flush=True)
print(
    f"Hourly={hourly['Trips'].sum():,} Daily={daily['Trips'].sum():,} Zone={zone_sum['Trips'].sum():,} Pay={pay_sum['Trips'].sum():,}",
    flush=True,
)
ams = (
    hourly["Trips"].sum()
    == daily["Trips"].sum()
    == zone_sum["Trips"].sum()
    == pay_sum["Trips"].sum()
)
print(f"All match: {ams}", flush=True)
