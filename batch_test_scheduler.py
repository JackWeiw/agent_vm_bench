#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Multi-VM Agent Batch Test Scheduler

Orchestrates multiple test runs with different parameter combinations:
1. Define test parameter matrix (VM count, ratio, active percent)
2. Generate test task list
3. For each task: generate temp config, call auto_vm_test.py, collect results
4. Generate summary report comparing all test results

Usage:
    python batch_test_scheduler.py --config batch_config.yaml
"""

import os
import sys
import time
import subprocess
import argparse
import yaml
import json
import shutil
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Tuple
from dataclasses import dataclass


@dataclass
class TestTask:
    """Single test task definition"""
    task_id: str
    vm_count: int
    ratio: float
    active_percent: float
    config_file: str = ""
    result_dir: str = ""
    success: bool = False
    error_msg: str = ""


def load_batch_config(config_path: str) -> Dict:
    """Load batch config YAML file"""
    with open(config_path) as f:
        return yaml.safe_load(f)


def generate_task_id(vm_count: int, ratio: float, active_percent: float) -> str:
    """Generate unique task ID from parameters"""
    return f"vm{vm_count}_ratio{ratio}_active{active_percent}"


def generate_test_tasks(batch_config: Dict) -> List[TestTask]:
    """Generate all test tasks from parameter matrix"""
    matrix = batch_config["test_matrix"]
    fixed = batch_config["fixed_params"]

    vm_counts = matrix["vm_counts"]
    ratios = matrix["ratios"]
    active_percentages = matrix["active_percentages"]

    tasks = []

    for vm_count in vm_counts:
        for ratio in ratios:
            for active_percent in active_percentages:
                task = TestTask(
                    task_id=generate_task_id(vm_count, ratio, active_percent),
                    vm_count=vm_count,
                    ratio=ratio,
                    active_percent=active_percent
                )
                tasks.append(task)

    return tasks


def generate_temp_config(template_path: str, task: TestTask, fixed_params: Dict, base_dir: str, output_dir: str) -> str:
    """Generate temporary config file for a test task"""
    # Load template
    with open(template_path) as f:
        template_content = f.read()

    # Replace dynamic parameters
    replacements = {
        "{{VM_COUNT}}": str(task.vm_count),
        "{{START_IP}}": fixed_params["start_ip"],
        "{{SWAP_SIZE_GB}}": str(fixed_params["swap_size_gb"]),
        "{{RATIO}}": str(task.ratio),
        "{{ACTIVE_PERCENT}}": str(task.active_percent),
        "{{DURATION}}": str(fixed_params["duration"]),
        "{{BASE_DIR}}": base_dir
    }

    for key, value in replacements.items():
        template_content = template_content.replace(key, value)

    # Write to temp file
    os.makedirs(output_dir, exist_ok=True)
    temp_config_path = os.path.join(output_dir, f"config_{task.task_id}.yaml")

    with open(temp_config_path, "w") as f:
        f.write(template_content)

    return temp_config_path


def run_single_test(task: TestTask, config_file: str, batch_log_file: str) -> Tuple[bool, str]:
    """Run single test via auto_vm_test.py"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(batch_log_file, "a") as f:
        f.write(f"\n[{timestamp}] Starting test: {task.task_id}\n")
        f.write(f"  VM count: {task.vm_count}\n")
        f.write(f"  Ratio: {task.ratio}\n")
        f.write(f"  Active percent: {task.active_percent}\n")
        f.write(f"  Config file: {config_file}\n")

    print(f"\n{'=' * 60}")
    print(f"Running test: {task.task_id}")
    print(f"{'=' * 60}")

    # Execute auto_vm_test.py
    cmd = ["python", "auto_vm_test.py", "--config", config_file]

    print(f"[DEBUG] Command: {' '.join(cmd)}")
    print(f"[DEBUG] Working dir: {os.getcwd()}")

    with open(batch_log_file, "a") as f:
        f.write(f"[DEBUG] Command: {' '.join(cmd)}\n")
        f.write(f"[DEBUG] Working dir: {os.getcwd()}\n")

    result = subprocess.run(
        cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True
    )

    # Print and log stdout
    print(f"[DEBUG] Return code: {result.returncode}")
    print(f"[DEBUG] STDOUT ({len(result.stdout)} chars):")
    print(result.stdout)

    with open(batch_log_file, "a") as f:
        f.write(f"[DEBUG] Return code: {result.returncode}\n")
        f.write(f"[DEBUG] STDOUT:\n{result.stdout}\n")

    # Print and log stderr
    if result.stderr:
        print(f"[DEBUG] STDERR:")
        print(result.stderr)
        with open(batch_log_file, "a") as f:
            f.write(f"[DEBUG] STDERR:\n{result.stderr}\n")

    # Parse result directory from output
    result_dir = ""
    print(f"[DEBUG] Parsing result directory from output...")
    for line in result.stdout.split("\n"):
        print(f"[DEBUG] Line: '{line}'")
        if "Result directory:" in line or "Result dir:" in line or "Test result saved to:" in line:
            # Try different parsing patterns
            if "Result directory:" in line or "Result dir:" in line:
                parts = line.split(":")
                if len(parts) >= 2:
                    result_dir = parts[-1].strip()
            elif "Test result saved to:" in line:
                parts = line.split("saved to:")
                if len(parts) >= 2:
                    result_dir = parts[-1].strip()
            print(f"[DEBUG] Found result_dir: {result_dir}")
            break

    success = result.returncode == 0

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(batch_log_file, "a") as f:
        f.write(f"\n[{timestamp}] Test completed: {task.task_id}\n")
        f.write(f"  Success: {success}\n")
        f.write(f"  Result dir: {result_dir}\n")

    print(f"\nTest {task.task_id}: {'SUCCESS' if success else 'FAILED'}")
    print(f"Result: {result_dir}")

    return success, result_dir


def move_config_to_result(config_file: str, result_dir: str, task_id: str):
    """Move temp config file to result directory with task_id suffix"""
    if not result_dir or not os.path.exists(result_dir):
        return

    # Use task_id in filename to avoid collision with auto_vm_test.py's config.yaml
    dest = os.path.join(result_dir, f"config_{task_id}.yaml")
    shutil.move(config_file, dest)


def extract_browser_metrics_from_report(result_dir: str) -> Dict:
    """Extract browser metrics from vm_bench_lite bench_report file"""
    vm_bench_dir = os.path.join(result_dir, "vm_bench_lite")

    if not os.path.exists(vm_bench_dir):
        return {}

    # Find bench_report file
    bench_reports = list(Path(vm_bench_dir).glob("bench_report_*.txt"))
    if not bench_reports:
        return {}

    # Read the most recent report
    report_path = str(sorted(bench_reports)[-1])
    with open(report_path, encoding="utf-8") as f:
        content = f.read()

    # Parse metrics using regex
    metrics = {}

    # Success Rate: X.X%
    match = re.search(r"Success Rate:\s+([\d.]+)%", content)
    if match:
        metrics["success_rate"] = float(match.group(1))

    # Avg Latency: X.Xms
    match = re.search(r"Avg Latency:\s+([\d.]+)ms", content)
    if match:
        metrics["avg_latency"] = float(match.group(1))

    # P99 Latency: X.Xms
    match = re.search(r"P99 Latency:\s+([\d.]+)ms", content)
    if match:
        metrics["p99_latency"] = float(match.group(1))

    # Total Tasks
    match = re.search(r"Total Tasks:\s+(\d+)", content)
    if match:
        metrics["total_tasks"] = int(match.group(1))

    return metrics


def extract_qemu_metrics_from_excel(result_dir: str) -> Dict:
    """Extract QEMU and performance metrics from vm_monitor analysis_report.xlsx

    Extracts ALL metrics from:
    - Summary sheet (VM CPU stats)
    - DevKit_TopDown sheet (13 metrics)
    - DevKit_Memory sheet (6 metrics + NUMA bandwidth)
    - KSys sheet (latency, IPC, topdown)
    - UBWatch_Latency sheet (8 metrics)
    """
    vm_dir = os.path.join(result_dir, "vm_monitor")
    # Backward compatibility: also check qemu_monitor directory
    if not os.path.exists(vm_dir):
        vm_dir = os.path.join(result_dir, "qemu_monitor")
    excel_path = os.path.join(vm_dir, "analysis_report.xlsx")

    if not os.path.exists(excel_path):
        return {}

    metrics = {}

    try:
        import pandas as pd

        # ========== Summary sheet ==========
        try:
            df_summary = pd.read_excel(excel_path, sheet_name="Summary")
            for idx, row in df_summary.iterrows():
                metric = str(row["Metric"]).strip() if pd.notna(row["Metric"]) else ""
                value = row["Value"]
                if metric == "VM Avg CPU":
                    metrics["avg_cpu_percent"] = float(value) if pd.notna(value) else 0
                elif metric == "VM Peak Total CPU":
                    metrics["max_cpu_percent"] = float(value) if pd.notna(value) else 0
        except Exception:
            pass

        # ========== DevKit_TopDown sheet (13 metrics) ==========
        try:
            df_topdown = pd.read_excel(excel_path, sheet_name="DevKit_TopDown")
            for idx, row in df_topdown.iterrows():
                metric = str(row["Metric"]).strip() if pd.notna(row["Metric"]) else ""
                value = row["Value"]
                # Map metric names to keys
                key_map = {
                    "Cycles Avg": "td_cycles_avg",
                    "Instructions Avg": "td_instructions_avg",
                    "IPC Avg": "td_ipc_avg",
                    "IPC Max": "td_ipc_max",
                    "IPC Min": "td_ipc_min",
                    "Bad Speculation (%)": "td_bad_speculation",
                    "Frontend Bound (%)": "td_frontend_bound",
                    "Retiring (%)": "td_retiring",
                    "Backend Bound (%)": "td_backend_bound",
                    "L3 Bound (%)": "td_l3_bound",
                    "Mem Bound (%)": "td_mem_bound",
                    "Latency Bound (%)": "td_latency_bound",
                    "Bandwidth Bound (%)": "td_bandwidth_bound",
                }
                if metric in key_map:
                    metrics[key_map[metric]] = float(value) if pd.notna(value) else 0
        except Exception:
            pass

        # ========== DevKit_Memory sheet (6 metrics + L3 hit rate) ==========
        try:
            df_mem = pd.read_excel(excel_path, sheet_name="DevKit_Memory")
            for idx, row in df_mem.iterrows():
                metric = str(row["Metric"]).strip() if pd.notna(row["Metric"]) else ""
                value = row["Value"]
                key_map = {
                    "L1D Miss (%)": "mem_l1d_miss",
                    "L1I Miss (%)": "mem_l1i_miss",
                    "L2D Miss (%)": "mem_l2d_miss",
                    "L2I Miss (%)": "mem_l2i_miss",
                    "DDR Read (MB/s)": "mem_ddr_read",
                    "DDR Write (MB/s)": "mem_ddr_write",
                }
                if metric in key_map:
                    metrics[key_map[metric]] = float(value) if pd.notna(value) else 0
                # Extract L3 hit rate: NUMA0 L3 Hit Rate (%), NUMA1 L3 Hit Rate (%), etc.
                elif "L3 Hit Rate" in metric:
                    # Parse "NUMA0 L3 Hit Rate (%)" -> numa0_l3_hit_rate
                    numa_match = re.match(r"NUMA(\d+)\s+L3 Hit Rate", metric)
                    if numa_match:
                        node_id = numa_match.group(1)
                        metrics[f"numa{node_id}_l3_hit_rate"] = float(value) if pd.notna(value) else 0
        except Exception:
            pass

        # ========== NUMA_Bandwidth sheet ==========
        try:
            df_numa = pd.read_excel(excel_path, sheet_name="NUMA_Bandwidth")
            # Store per-node bandwidth, also compute total
            total_read = 0.0
            total_write = 0.0
            for idx, row in df_numa.iterrows():
                node = str(row["NUMA Node"]).strip() if pd.notna(row["NUMA Node"]) else ""
                read = float(row["Read (MB/s)"]) if pd.notna(row["Read (MB/s)"]) else 0
                write = float(row["Write (MB/s)"]) if pd.notna(row["Write (MB/s)"]) else 0
                if node:
                    metrics[f"numa{node}_read"] = read
                    metrics[f"numa{node}_write"] = write
                    total_read += read
                    total_write += write
            metrics["numa_total_read"] = total_read
            metrics["numa_total_write"] = total_write
        except Exception:
            pass

        # ========== KSys sheet ==========
        try:
            df_ksys = pd.read_excel(excel_path, sheet_name="KSys")
            for idx, row in df_ksys.iterrows():
                metric = str(row["Metric"]).strip() if pd.notna(row["Metric"]) else ""
                value = row["Value"]
                key_map = {
                    "L2 Miss Latency Max": "ksys_l2_miss_latency_max",
                    "L2 Miss Latency Min": "ksys_l2_miss_latency_min",
                    "L2 Miss Latency Avg": "ksys_l2_miss_latency_avg",
                    "L3 Miss Latency Max": "ksys_l3_miss_latency_max",
                    "L3 Miss Latency Min": "ksys_l3_miss_latency_min",
                    "L3 Miss Latency Avg": "ksys_l3_miss_latency_avg",
                    "IPC": "ksys_ipc",
                    "Retiring (%)": "ksys_retiring",
                    "Frontend Bound (%)": "ksys_frontend_bound",
                    "Bad Speculation (%)": "ksys_bad_speculation",
                    "Backend Bound (%)": "ksys_backend_bound",
                }
                if metric in key_map:
                    metrics[key_map[metric]] = float(value) if pd.notna(value) else 0
        except Exception:
            pass

        # ========== UBWatch_Latency sheet ==========
        try:
            df_ub = pd.read_excel(excel_path, sheet_name="UBWatch_Latency")
            for idx, row in df_ub.iterrows():
                metric = str(row["Metric"]).strip() if pd.notna(row["Metric"]) else ""
                value = row["Value"]
                key_map = {
                    "Samples": "ub_samples",
                    "Avg Read (ns)": "ub_avg_read_ns",
                    "Avg Write (ns)": "ub_avg_write_ns",
                    "Min Read (ns)": "ub_min_read_ns",
                    "Min Write (ns)": "ub_min_write_ns",
                    "Max Read (ns)": "ub_max_read_ns",
                    "Max Write (ns)": "ub_max_write_ns",
                }
                if metric in key_map:
                    # Handle numeric values, skip "Latency Path" which is a string
                    try:
                        metrics[key_map[metric]] = float(value) if pd.notna(value) else 0
                    except (ValueError, TypeError):
                        pass
        except Exception:
            pass

        # ========== UBWatch_Bandwidth sheet (per-chip per-port bandwidth) ==========
        try:
            df_bw = pd.read_excel(excel_path, sheet_name="UBWatch_Bandwidth")
            # Calculate totals across all chip+port combinations
            total_avg_wr = 0.0
            total_avg_rd = 0.0
            total_avg_sum = 0.0
            total_max_wr = 0.0
            total_max_rd = 0.0
            total_max_sum = 0.0

            for idx, row in df_bw.iterrows():
                chip = int(row["Chip"]) if pd.notna(row["Chip"]) else None
                ports = str(row["Ports"]).strip() if pd.notna(row["Ports"]) else ""
                if chip is not None and ports:
                    # Normalize ports name: "0&1" -> "p01", "8" -> "p8"
                    port_key = "p" + ports.replace("&", "")

                    avg_wr = float(row["Avg Write (MB/s)"]) if pd.notna(row["Avg Write (MB/s)"]) else 0
                    avg_rd = float(row["Avg Read (MB/s)"]) if pd.notna(row["Avg Read (MB/s)"]) else 0
                    avg_sum = float(row["Avg Sum (MB/s)"]) if pd.notna(row["Avg Sum (MB/s)"]) else 0
                    max_wr = float(row["Max Write (MB/s)"]) if pd.notna(row["Max Write (MB/s)"]) else 0
                    max_rd = float(row["Max Read (MB/s)"]) if pd.notna(row["Max Read (MB/s)"]) else 0
                    max_sum = float(row["Max Sum (MB/s)"]) if pd.notna(row["Max Sum (MB/s)"]) else 0

                    # Per-chip+port metrics (e.g., ub_bw_chip0_p01_avg_wr_mb_s)
                    key_prefix = f"ub_bw_chip{chip}_{port_key}"
                    metrics[f"{key_prefix}_avg_wr_mb_s"] = avg_wr
                    metrics[f"{key_prefix}_avg_rd_mb_s"] = avg_rd
                    metrics[f"{key_prefix}_avg_sum_mb_s"] = avg_sum
                    metrics[f"{key_prefix}_max_wr_mb_s"] = max_wr
                    metrics[f"{key_prefix}_max_rd_mb_s"] = max_rd
                    metrics[f"{key_prefix}_max_sum_mb_s"] = max_sum

                    # Accumulate totals
                    total_avg_wr += avg_wr
                    total_avg_rd += avg_rd
                    total_avg_sum += avg_sum
                    total_max_wr = max(total_max_wr, max_wr)
                    total_max_rd = max(total_max_rd, max_rd)
                    total_max_sum = max(total_max_sum, max_sum)

            # Total bandwidth metrics (sum of all chip+port combinations)
            metrics["ub_bw_total_avg_wr_mb_s"] = total_avg_wr
            metrics["ub_bw_total_avg_rd_mb_s"] = total_avg_rd
            metrics["ub_bw_total_avg_sum_mb_s"] = total_avg_sum
            metrics["ub_bw_total_max_wr_mb_s"] = total_max_wr
            metrics["ub_bw_total_max_rd_mb_s"] = total_max_rd
            metrics["ub_bw_total_max_sum_mb_s"] = total_max_sum
        except Exception:
            pass

        # ========== SMAPBW_Summary sheet (SMAP migration bandwidth) ==========
        try:
            df_smap = pd.read_excel(excel_path, sheet_name="SMAPBW_Summary")
            for idx, row in df_smap.iterrows():
                metric = str(row["Metric"]).strip() if pd.notna(row["Metric"]) else ""
                value = row["Value"]
                key_map = {
                    "Total Cycles": "smapbw_total_cycles",
                    "Total Pages": "smapbw_total_pages",
                    "Avg Bandwidth (GB/s)": "smapbw_avg_bandwidth_gb_s",
                    "Min Bandwidth (GB/s)": "smapbw_min_bandwidth_gb_s",
                    "Max Bandwidth (GB/s)": "smapbw_max_bandwidth_gb_s",
                }
                if metric in key_map:
                    metrics[key_map[metric]] = float(value) if pd.notna(value) else 0
        except Exception:
            pass

        # ========== SMAPBW_Cycles sheet (per-cycle bandwidth details) ==========
        try:
            df_smap_cycles = pd.read_excel(excel_path, sheet_name="SMAPBW_Cycles")
            # Calculate per-cycle statistics
            cycles_list = []
            for idx, row in df_smap_cycles.iterrows():
                cycle_no = int(row["Cycle"]) if pd.notna(row["Cycle"]) else 0
                bandwidth = float(row["Bandwidth (GB/s)"]) if pd.notna(row["Bandwidth (GB/s)"]) else 0
                cycles_list.append(bandwidth)
                # Store per-cycle bandwidth
                metrics[f"smapbw_cycle{cycle_no}_bandwidth_gb_s"] = bandwidth
            # Additional statistics from cycles if available
            if cycles_list:
                metrics["smapbw_cycle_count"] = len(cycles_list)
                if "smapbw_avg_bandwidth_gb_s" not in metrics or metrics["smapbw_avg_bandwidth_gb_s"] == 0:
                    metrics["smapbw_avg_bandwidth_gb_s"] = sum(cycles_list) / len(cycles_list)
                if "smapbw_min_bandwidth_gb_s" not in metrics or metrics["smapbw_min_bandwidth_gb_s"] == 0:
                    metrics["smapbw_min_bandwidth_gb_s"] = min(cycles_list)
                if "smapbw_max_bandwidth_gb_s" not in metrics or metrics["smapbw_max_bandwidth_gb_s"] == 0:
                    metrics["smapbw_max_bandwidth_gb_s"] = max(cycles_list)
        except Exception:
            pass

        # ========== Getfre_Summary sheet (Core Frequency Summary) ==========
        try:
            df_getfre_sum = pd.read_excel(excel_path, sheet_name="Getfre_Summary")
            for idx, row in df_getfre_sum.iterrows():
                numa = int(row["NUMA"]) if pd.notna(row["NUMA"]) else None
                if numa is not None:
                    avg_freq = float(row["Avg Frequency (MHz)"]) if pd.notna(row["Avg Frequency (MHz)"]) else 0
                    min_freq = float(row["Min Frequency (MHz)"]) if pd.notna(row["Min Frequency (MHz)"]) else 0
                    max_freq = float(row["Max Frequency (MHz)"]) if pd.notna(row["Max Frequency (MHz)"]) else 0
                    sample_count = int(row["Sample Count"]) if pd.notna(row["Sample Count"]) else 0
                    core_count = int(row["Core Count"]) if pd.notna(row["Core Count"]) else 0

                    # Per-NUMA getfre metrics
                    metrics[f"getfre_numa{numa}_avg_mhz"] = avg_freq
                    metrics[f"getfre_numa{numa}_min_mhz"] = min_freq
                    metrics[f"getfre_numa{numa}_max_mhz"] = max_freq
                    metrics[f"getfre_numa{numa}_sample_count"] = sample_count
                    metrics[f"getfre_numa{numa}_core_count"] = core_count
        except Exception:
            pass

        # ========== Getfre_NUMA sheets (Per-core frequency details) ==========
        # Check for Getfre_NUMA0, Getfre_NUMA1, etc.
        try:
            xl = pd.ExcelFile(excel_path)
            getfre_numa_sheets = [s for s in xl.sheet_names if s.startswith("Getfre_NUMA")]

            for sheet_name in getfre_numa_sheets:
                # Extract NUMA ID from sheet name (e.g., "Getfre_NUMA0" -> 0)
                numa_id = sheet_name.replace("Getfre_NUMA", "")

                df_cores = pd.read_excel(excel_path, sheet_name=sheet_name)
                # Calculate per-NUMA core frequency statistics
                freq_avg_list = []
                freq_min_list = []
                freq_max_list = []

                for idx, row in df_cores.iterrows():
                    core_id = int(row["Core ID"]) if pd.notna(row["Core ID"]) else None
                    if core_id is not None:
                        avg_freq = float(row["Avg Frequency (MHz)"]) if pd.notna(row["Avg Frequency (MHz)"]) else 0
                        min_freq = float(row["Min Frequency (MHz)"]) if pd.notna(row["Min Frequency (MHz)"]) else 0
                        max_freq = float(row["Max Frequency (MHz)"]) if pd.notna(row["Max Frequency (MHz)"]) else 0

                        freq_avg_list.append(avg_freq)
                        freq_min_list.append(min_freq)
                        freq_max_list.append(max_freq)

                        # Store top frequency variance cores (optional, for detailed analysis)
                        if max_freq - min_freq > 100:  # Only store if variance > 100 MHz
                            metrics[f"getfre_numa{numa_id}_core{core_id}_avg_mhz"] = avg_freq
                            metrics[f"getfre_numa{numa_id}_core{core_id}_variance_mhz"] = max_freq - min_freq

                # Calculate overall statistics for this NUMA from per-core data
                if freq_avg_list:
                    # Weighted average across cores (each core's avg)
                    overall_avg = sum(freq_avg_list) / len(freq_avg_list)
                    overall_min = min(freq_min_list)
                    overall_max = max(freq_max_list)
                    overall_variance = overall_max - overall_min

                    # Store as summary metrics (may override Getfre_Summary values)
                    metrics[f"getfre_numa{numa_id}_overall_avg_mhz"] = overall_avg
                    metrics[f"getfre_numa{numa_id}_overall_min_mhz"] = overall_min
                    metrics[f"getfre_numa{numa_id}_overall_max_mhz"] = overall_max
                    metrics[f"getfre_numa{numa_id}_overall_variance_mhz"] = overall_variance
                    metrics[f"getfre_numa{numa_id}_core_count"] = len(freq_avg_list)
        except Exception:
            pass

    except ImportError:
        # pandas not available, try openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)

            # Helper function to extract metrics from a sheet
            def extract_sheet_metrics(ws, key_map):
                result = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        metric = str(row[0]).strip()
                        if metric in key_map:
                            try:
                                result[key_map[metric]] = float(row[1]) if row[1] else 0
                            except (ValueError, TypeError):
                                pass
                return result

            # Summary sheet
            if "Summary" in wb.sheetnames:
                ws = wb["Summary"]
                key_map = {"VM Avg CPU": "avg_cpu_percent", "VM Peak Total CPU": "max_cpu_percent"}
                metrics.update(extract_sheet_metrics(ws, key_map))

            # DevKit_TopDown
            if "DevKit_TopDown" in wb.sheetnames:
                ws = wb["DevKit_TopDown"]
                key_map = {
                    "Cycles Avg": "td_cycles_avg",
                    "Instructions Avg": "td_instructions_avg",
                    "IPC Avg": "td_ipc_avg",
                    "IPC Max": "td_ipc_max",
                    "IPC Min": "td_ipc_min",
                    "Bad Speculation (%)": "td_bad_speculation",
                    "Frontend Bound (%)": "td_frontend_bound",
                    "Retiring (%)": "td_retiring",
                    "Backend Bound (%)": "td_backend_bound",
                    "L3 Bound (%)": "td_l3_bound",
                    "Mem Bound (%)": "td_mem_bound",
                    "Latency Bound (%)": "td_latency_bound",
                    "Bandwidth Bound (%)": "td_bandwidth_bound",
                }
                metrics.update(extract_sheet_metrics(ws, key_map))

            # DevKit_Memory
            if "DevKit_Memory" in wb.sheetnames:
                ws = wb["DevKit_Memory"]
                key_map = {
                    "L1D Miss (%)": "mem_l1d_miss",
                    "L1I Miss (%)": "mem_l1i_miss",
                    "L2D Miss (%)": "mem_l2d_miss",
                    "L2I Miss (%)": "mem_l2i_miss",
                    "DDR Read (MB/s)": "mem_ddr_read",
                    "DDR Write (MB/s)": "mem_ddr_write",
                }
                metrics.update(extract_sheet_metrics(ws, key_map))

            # NUMA_Bandwidth
            if "NUMA_Bandwidth" in wb.sheetnames:
                ws = wb["NUMA_Bandwidth"]
                total_read = 0.0
                total_write = 0.0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        node = str(row[0]).strip()
                        try:
                            read = float(row[1]) if row[1] else 0
                            write = float(row[2]) if row[2] else 0
                            metrics[f"numa{node}_read"] = read
                            metrics[f"numa{node}_write"] = write
                            total_read += read
                            total_write += write
                        except (ValueError, TypeError):
                            pass
                metrics["numa_total_read"] = total_read
                metrics["numa_total_write"] = total_write

            # KSys
            if "KSys" in wb.sheetnames:
                ws = wb["KSys"]
                key_map = {
                    "L2 Miss Latency Max": "ksys_l2_miss_latency_max",
                    "L2 Miss Latency Min": "ksys_l2_miss_latency_min",
                    "L2 Miss Latency Avg": "ksys_l2_miss_latency_avg",
                    "L3 Miss Latency Max": "ksys_l3_miss_latency_max",
                    "L3 Miss Latency Min": "ksys_l3_miss_latency_min",
                    "L3 Miss Latency Avg": "ksys_l3_miss_latency_avg",
                    "IPC": "ksys_ipc",
                    "Retiring (%)": "ksys_retiring",
                    "Frontend Bound (%)": "ksys_frontend_bound",
                    "Bad Speculation (%)": "ksys_bad_speculation",
                    "Backend Bound (%)": "ksys_backend_bound",
                }
                metrics.update(extract_sheet_metrics(ws, key_map))

            # UBWatch_Latency
            if "UBWatch_Latency" in wb.sheetnames:
                ws = wb["UBWatch_Latency"]
                key_map = {
                    "Samples": "ub_samples",
                    "Avg Read (ns)": "ub_avg_read_ns",
                    "Avg Write (ns)": "ub_avg_write_ns",
                    "Min Read (ns)": "ub_min_read_ns",
                    "Min Write (ns)": "ub_min_write_ns",
                    "Max Read (ns)": "ub_max_read_ns",
                    "Max Write (ns)": "ub_max_write_ns",
                }
                metrics.update(extract_sheet_metrics(ws, key_map))

            # UBWatch_Bandwidth (per-chip per-port bandwidth)
            if "UBWatch_Bandwidth" in wb.sheetnames:
                ws = wb["UBWatch_Bandwidth"]
                total_avg_wr = 0.0
                total_avg_rd = 0.0
                total_avg_sum = 0.0
                total_max_wr = 0.0
                total_max_rd = 0.0
                total_max_sum = 0.0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:  # Chip column
                        try:
                            chip = int(row[0])
                            ports = str(row[1]).strip() if row[1] else ""  # Ports column
                            if ports:
                                # Normalize ports name: "0&1" -> "p01", "8" -> "p8"
                                port_key = "p" + ports.replace("&", "")

                                avg_wr = float(row[2]) if row[2] else 0  # Avg Write
                                avg_rd = float(row[3]) if row[3] else 0  # Avg Read
                                avg_sum = float(row[4]) if row[4] else 0  # Avg Sum
                                max_wr = float(row[5]) if row[5] else 0  # Max Write
                                max_rd = float(row[6]) if row[6] else 0  # Max Read
                                max_sum = float(row[7]) if row[7] else 0  # Max Sum

                                # Per-chip+port metrics (e.g., ub_bw_chip0_p01_avg_wr_mb_s)
                                key_prefix = f"ub_bw_chip{chip}_{port_key}"
                                metrics[f"{key_prefix}_avg_wr_mb_s"] = avg_wr
                                metrics[f"{key_prefix}_avg_rd_mb_s"] = avg_rd
                                metrics[f"{key_prefix}_avg_sum_mb_s"] = avg_sum
                                metrics[f"{key_prefix}_max_wr_mb_s"] = max_wr
                                metrics[f"{key_prefix}_max_rd_mb_s"] = max_rd
                                metrics[f"{key_prefix}_max_sum_mb_s"] = max_sum

                                # Accumulate totals
                                total_avg_wr += avg_wr
                                total_avg_rd += avg_rd
                                total_avg_sum += avg_sum
                                total_max_wr = max(total_max_wr, max_wr)
                                total_max_rd = max(total_max_rd, max_rd)
                                total_max_sum = max(total_max_sum, max_sum)
                        except (ValueError, TypeError):
                            pass

                # Total bandwidth metrics
                metrics["ub_bw_total_avg_wr_mb_s"] = total_avg_wr
                metrics["ub_bw_total_avg_rd_mb_s"] = total_avg_rd
                metrics["ub_bw_total_avg_sum_mb_s"] = total_avg_sum
                metrics["ub_bw_total_max_wr_mb_s"] = total_max_wr
                metrics["ub_bw_total_max_rd_mb_s"] = total_max_rd
                metrics["ub_bw_total_max_sum_mb_s"] = total_max_sum

            # SMAPBW_Summary (SMAP migration bandwidth summary)
            if "SMAPBW_Summary" in wb.sheetnames:
                ws = wb["SMAPBW_Summary"]
                key_map = {
                    "Total Cycles": "smapbw_total_cycles",
                    "Total Pages": "smapbw_total_pages",
                    "Avg Bandwidth (GB/s)": "smapbw_avg_bandwidth_gb_s",
                    "Min Bandwidth (GB/s)": "smapbw_min_bandwidth_gb_s",
                    "Max Bandwidth (GB/s)": "smapbw_max_bandwidth_gb_s",
                }
                metrics.update(extract_sheet_metrics(ws, key_map))

            # SMAPBW_Cycles (per-cycle bandwidth details)
            if "SMAPBW_Cycles" in wb.sheetnames:
                ws = wb["SMAPBW_Cycles"]
                cycles_list = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:  # Cycle column
                        try:
                            cycle_no = int(row[0])
                            bandwidth = float(row[3]) if row[3] else 0  # Bandwidth (GB/s) column
                            cycles_list.append(bandwidth)
                            metrics[f"smapbw_cycle{cycle_no}_bandwidth_gb_s"] = bandwidth
                        except (ValueError, TypeError):
                            pass
                if cycles_list:
                    metrics["smapbw_cycle_count"] = len(cycles_list)
                    if "smapbw_avg_bandwidth_gb_s" not in metrics or metrics.get("smapbw_avg_bandwidth_gb_s", 0) == 0:
                        metrics["smapbw_avg_bandwidth_gb_s"] = sum(cycles_list) / len(cycles_list)
                    if "smapbw_min_bandwidth_gb_s" not in metrics or metrics.get("smapbw_min_bandwidth_gb_s", 0) == 0:
                        metrics["smapbw_min_bandwidth_gb_s"] = min(cycles_list)
                    if "smapbw_max_bandwidth_gb_s" not in metrics or metrics.get("smapbw_max_bandwidth_gb_s", 0) == 0:
                        metrics["smapbw_max_bandwidth_gb_s"] = max(cycles_list)

            # Getfre_Summary (Core Frequency Summary)
            if "Getfre_Summary" in wb.sheetnames:
                ws = wb["Getfre_Summary"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:  # NUMA column
                        try:
                            numa = int(row[0])
                            avg_freq = float(row[1]) if row[1] else 0  # Avg Frequency (MHz)
                            min_freq = float(row[2]) if row[2] else 0  # Min Frequency (MHz)
                            max_freq = float(row[3]) if row[3] else 0  # Max Frequency (MHz)
                            sample_count = int(row[4]) if row[4] else 0  # Sample Count
                            core_count = int(row[5]) if row[5] else 0  # Core Count

                            metrics[f"getfre_numa{numa}_avg_mhz"] = avg_freq
                            metrics[f"getfre_numa{numa}_min_mhz"] = min_freq
                            metrics[f"getfre_numa{numa}_max_mhz"] = max_freq
                            metrics[f"getfre_numa{numa}_sample_count"] = sample_count
                            metrics[f"getfre_numa{numa}_core_count"] = core_count
                        except (ValueError, TypeError):
                            pass

            # Getfre_NUMA sheets (Per-core frequency details)
            getfre_numa_sheets = [s for s in wb.sheetnames if s.startswith("Getfre_NUMA")]
            for sheet_name in getfre_numa_sheets:
                numa_id = sheet_name.replace("Getfre_NUMA", "")
                ws = wb[sheet_name]
                freq_avg_list = []
                freq_min_list = []
                freq_max_list = []

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:  # Core ID column
                        try:
                            core_id = int(row[0])
                            avg_freq = float(row[1]) if row[1] else 0  # Avg Frequency
                            min_freq = float(row[2]) if row[2] else 0  # Min Frequency
                            max_freq = float(row[3]) if row[3] else 0  # Max Frequency

                            freq_avg_list.append(avg_freq)
                            freq_min_list.append(min_freq)
                            freq_max_list.append(max_freq)

                            # Store variance cores
                            if max_freq - min_freq > 100:
                                metrics[f"getfre_numa{numa_id}_core{core_id}_avg_mhz"] = avg_freq
                                metrics[f"getfre_numa{numa_id}_core{core_id}_variance_mhz"] = max_freq - min_freq
                        except (ValueError, TypeError):
                            pass

                # Calculate overall statistics
                if freq_avg_list:
                    metrics[f"getfre_numa{numa_id}_overall_avg_mhz"] = sum(freq_avg_list) / len(freq_avg_list)
                    metrics[f"getfre_numa{numa_id}_overall_min_mhz"] = min(freq_min_list)
                    metrics[f"getfre_numa{numa_id}_overall_max_mhz"] = max(freq_max_list)
                    metrics[f"getfre_numa{numa_id}_overall_variance_mhz"] = max(freq_max_list) - min(freq_min_list)
                    metrics[f"getfre_numa{numa_id}_core_count"] = len(freq_avg_list)

        except ImportError:
            print(f"WARNING: Neither pandas nor openpyxl available for {result_dir}")
        except Exception as e:
            print(f"WARNING: Failed to read Excel for {result_dir}: {e}")

    return metrics


def collect_all_results(tasks: List[TestTask], base_dir: str) -> Dict[str, Any]:
    """Collect results from all test runs by parsing raw result files"""
    all_metrics = []

    for task in tasks:
        if not task.result_dir:
            continue

        if not os.path.exists(task.result_dir):
            continue

        # Extract browser metrics from vm_bench_lite report
        browser_metrics = extract_browser_metrics_from_report(task.result_dir)

        # Extract ALL QEMU and performance metrics from analysis_report.xlsx
        qemu_metrics = extract_qemu_metrics_from_excel(task.result_dir)

        # Try to get parameters from config.yaml in result dir
        params = {}
        config_path = os.path.join(task.result_dir, "config.yaml")
        if os.path.exists(config_path):
            try:
                with open(config_path) as f:
                    config = yaml.safe_load(f)
                params = {
                    "vm_count": int(config.get("vm", {}).get("count", 0)),
                    "ratio": float(config.get("smap_tool", {}).get("ratio", 0)),
                    "active_percent": float(config.get("test", {}).get("active_percent", 0)),
                }
            except Exception:
                # Fallback to task parameters
                params = {
                    "vm_count": task.vm_count,
                    "ratio": task.ratio,
                    "active_percent": task.active_percent,
                }
        else:
            params = {
                "vm_count": task.vm_count,
                "ratio": task.ratio,
                "active_percent": task.active_percent,
            }

        # Build metrics dict - pass ALL extracted metrics
        metrics = {
            "task_id": task.task_id,
            "success": task.success,
            "parameters": params,
            "browser_metrics": browser_metrics,
            "qemu_metrics": qemu_metrics,  # All metrics from Excel
        }

        all_metrics.append(metrics)

    return {
        "tasks": all_metrics,
        "total_tests": len(tasks),
        "successful_tests": sum(1 for t in tasks if t.success),
        "failed_tests": sum(1 for t in tasks if not t.success),
    }


def generate_summary_report(results: Dict, output_path: str):
    """Generate Excel summary report with ALL metrics from all sheets

    Output format:
    Row 1: Source headers (merged cells showing data source)
    Row 2: Column headers
    Row 3+: Data rows
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Build DataFrame with all metrics
        rows = []
        for task_data in results["tasks"]:
            params = task_data.get("parameters", {})
            browser = task_data.get("browser_metrics", {})
            qemu = task_data.get("qemu_metrics", {})

            # Basic test info
            row = {
                "test_id": task_data.get("task_id", ""),
                "vm_count": params.get("vm_count", 0),
                "ratio": params.get("ratio", 0),
                "active_percent": params.get("active_percent", 0),
                "active_vm_count": int(params.get("vm_count", 0) * params.get("active_percent", 0)),
                "success": task_data.get("success", False),
            }

            # Browser metrics
            row["browser_total_tasks"] = browser.get("total_tasks", 0)
            row["browser_success_rate"] = browser.get("success_rate", 0)
            row["browser_avg_latency_ms"] = browser.get("avg_latency", 0)
            row["browser_p99_latency_ms"] = browser.get("p99_latency", 0)

            # VM CPU metrics (from Summary sheet)
            row["vm_avg_cpu_percent"] = qemu.get("avg_cpu_percent", 0)
            row["vm_max_cpu_percent"] = qemu.get("max_cpu_percent", 0)

            # DevKit TopDown metrics (13 metrics)
            row["td_cycles_avg"] = qemu.get("td_cycles_avg", 0)
            row["td_instructions_avg"] = qemu.get("td_instructions_avg", 0)
            row["td_ipc_avg"] = qemu.get("td_ipc_avg", 0)
            row["td_ipc_max"] = qemu.get("td_ipc_max", 0)
            row["td_ipc_min"] = qemu.get("td_ipc_min", 0)
            row["td_bad_speculation_percent"] = qemu.get("td_bad_speculation", 0)
            row["td_frontend_bound_percent"] = qemu.get("td_frontend_bound", 0)
            row["td_retiring_percent"] = qemu.get("td_retiring", 0)
            row["td_backend_bound_percent"] = qemu.get("td_backend_bound", 0)
            row["td_l3_bound_percent"] = qemu.get("td_l3_bound", 0)
            row["td_mem_bound_percent"] = qemu.get("td_mem_bound", 0)
            row["td_latency_bound_percent"] = qemu.get("td_latency_bound", 0)
            row["td_bandwidth_bound_percent"] = qemu.get("td_bandwidth_bound", 0)

            # DevKit Memory metrics (6 metrics)
            row["mem_l1d_miss_percent"] = qemu.get("mem_l1d_miss", 0)
            row["mem_l1i_miss_percent"] = qemu.get("mem_l1i_miss", 0)
            row["mem_l2d_miss_percent"] = qemu.get("mem_l2d_miss", 0)
            row["mem_l2i_miss_percent"] = qemu.get("mem_l2i_miss", 0)
            row["mem_ddr_read_mb_s"] = qemu.get("mem_ddr_read", 0)
            row["mem_ddr_write_mb_s"] = qemu.get("mem_ddr_write", 0)

            # NUMA Bandwidth (total)
            row["numa_total_read_mb_s"] = qemu.get("numa_total_read", 0)
            row["numa_total_write_mb_s"] = qemu.get("numa_total_write", 0)
            # Per-node NUMA bandwidth (dynamically add if exists)
            for key, value in qemu.items():
                if key.startswith("numa") and "_read" in key and key != "numa_total_read":
                    row[key] = value
                elif key.startswith("numa") and "_write" in key and key != "numa_total_write":
                    row[key] = value

            # KSys metrics
            row["ksys_l2_miss_latency_max"] = qemu.get("ksys_l2_miss_latency_max", 0)
            row["ksys_l2_miss_latency_min"] = qemu.get("ksys_l2_miss_latency_min", 0)
            row["ksys_l2_miss_latency_avg"] = qemu.get("ksys_l2_miss_latency_avg", 0)
            row["ksys_l3_miss_latency_max"] = qemu.get("ksys_l3_miss_latency_max", 0)
            row["ksys_l3_miss_latency_min"] = qemu.get("ksys_l3_miss_latency_min", 0)
            row["ksys_l3_miss_latency_avg"] = qemu.get("ksys_l3_miss_latency_avg", 0)
            row["ksys_ipc"] = qemu.get("ksys_ipc", 0)
            row["ksys_retiring_percent"] = qemu.get("ksys_retiring", 0)
            row["ksys_frontend_bound_percent"] = qemu.get("ksys_frontend_bound", 0)
            row["ksys_bad_speculation_percent"] = qemu.get("ksys_bad_speculation", 0)
            row["ksys_backend_bound_percent"] = qemu.get("ksys_backend_bound", 0)

            # UBWatch Latency metrics
            row["ub_samples"] = qemu.get("ub_samples", 0)
            row["ub_avg_read_ns"] = qemu.get("ub_avg_read_ns", 0)
            row["ub_avg_write_ns"] = qemu.get("ub_avg_write_ns", 0)
            row["ub_min_read_ns"] = qemu.get("ub_min_read_ns", 0)
            row["ub_min_write_ns"] = qemu.get("ub_min_write_ns", 0)
            row["ub_max_read_ns"] = qemu.get("ub_max_read_ns", 0)
            row["ub_max_write_ns"] = qemu.get("ub_max_write_ns", 0)

            # UBWatch Bandwidth metrics - Total (6 metrics)
            row["ub_bw_total_avg_wr_mb_s"] = qemu.get("ub_bw_total_avg_wr_mb_s", 0)
            row["ub_bw_total_avg_rd_mb_s"] = qemu.get("ub_bw_total_avg_rd_mb_s", 0)
            row["ub_bw_total_avg_sum_mb_s"] = qemu.get("ub_bw_total_avg_sum_mb_s", 0)
            row["ub_bw_total_max_wr_mb_s"] = qemu.get("ub_bw_total_max_wr_mb_s", 0)
            row["ub_bw_total_max_rd_mb_s"] = qemu.get("ub_bw_total_max_rd_mb_s", 0)
            row["ub_bw_total_max_sum_mb_s"] = qemu.get("ub_bw_total_max_sum_mb_s", 0)

            # UBWatch Bandwidth - Per-chip metrics (dynamically add if exists)
            for key, value in qemu.items():
                if key.startswith("ub_bw_chip"):
                    row[key] = value

            # SMAPBW Summary metrics (5 metrics)
            row["smapbw_total_cycles"] = qemu.get("smapbw_total_cycles", 0)
            row["smapbw_total_pages"] = qemu.get("smapbw_total_pages", 0)
            row["smapbw_avg_bandwidth_gb_s"] = qemu.get("smapbw_avg_bandwidth_gb_s", 0)
            row["smapbw_min_bandwidth_gb_s"] = qemu.get("smapbw_min_bandwidth_gb_s", 0)
            row["smapbw_max_bandwidth_gb_s"] = qemu.get("smapbw_max_bandwidth_gb_s", 0)

            # Getfre Core Frequency metrics - Per-NUMA summary
            # Dynamically add getfre metrics for each NUMA found
            for key, value in qemu.items():
                if key.startswith("getfre_numa") and ("_avg_mhz" in key or "_min_mhz" in key or "_max_mhz" in key):
                    row[key] = value
                elif key.startswith("getfre_numa") and "_overall_variance_mhz" in key:
                    row[key] = value
                elif key.startswith("getfre_numa") and "_core_count" in key:
                    row[key] = value

            rows.append(row)

        df = pd.DataFrame(rows)

        # Sort by vm_count, ratio, active_percent ascending
        df = df.sort_values(by=['vm_count', 'ratio', 'active_percent'], ascending=[True, True, True])

        # Save to Excel first (creates basic structure)
        df.to_excel(output_path, index=False, sheet_name="Summary")

        # Now add source header row with merged cells
        wb = load_workbook(output_path)
        ws = wb["Summary"]

        # Define column groupings with their sources
        # Format: (source_name, start_col, end_col) - columns are 1-indexed
        column_sources = []

        # Track column positions
        col_idx = 1

        # Basic info (test_id, vm_count, ratio, active_percent, active_vm_count, success) - 6 columns
        column_sources.append(("Basic Info (目录/config.yaml)", col_idx, col_idx + 5))
        col_idx += 6

        # Browser metrics - 4 columns
        column_sources.append(("Browser (vm_bench_lite/bench_report.txt)", col_idx, col_idx + 3))
        col_idx += 4

        # VM CPU - 2 columns
        column_sources.append(("VM CPU (Excel: Summary)", col_idx, col_idx + 1))
        col_idx += 2

        # DevKit TopDown - 13 columns
        column_sources.append(("DevKit TopDown (Excel: DevKit_TopDown)", col_idx, col_idx + 12))
        col_idx += 13

        # DevKit Memory - 6 columns
        column_sources.append(("DevKit Memory (Excel: DevKit_Memory)", col_idx, col_idx + 5))
        col_idx += 6

        # NUMA Bandwidth - 2 columns (base) + dynamic per-node columns
        numa_end = col_idx + 1
        # Check if there are per-node numa columns
        numa_cols = [c for c in df.columns if c.startswith("numa") and c not in ["numa_total_read_mb_s", "numa_total_write_mb_s"]]
        numa_end += len(numa_cols)
        column_sources.append(("NUMA Bandwidth (Excel: NUMA_Bandwidth)", col_idx, numa_end))
        col_idx = numa_end + 1

        # KSys - 11 columns
        column_sources.append(("KSys (Excel: KSys)", col_idx, col_idx + 10))
        col_idx += 11

        # UBWatch Latency - 7 columns
        column_sources.append(("UBWatch Latency (Excel: UBWatch_Latency)", col_idx, col_idx + 6))
        col_idx += 7

        # UBWatch Bandwidth Total - 6 columns + dynamic per-chip+port columns
        ub_bw_total_end = col_idx + 5
        # Check if there are per-chip+port ub_bw columns
        ub_bw_chip_cols = [c for c in df.columns if c.startswith("ub_bw_chip")]
        ub_bw_total_end += len(ub_bw_chip_cols)
        column_sources.append(("UBWatch Bandwidth (Excel: UBWatch_Bandwidth)", col_idx, ub_bw_total_end))
        col_idx = ub_bw_total_end + 1

        # SMAPBW Summary - 5 columns (summary data only)
        column_sources.append(("SMAPBW (Excel: SMAPBW_Summary)", col_idx, col_idx + 4))
        col_idx += 5

        # Getfre Core Frequency - dynamic columns (per-NUMA metrics)
        getfre_cols = [c for c in df.columns if c.startswith("getfre_numa")]
        if getfre_cols:
            getfre_end = col_idx + len(getfre_cols) - 1
            column_sources.append(("Getfre CoreFreq (Excel: Getfre_Summary)", col_idx, getfre_end))

        # Insert source header row at row 1
        ws.insert_rows(1)

        # Style definitions
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Color fills for different sources
        source_fills = {
            "Basic": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),  # Blue
            "Browser": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),  # Green
            "VM CPU": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # Orange
            "DevKit TopDown": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),  # Dark Orange
            "DevKit Memory": PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid"),  # Gray
            "NUMA": PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),  # Light Blue
            "KSys": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),  # Purple
            "UBWatch Latency": PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid"),  # Brown
            "UBWatch Bandwidth": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),  # Dark Green
            "SMAPBW": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),  # Red/Pink
            "Getfre": PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),  # Cyan
        }

        # Create merged cells for source headers
        for source_name, start_col, end_col in column_sources:
            if start_col <= end_col:
                # Merge cells
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

                # Set value and style
                cell = ws.cell(row=1, column=start_col)
                cell.value = source_name
                cell.alignment = center_align
                cell.font = header_font_white

                # Find matching fill color
                fill = header_fill
                for key, color_fill in source_fills.items():
                    if key in source_name:
                        fill = color_fill
                        break
                cell.fill = fill

                # Add border to merged range
                for c in range(start_col, end_col + 1):
                    ws.cell(row=1, column=c).border = thin_border

        # Style the column header row (row 2 now, after insert)
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=2, column=c)
            cell.font = header_font
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.alignment = center_align
            cell.border = thin_border

        # Set row height for header rows
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20

        # Freeze panes (freeze first two rows)
        ws.freeze_panes = "A3"

        # Save the workbook
        wb.save(output_path)

        print(f"\nSummary report saved to: {output_path}")
        print(f"  - Row 1: Data source headers (merged cells)")
        print(f"  - Row 2: Column names")
        print(f"  - Row 3+: Test data")

    except ImportError as e:
        print(f"WARNING: Required libraries not available ({e}), saving as JSON instead")
        json_path = output_path.replace(".xlsx", ".json")
        with open(json_path, "w") as f:
            json.dump(results, f, indent=2)
        print(f"Summary saved to: {json_path}")


def scan_existing_results(result_base_dir: str) -> List[TestTask]:
    """Scan existing test result directories and create TestTask list"""
    tasks = []

    if not os.path.exists(result_base_dir):
        print(f"ERROR: Result directory not found: {result_base_dir}")
        return tasks

    # Scan subdirectories matching pattern: vm{n}_ratio{ratio}_active{percent}_timestamp
    pattern = re.compile(r"^vm(\d+)_ratio([\d.]+)_active([\d.]+)_\d{8}_\d{6}$")

    for entry in os.listdir(result_base_dir):
        entry_path = os.path.join(result_base_dir, entry)

        if not os.path.isdir(entry_path):
            continue

        # Try to match pattern
        match = pattern.match(entry)
        if match:
            vm_count = int(match.group(1))
            ratio = float(match.group(2))
            active_percent = float(match.group(3))

            # Check if result files exist (backward compatibility: check both vm_monitor and qemu_monitor)
            has_vm = os.path.exists(os.path.join(entry_path, "vm_monitor", "analysis_report.xlsx")) or \
                     os.path.exists(os.path.join(entry_path, "qemu_monitor", "analysis_report.xlsx"))
            has_bench = os.path.exists(os.path.join(entry_path, "vm_bench_lite"))

            # Consider success if both result directories exist
            success = has_vm or has_bench

            task = TestTask(
                task_id=entry,
                vm_count=vm_count,
                ratio=ratio,
                active_percent=active_percent,
                result_dir=entry_path,
                success=success
            )
            tasks.append(task)
            print(f"  Found: {entry} (vm={vm_count}, ratio={ratio}, active={active_percent})")

    # Sort by task_id
    tasks.sort(key=lambda t: t.task_id)

    return tasks


def offline_summary(result_base_dir: str, output_path: str = None):
    """Generate batch summary from existing test results (offline mode)"""
    print("=" * 60)
    print("Offline Batch Summary Generator")
    print("=" * 60)
    print(f"Scanning result directory: {result_base_dir}")

    # Scan existing results
    tasks = scan_existing_results(result_base_dir)

    if not tasks:
        print("ERROR: No valid test result directories found")
        return

    print(f"\nFound {len(tasks)} test results")

    # Summary of what we found
    success_count = sum(1 for t in tasks if t.success)
    print(f"  Complete results (has data files): {success_count}")
    print(f"  Incomplete results: {len(tasks) - success_count}")

    # Collect results
    print("\n" + "=" * 60)
    print("Extracting metrics from result files...")
    print("=" * 60)

    results = collect_all_results(tasks, result_base_dir)

    # Generate output path if not specified
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(result_base_dir, f"batch_summary_{timestamp}.xlsx")

    # Generate summary report
    print("\nGenerating summary report...")
    generate_summary_report(results, output_path)

    # Print summary
    print("\n" + "=" * 60)
    print("Offline Summary Complete")
    print("=" * 60)
    print(f"Total results processed: {len(tasks)}")
    print(f"Successful: {results['successful_tests']}")
    print(f"Failed/Incomplete: {results['failed_tests']}")
    print(f"Summary report: {output_path}")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(description="Multi-VM Agent Batch Test Scheduler")
    parser.add_argument("--config", default="batch_config.yaml", help="Batch config YAML file")
    parser.add_argument("--dry-run", action="store_true", help="Print tasks without executing")
    parser.add_argument("--offline", action="store_true", help="Generate summary from existing results (offline mode)")
    parser.add_argument("--result-dir", help="Result directory for offline mode (default: from config)")
    parser.add_argument("--output", help="Output Excel path for offline mode")

    args = parser.parse_args()

    # Offline mode: generate summary from existing results
    if args.offline:
        # Determine result directory
        if args.result_dir:
            result_dir = args.result_dir
        else:
            # Try to load from config
            try:
                batch_config = load_batch_config(args.config)
                result_dir = batch_config["result"]["base_dir"]
            except Exception:
                print("ERROR: Cannot determine result directory. Use --result-dir or provide valid --config")
                return

        offline_summary(result_dir, args.output)
        return

    # Load config
    batch_config = load_batch_config(args.config)

    # Generate tasks
    tasks = generate_test_tasks(batch_config)

    print("=" * 60)
    print("Multi-VM Agent Batch Test Scheduler")
    print("=" * 60)
    print(f"Total tasks: {len(tasks)}")
    print(f"VM counts: {batch_config['test_matrix']['vm_counts']}")
    print(f"Ratios: {batch_config['test_matrix']['ratios']}")
    print(f"Active percentages: {batch_config['test_matrix']['active_percentages']}")
    print("=" * 60)

    # Print task list
    print("\nTask list:")
    for i, task in enumerate(tasks):
        print(f"  [{i+1}] {task.task_id}")

    if args.dry_run:
        print("\n[Dry run mode] Tasks listed but not executed")
        return

    # Create directories
    base_dir = batch_config["result"]["base_dir"]
    temp_config_dir = os.path.join(base_dir, "temp_configs")
    os.makedirs(temp_config_dir, exist_ok=True)

    # Batch log file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    batch_log_file = os.path.join(base_dir, f"batch_log_{timestamp}.txt")

    with open(batch_log_file, "w") as f:
        f.write("=" * 60 + "\n")
        f.write("Multi-VM Agent Batch Test Log\n")
        f.write(f"Started: {timestamp}\n")
        f.write(f"Total tasks: {len(tasks)}\n")
        f.write("=" * 60 + "\n")

    print(f"\nBatch log: {batch_log_file}")

    # Template path
    template_path = batch_config["result"]["template_path"]
    fixed_params = batch_config["fixed_params"]
    scheduler_config = batch_config["scheduler"]

    # Execute tests
    start_time = time.time()

    for i, task in enumerate(tasks):
        print(f"\n{'=' * 60}")
        print(f"Task [{i+1}/{len(tasks)}]: {task.task_id}")
        print(f"{'=' * 60}")

        # Generate temp config
        task.config_file = generate_temp_config(
            template_path, task, fixed_params, base_dir, temp_config_dir
        )

        # Run test
        success, result_dir = run_single_test(task, task.config_file, batch_log_file)

        task.success = success
        task.result_dir = result_dir

        # Move config to result directory
        if result_dir:
            move_config_to_result(task.config_file, result_dir, task.task_id)
        else:
            # Keep failed config for debugging in temp_configs
            pass

        # Handle failure
        if not success and not scheduler_config["continue_on_failure"]:
            print(f"\nTest failed, stopping batch execution")
            break

    elapsed = time.time() - start_time

    # Collect results
    print("\n" + "=" * 60)
    print("Collecting results...")
    print("=" * 60)

    results = collect_all_results(tasks, base_dir)

    # Generate summary report
    summary_path = os.path.join(base_dir, f"batch_summary_{timestamp}.xlsx")
    generate_summary_report(results, summary_path)

    # Final summary
    with open(batch_log_file, "a") as f:
        f.write("\n" + "=" * 60 + "\n")
        f.write("Batch Test Completed\n")
        f.write(f"Total time: {elapsed:.1f} seconds\n")
        f.write(f"Successful: {results['successful_tests']}\n")
        f.write(f"Failed: {results['failed_tests']}\n")
        f.write("=" * 60 + "\n")

    print("\n" + "=" * 60)
    print("Batch Test Summary")
    print("=" * 60)
    print(f"Total time: {elapsed:.1f} seconds")
    print(f"Successful tests: {results['successful_tests']}/{len(tasks)}")
    print(f"Failed tests: {results['failed_tests']}/{len(tasks)}")
    print(f"Batch log: {batch_log_file}")
    print(f"Summary report: {summary_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()