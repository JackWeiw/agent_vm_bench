# qemu_monitor/exporters.py
"""
Data Export Module

Exports monitoring data and parsed logs to Excel/CSV formats with charts.
Handles backward compatibility for pandas availability.
"""

import os
from datetime import datetime
from typing import Dict, List, Optional, Any, TYPE_CHECKING

# Try to import pandas for Excel export
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Internal dependency
from .parsers import parse_all_logs

# Type checking import (avoid circular dependency)
if TYPE_CHECKING:
    from .monitor import QEMUMonitor


def export_to_excel(monitor: 'QEMUMonitor', log_dir: str, numa_nodes: list = None,
                    output_file: str = None, capture_results: dict = None) -> str:
    """Export all monitoring and parsed log data to Excel

    Args:
        monitor: QEMUMonitor instance with collected data
        log_dir: directory containing log files
        numa_nodes: list of NUMA nodes monitored
        output_file: Excel output filename (default: analysis_report.xlsx in log_dir)
        capture_results: LogCapture results (optional)

    Returns:
        Path to generated Excel file
    """
    if not PANDAS_AVAILABLE:
        print("[WARN] pandas not available, skipping Excel export")
        print("  Install with: pip install pandas openpyxl")
        return None

    if output_file is None:
        output_file = os.path.join(log_dir, 'analysis_report.xlsx')

    # Parse log files
    parsed_logs = parse_all_logs(log_dir, numa_nodes)

    # Create writer for multi-sheet Excel (openpyxl required)
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # ========== Sheet 1: Summary Overview ==========
            summary_data = {
                'Metric': [],
                'Value': [],
                'Unit': []
            }

            # Test metadata
            summary_data['Metric'].extend(['Test Date', 'Duration', 'Sampling Interval', 'NUMA Nodes'])
            summary_data['Value'].extend([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                f"{monitor.data[0]['timestamp'] if monitor.data else 'N/A'} ~ {monitor.data[-1]['timestamp'] if monitor.data else 'N/A'}",
                'N/A',  # Interval not stored in monitor
                ','.join(map(str, numa_nodes)) if numa_nodes else 'N/A'
            ])
            summary_data['Unit'].extend(['', '', 'seconds', ''])

            # Host machine stats
            if monitor.host_cpu_history:
                host_cpu_avg = round(sum(monitor.host_cpu_history)/len(monitor.host_cpu_history), 1)
                summary_data['Metric'].extend(['Host Avg CPU', 'Host Peak CPU', 'Host Avg Memory', 'Host Peak Memory'])
                summary_data['Value'].extend([
                    host_cpu_avg, round(monitor.peak_host_cpu, 1),
                    round(sum(h['used_mb'] for h in monitor.host_mem_history)/len(monitor.host_mem_history), 0),
                    round(monitor.peak_host_mem_mb, 0)
                ])
                summary_data['Unit'].extend(['%', '%', 'MB', 'MB'])

            # Hugepage stats
            summary_data['Metric'].extend(['Hugepage Total', 'Hugepage Avg Used', 'Hugepage Peak Used', 'Hugepage Peak Usage %'])
            summary_data['Value'].extend([
                round(monitor.hugepage_total_mb, 0),
                round(sum(monitor.hugepage_used_history)/len(monitor.hugepage_used_history), 0) if monitor.hugepage_used_history else 0,
                round(monitor.peak_hugepage_used_mb, 0),
                round((monitor.peak_hugepage_used_mb / monitor.hugepage_total_mb * 100), 1) if monitor.hugepage_total_mb > 0 else 0
            ])
            summary_data['Unit'].extend(['MB', 'MB', 'MB', '%'])

            # Swap stats
            if monitor.swap_history:
                swap_avg = round(sum(s['used_mb'] for s in monitor.swap_history)/len(monitor.swap_history), 0)
                swap_total = monitor.swap_history[0]['total_mb'] if monitor.swap_history else 0
                swap_peak_pct = round(monitor.peak_swap_used_mb / swap_total * 100, 1) if swap_total > 0 else 0
                summary_data['Metric'].extend(['Swap Total', 'Swap Avg Used', 'Swap Peak Used', 'Swap Peak Usage %'])
                summary_data['Value'].extend([round(swap_total, 0), swap_avg, round(monitor.peak_swap_used_mb, 0), swap_peak_pct])
                summary_data['Unit'].extend(['MB', 'MB', 'MB', '%'])

            # VM stats
            vm_stats = monitor.calculate_vm_stats()
            overall_stats = monitor.calculate_overall_stats(vm_stats)
            summary_data['Metric'].extend(['Total VMs', 'Alive VMs at End', 'VM Avg CPU', 'VM Peak Total CPU', 'Total Avg Memory'])
            summary_data['Value'].extend([
                overall_stats['total_vms'], monitor.last_vm_count,
                overall_stats['overall_avg_cpu'], round(monitor.peak_total_cpu, 1),
                overall_stats['total_avg_memory_mb']
            ])
            summary_data['Unit'].extend(['', '', '%', '%', 'MB'])

            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)

            # ========== Sheet 2: NUMA CPU Stats ==========
            numa_cpu_data = {'NUMA Node': [], 'Avg CPU (%)': [], 'Peak CPU (%)': []}
            for node in sorted(monitor.numa_cpu_history.keys()):
                hist = monitor.numa_cpu_history[node]
                if hist:
                    numa_cpu_data['NUMA Node'].append(node)
                    numa_cpu_data['Avg CPU (%)'].append(round(sum(hist)/len(hist), 1))
                    numa_cpu_data['Peak CPU (%)'].append(round(monitor.numa_cpu_peak[node], 1))
            if numa_cpu_data['NUMA Node']:
                pd.DataFrame(numa_cpu_data).to_excel(writer, sheet_name='NUMA_CPU', index=False)

            # ========== Sheet 3: NUMA Memory Stats ==========
            numa_mem_data = {'NUMA Node': [], 'Avg Used (MB)': [], 'Peak Used (MB)': [], 'Avg Usage (%)': []}
            numa_summary = {}
            from collections import defaultdict
            numa_summary = defaultdict(lambda: {'used': [], 'usage': []})
            for entry in monitor.numa_memory_history:
                for n in entry['nodes']:
                    numa_summary[n['node']]['used'].append(n['used'])
                    numa_summary[n['node']]['usage'].append(n['usage'])
            for node_id in sorted(numa_summary.keys()):
                data = numa_summary[node_id]
                numa_mem_data['NUMA Node'].append(node_id)
                numa_mem_data['Avg Used (MB)'].append(round(sum(data['used'])/len(data['used']), 0) if data['used'] else 0)
                numa_mem_data['Peak Used (MB)'].append(round(max(data['used']), 0) if data['used'] else 0)
                numa_mem_data['Avg Usage (%)'].append(round(sum(data['usage'])/len(data['usage']), 1) if data['usage'] else 0)
            if numa_mem_data['NUMA Node']:
                pd.DataFrame(numa_mem_data).to_excel(writer, sheet_name='NUMA_Memory', index=False)

            # ========== Sheet 4: Hugepage Per NUMA ==========
            if monitor.hugepage_per_numa_history:
                hp_data = {'NUMA Node': [], 'Avg Total (MB)': [], 'Avg Used (MB)': [], 'Avg Usage (%)': []}
                hp_summary = defaultdict(lambda: {'total': [], 'used': [], 'usage': []})
                for entry in monitor.hugepage_per_numa_history:
                    for node_id, data in entry['nodes'].items():
                        hp_summary[node_id]['total'].append(data['total_mb'])
                        hp_summary[node_id]['used'].append(data['used_mb'])
                        hp_summary[node_id]['usage'].append(data['usage_pct'])
                for node_id in sorted(hp_summary.keys()):
                    data = hp_summary[node_id]
                    hp_data['NUMA Node'].append(node_id)
                    hp_data['Avg Total (MB)'].append(round(sum(data['total'])/len(data['total']), 0) if data['total'] else 0)
                    hp_data['Avg Used (MB)'].append(round(sum(data['used'])/len(data['used']), 0) if data['used'] else 0)
                    hp_data['Avg Usage (%)'].append(round(sum(data['usage'])/len(data['usage']), 1) if data['usage'] else 0)
                pd.DataFrame(hp_data).to_excel(writer, sheet_name='Hugepage_Per_NUMA', index=False)

            # ========== Sheet 5: VM Statistics ==========
            if vm_stats:
                vm_data = {
                    'VM Name': [v['vm_name'] for v in vm_stats],
                    'PID': [v['pid'] for v in vm_stats],
                    'Samples': [v['sample_count'] for v in vm_stats],
                    'Avg CPU (%)': [v['avg_cpu'] for v in vm_stats],
                    'Max CPU (%)': [v['max_cpu'] for v in vm_stats],
                    'Avg Memory (MB)': [v['avg_memory_mb'] for v in vm_stats],
                    'Max Memory (MB)': [v['max_memory_mb'] for v in vm_stats],
                    'Avg Hugepage (MB)': [v.get('avg_huge_mb', 0) for v in vm_stats],
                }
                pd.DataFrame(vm_data).to_excel(writer, sheet_name='VM_Stats', index=False)

            # ========== Sheet 6: DevKit Top-Down ==========
            if 'devkit_top_down' in parsed_logs and 'error' not in parsed_logs['devkit_top_down']:
                td = parsed_logs['devkit_top_down']
                td_data = {
                    'Metric': ['Cycles Avg', 'Instructions Avg', 'IPC Avg',
                              'Bad Speculation (%)', 'Frontend Bound (%)', 'Retiring (%)', 'Backend Bound (%)',
                              'L3 Bound (%)', 'Mem Bound (%)', 'Latency Bound (%)', 'Bandwidth Bound (%)',
                              'IPC Max', 'IPC Min'],
                    'Value': [
                        td.get('cycles_avg', 0), td.get('instructions_avg', 0), td.get('ipc_avg', 0),
                        td.get('bad_speculation_avg', 0), td.get('frontend_bound_avg', 0),
                        td.get('retiring_avg', 0), td.get('backend_bound_avg', 0),
                        td.get('l3_bound_avg', 0), td.get('mem_bound_avg', 0),
                        td.get('mem_latency_bound_avg', 0), td.get('mem_bandwidth_bound_avg', 0),
                        td.get('ipc_max', 0), td.get('ipc_min', 0)
                    ],
                    'Report Count': [td.get('report_count', 0)] * 13
                }
                pd.DataFrame(td_data).to_excel(writer, sheet_name='DevKit_TopDown', index=False)

                # Timeline sheet (with safety check)
                if td.get('timeline') and td['timeline'].get('timestamp'):
                    try:
                        timeline_data = td['timeline']
                        # Ensure all arrays have same length
                        min_len = min(len(timeline_data.get(k, [])) for k in timeline_data.keys())
                        # Truncate all arrays to same length
                        for k in timeline_data:
                            if isinstance(timeline_data[k], list):
                                timeline_data[k] = timeline_data[k][:min_len]
                        timeline_df = pd.DataFrame(timeline_data)
                        timeline_df.to_excel(writer, sheet_name='TopDown_Timeline', index=False)
                    except Exception as e:
                        print(f"  Warning: TopDown_Timeline creation failed: {e}")

            # ========== Sheet 7: DevKit Memory ==========
            if 'devkit_mem' in parsed_logs and 'error' not in parsed_logs['devkit_mem']:
                mem = parsed_logs['devkit_mem']
                mem_data = {
                    'Metric': ['L1D Miss (%)', 'L1I Miss (%)', 'L2D Miss (%)', 'L2I Miss (%)',
                              'DDR Write (MB/s)', 'DDR Read (MB/s)'],
                    'Value': [
                        mem.get('cache_miss', {}).get('L1D', 0),
                        mem.get('cache_miss', {}).get('L1I', 0),
                        mem.get('cache_miss', {}).get('L2D', 0),
                        mem.get('cache_miss', {}).get('L2I', 0),
                        mem.get('ddr_bandwidth_system', {}).get('write', 0),
                        mem.get('ddr_bandwidth_system', {}).get('read', 0)
                    ],
                    'Report Count': [mem.get('report_count', 0)] * 6
                }
                pd.DataFrame(mem_data).to_excel(writer, sheet_name='DevKit_Memory', index=False)

                # NUMA Bandwidth as separate table
                numa_bw = mem.get('numa_bandwidth', {})
                if numa_bw:
                    bw_data = {'NUMA Node': [], 'Read (MB/s)': [], 'Write (MB/s)': []}
                    for node_id in sorted(numa_bw.keys()):
                        bw_data['NUMA Node'].append(node_id)
                        bw_data['Read (MB/s)'].append(numa_bw[node_id].get('read', 0))
                        bw_data['Write (MB/s)'].append(numa_bw[node_id].get('write', 0))
                    pd.DataFrame(bw_data).to_excel(writer, sheet_name='NUMA_Bandwidth', index=False)

                # Memory Timeline sheet (with safety check)
                if mem.get('timeline') and mem['timeline'].get('timestamp'):
                    try:
                        timeline_data = mem['timeline']
                        # Ensure all arrays have same length
                        min_len = min(len(timeline_data.get(k, [])) for k in timeline_data.keys())
                        # Truncate all arrays to same length
                        for k in timeline_data:
                            if isinstance(timeline_data[k], list):
                                timeline_data[k] = timeline_data[k][:min_len]
                        mem_timeline_df = pd.DataFrame(timeline_data)
                        mem_timeline_df.to_excel(writer, sheet_name='Memory_Timeline', index=False)
                    except Exception as e:
                        print(f"  Warning: Memory_Timeline creation failed: {e}")

            # ========== Sheet 8: KSys ==========
            if 'ksys' in parsed_logs and 'error' not in parsed_logs['ksys']:
                ksys = parsed_logs['ksys']
                ksys_data = {'Metric': [], 'Value': []}

                # Latency
                l2 = ksys.get('l2_miss_latency', {})
                l3 = ksys.get('l3_miss_latency', {})
                if l2:
                    ksys_data['Metric'].extend(['L2 Miss Latency Max', 'L2 Miss Latency Min', 'L2 Miss Latency Avg'])
                    ksys_data['Value'].extend([l2.get('cycles_max', 0), l2.get('cycles_min', 0), l2.get('cycles_avg', 0)])
                if l3:
                    ksys_data['Metric'].extend(['L3 Miss Latency Max', 'L3 Miss Latency Min', 'L3 Miss Latency Avg'])
                    ksys_data['Value'].extend([l3.get('cycles_max', 0), l3.get('cycles_min', 0), l3.get('cycles_avg', 0)])

                # IPC
                if ksys.get('ipc'):
                    ksys_data['Metric'].append('IPC')
                    ksys_data['Value'].append(ksys.get('ipc', 0))

                # Topdown
                td = ksys.get('topdown', {})
                if td:
                    for key, label in [('retiring', 'Retiring (%)'), ('frontend_bound', 'Frontend Bound (%)'),
                                       ('bad_speculation', 'Bad Speculation (%)'), ('backend_bound', 'Backend Bound (%)')]:
                        if td.get(key):
                            ksys_data['Metric'].append(label)
                            ksys_data['Value'].append(td[key])

                if ksys_data['Metric']:
                    pd.DataFrame(ksys_data).to_excel(writer, sheet_name='KSys', index=False)

            # ========== Sheet 9: UB Watch ==========
            if 'ub_watch' in parsed_logs and 'error' not in parsed_logs['ub_watch']:
                ub = parsed_logs['ub_watch']
                ub_data = {'Metric': [], 'Value': []}

                lat = ub.get('latency', {})
                if lat:
                    ub_data['Metric'].extend(['Latency Path', 'Samples', 'Avg Read (ns)', 'Avg Write (ns)',
                                             'Min Read (ns)', 'Min Write (ns)', 'Max Read (ns)', 'Max Write (ns)'])
                    ub_data['Value'].extend([
                        lat.get('path', 'N/A'), lat.get('samples', 0),
                        lat.get('avg_r', 0), lat.get('avg_w', 0),
                        lat.get('min_r', 0), lat.get('min_w', 0),
                        lat.get('max_r', 0), lat.get('max_w', 0)
                    ])

                if ub_data['Metric']:
                    pd.DataFrame(ub_data).to_excel(writer, sheet_name='UBWatch_Latency', index=False)

                # Bandwidth
                bw_list = ub.get('bandwidth', [])
                if bw_list:
                    bw_data = {
                        'Chip': [bw['chip'] for bw in bw_list],
                        'Ports': [bw['ports'] for bw in bw_list],
                        'Avg Write (MB/s)': [bw['avg_wr'] for bw in bw_list],
                        'Avg Read (MB/s)': [bw['avg_rd'] for bw in bw_list],
                        'Avg Sum (MB/s)': [bw['avg_sum'] for bw in bw_list],
                        'Max Write (MB/s)': [bw['max_wr'] for bw in bw_list],
                        'Max Read (MB/s)': [bw['max_rd'] for bw in bw_list],
                        'Max Sum (MB/s)': [bw['max_sum'] for bw in bw_list]
                    }
                    pd.DataFrame(bw_data).to_excel(writer, sheet_name='UBWatch_Bandwidth', index=False)

            # ========== Sheet 10: SMAP BW ==========
            if 'smap_bw' in parsed_logs and 'error' not in parsed_logs['smap_bw']:
                smap = parsed_logs['smap_bw']

                # Summary sheet
                summary = smap.get('summary', {})
                if summary:
                    smap_summary = {
                        'Metric': ['Total Cycles', 'Total Pages', 'Avg Bandwidth (GB/s)',
                                   'Min Bandwidth (GB/s)', 'Max Bandwidth (GB/s)'],
                        'Value': [
                            summary.get('total_cycles', 0),
                            summary.get('total_pages', 0),
                            summary.get('avg_bandwidth_gb_s', 0),
                            summary.get('min_bandwidth_gb_s', 0),
                            summary.get('max_bandwidth_gb_s', 0)
                        ]
                    }
                    pd.DataFrame(smap_summary).to_excel(writer, sheet_name='SMAPBW_Summary', index=False)

                # Per-cycle sheet with direction columns spread out
                cycles = smap.get('cycles', [])
                all_directions = smap.get('all_directions', set())
                if cycles and all_directions:
                    # Sort directions for consistent column order
                    sorted_directions = sorted(all_directions)

                    # Build column data
                    cycle_data = {
                        'Cycle': [c['cycle_no'] for c in cycles],
                        'Pages': [c['total_pages'] for c in cycles],
                        'Duration (s)': [c['duration'] for c in cycles],
                        'Bandwidth (GB/s)': [c['bandwidth_gb_s'] for c in cycles],
                    }

                    # Add direction columns (e.g., "N0->N1_pages")
                    for from_node, to_node in sorted_directions:
                        col_name = f"N{from_node}->N{to_node}_pages"
                        cycle_data[col_name] = [
                            c['directions'].get((from_node, to_node), 0) for c in cycles
                        ]

                    pd.DataFrame(cycle_data).to_excel(writer, sheet_name='SMAPBW_Cycles', index=False)

            # ========== Sheet 11: getfre Core Frequency ==========
            if 'getfre' in parsed_logs:
                getfre_data = parsed_logs['getfre']

                # Summary sheet (all NUMA averages)
                getfre_summary = {
                    'NUMA': [],
                    'Avg Frequency (MHz)': [],
                    'Min Frequency (MHz)': [],
                    'Max Frequency (MHz)': [],
                    'Sample Count': [],
                    'Core Count': []
                }
                for numa_id in sorted(getfre_data.keys()):
                    gf = getfre_data[numa_id]
                    if 'error' not in gf:
                        getfre_summary['NUMA'].append(numa_id)
                        getfre_summary['Avg Frequency (MHz)'].append(gf.get('numa_avg', 0))
                        getfre_summary['Min Frequency (MHz)'].append(gf.get('numa_min', 0))
                        getfre_summary['Max Frequency (MHz)'].append(gf.get('numa_max', 0))
                        getfre_summary['Sample Count'].append(gf.get('sample_count', 0))
                        getfre_summary['Core Count'].append(len(gf.get('core_stats', {})))

                if getfre_summary['NUMA']:
                    pd.DataFrame(getfre_summary).to_excel(writer, sheet_name='Getfre_Summary', index=False)

                # Per-core sheet (per NUMA)
                for numa_id in sorted(getfre_data.keys()):
                    gf = getfre_data[numa_id]
                    if 'error' not in gf:
                        core_stats = gf.get('core_stats', {})
                        if core_stats:
                            core_data = {
                                'Core ID': sorted(core_stats.keys()),
                                'Avg Frequency (MHz)': [core_stats[c]['avg'] for c in sorted(core_stats.keys())],
                                'Min Frequency (MHz)': [core_stats[c]['min'] for c in sorted(core_stats.keys())],
                                'Max Frequency (MHz)': [core_stats[c]['max'] for c in sorted(core_stats.keys())],
                                'Sample Count': [core_stats[c]['count'] for c in sorted(core_stats.keys())]
                            }
                            pd.DataFrame(core_data).to_excel(writer,
                                sheet_name=f'Getfre_NUMA{numa_id}', index=False)

            # ========== Sheet 12: Raw VM Data Time Series ==========
            if monitor.data:
                raw_data = {
                    'Timestamp': [d['timestamp'] for d in monitor.data],
                    'VM Name': [d['vm_name'] for d in monitor.data],
                    'PID': [d['pid'] for d in monitor.data],
                    'CPU (%)': [d['cpu_percent'] for d in monitor.data],
                    'Memory (MB)': [d['memory_mb'] for d in monitor.data],
                    'Hugepage (MB)': [d.get('memory_huge_mb', 0) for d in monitor.data]
                }
                pd.DataFrame(raw_data).to_excel(writer, sheet_name='Raw_VM_Data', index=False)

        # ========== Add Charts (using openpyxl directly) ==========
        try:
            from openpyxl.chart import PieChart, LineChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            from openpyxl import load_workbook

            wb = load_workbook(output_file)

            # Chart 1: Top-down Pie Chart (four major categories)
            if 'DevKit_TopDown' in wb.sheetnames:
                ws = wb['DevKit_TopDown']
                # Find the rows for the four categories (rows 4-7: Bad Speculation, Frontend Bound, Retiring, Backend Bound)
                pie = PieChart()
                pie.title = "CPU Top-down Analysis"
                labels = Reference(ws, min_col=1, min_row=4, max_row=7)  # Metric names
                data = Reference(ws, min_col=2, min_row=3, max_row=7)   # Values (include header)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.width = 15
                pie.height = 10
                ws.add_chart(pie, "D2")

            # Chart 2: IPC Timeline Line Chart
            if 'TopDown_Timeline' in wb.sheetnames:
                ws = wb['TopDown_Timeline']
                if ws.max_row > 1:  # Has data
                    line = LineChart()
                    line.title = "IPC Over Time"
                    line.style = 10
                    line.y_axis.title = "IPC"
                    line.x_axis.title = "Time"
                    line.width = 18
                    line.height = 8

                    # Data: IPC column (column 2)
                    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
                    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # Timestamps
                    line.add_data(data, titles_from_data=True)
                    line.set_categories(cats)
                    ws.add_chart(line, "E2")

            # Chart 3: Memory Bound Breakdown Bar Chart
            if 'DevKit_TopDown' in wb.sheetnames:
                ws = wb['DevKit_TopDown']
                bar = BarChart()
                bar.title = "Memory Bound Breakdown"
                bar.style = 10
                bar.y_axis.title = "Percentage (%)"
                bar.width = 12
                bar.height = 8

                # Rows 8-11: L3 Bound, Mem Bound, Latency bound, Bandwidth bound
                data = Reference(ws, min_col=2, min_row=8, max_row=11)
                cats = Reference(ws, min_col=1, min_row=8, max_row=11)
                bar.add_data(data)
                bar.set_categories(cats)
                bar.shape = 4
                ws.add_chart(bar, "D14")

            # Chart 4: DDR Bandwidth Timeline
            if 'Memory_Timeline' in wb.sheetnames:
                ws = wb['Memory_Timeline']
                if ws.max_row > 1:
                    line2 = LineChart()
                    line2.title = "DDR Bandwidth Over Time"
                    line2.style = 13
                    line2.y_axis.title = "MB/s"
                    line2.x_axis.title = "Time"
                    line2.width = 18
                    line2.height = 8

                    # DDR Write and Read columns (columns 6 and 7)
                    data = Reference(ws, min_col=6, min_row=1, max_col=7, max_row=ws.max_row)
                    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                    line2.add_data(data, titles_from_data=True)
                    line2.set_categories(cats)
                    ws.add_chart(line2, "I2")

            # Chart 5: Cache Miss Comparison Bar Chart
            if 'DevKit_Memory' in wb.sheetnames:
                ws = wb['DevKit_Memory']
                bar2 = BarChart()
                bar2.title = "Cache Miss Rate Comparison"
                bar2.style = 10
                bar2.y_axis.title = "Miss Rate (%)"
                bar2.width = 12
                bar2.height = 8

                # Rows 2-5: L1D, L1I, L2D, L2I miss
                data = Reference(ws, min_col=2, min_row=2, max_row=5)
                cats = Reference(ws, min_col=1, min_row=2, max_row=5)
                bar2.add_data(data)
                bar2.set_categories(cats)
                ws.add_chart(bar2, "D2")

            wb.save(output_file)
            print(f"[OK] Charts added to Excel report")

        except Exception as e:
            print(f"[WARN] Chart generation failed (non-critical): {e}")

        print(f"[OK] Excel report exported: {output_file}")
        return output_file

    except ImportError:
        print("[WARN] openpyxl not available, skipping Excel export")
        print("  Install with: pip install openpyxl")
        return None
    except Exception as e:
        print(f"[WARN] Excel export failed: {e}")
        return None


def print_capture_summary(results: dict, log_dir: str, numa_nodes: list = None):
    """Print log collection summary at the end of monitoring

    Args:
        results: capture results from LogCapture.get_results()
        log_dir: log directory path
        numa_nodes: list of NUMA nodes for filtering devkit_mem bandwidth
    """
    print("\n" + "=" * 70)
    print("Log Collection Summary")
    print("=" * 70)

    print(f"Duration: {results['duration']}s")
    print(f"Output directory: {log_dir}/")

    # Success tools
    if results['success']:
        print(f"\n[OK] Success ({len(results['success'])} tools):")
        for tool in results['success']:
            log_file = results['log_files'].get(tool, '')
            size = 0
            if os.path.exists(log_file):
                size = os.path.getsize(log_file)
            print(f"   {tool}: {log_file} ({size} bytes)")

    # Startup failures
    if results['failed_startup']:
        print(f"\n[ERROR] Failed to start ({len(results['failed_startup'])} tools):")
        for tool in results['failed_startup']:
            print(f"   {tool}: path not configured or invalid")

    # Runtime failures
    if results['failed_runtime']:
        print(f"\n[ERROR] Failed during runtime ({len(results['failed_runtime'])} tools):")
        for fail in results['failed_runtime']:
            tool = fail['tool']
            returncode = fail.get('returncode', 'unknown')
            print(f"   {tool}: returncode={returncode}")

    # List all log files
    print(f"\nLog files in {log_dir}:")
    for tool, path in results['log_files'].items():
        if os.path.exists(path):
            size = os.path.getsize(path)
            print(f"   {os.path.basename(path)} ({size} bytes)")
        else:
            print(f"   {os.path.basename(path)} (not created)")

    # Parse and display log contents
    print("\n" + "=" * 70)
    print("Parsed Metrics Summary")
    print("=" * 70)

    parsed = parse_all_logs(log_dir, numa_nodes)

    # devkit_top_down results
    if 'devkit_top_down' in parsed and 'error' not in parsed['devkit_top_down']:
        td = parsed['devkit_top_down']
        print(f"\n[DevKit Top-Down] ({td.get('report_count', 0)} reports)")
        print(f"  Cycles Avg:       {td.get('cycles_avg', 0):,.0f}")
        print(f"  Instructions Avg: {td.get('instructions_avg', 0):,.0f}")
        print(f"  IPC Avg:          {td.get('ipc_avg', 0):.2f}")
        print(f"  Top-down Metrics:")
        print(f"    Bad Speculation:  {td.get('bad_speculation_avg', 0):.2f}%")
        print(f"    Frontend Bound:   {td.get('frontend_bound_avg', 0):.2f}%")
        print(f"    Retiring:         {td.get('retiring_avg', 0):.2f}%")
        print(f"    Backend Bound:    {td.get('backend_bound_avg', 0):.2f}%")
        print(f"  Memory Bound Details:")
        print(f"    L3 Bound:         {td.get('l3_bound_avg', 0):.2f}%")
        print(f"    Mem Bound:        {td.get('mem_bound_avg', 0):.2f}%")
        print(f"      Latency bound:    {td.get('mem_latency_bound_avg', 0):.2f}%")
        print(f"      Bandwidth bound:  {td.get('mem_bandwidth_bound_avg', 0):.2f}%")

    # devkit_mem results
    if 'devkit_mem' in parsed and 'error' not in parsed['devkit_mem']:
        mem = parsed['devkit_mem']
        print(f"\n[DevKit Memory] ({mem.get('report_count', 0)} reports)")
        cm = mem.get('cache_miss', {})
        print(f"  Cache Miss %:")
        print(f"    L1D: {cm.get('L1D', 0):.2f}%")
        print(f"    L1I: {cm.get('L1I', 0):.2f}%")
        print(f"    L2D: {cm.get('L2D', 0):.2f}%")
        print(f"    L2I: {cm.get('L2I', 0):.2f}%")
        ddr = mem.get('ddr_bandwidth_system', {})
        print(f"  DDR Bandwidth (system):")
        print(f"    Write: {ddr.get('write', 0):.2f} MB/s")
        print(f"    Read:  {ddr.get('read', 0):.2f} MB/s")
        numa_bw = mem.get('numa_bandwidth', {})
        if numa_bw:
            print(f"  NUMA Bandwidth (filtered):")
            for node_id in sorted(numa_bw.keys()):
                bw = numa_bw[node_id]
                print(f"    NUMA {node_id}: Read {bw.get('read', 0):.2f} MB/s | Write {bw.get('write', 0):.2f} MB/s")

    # ksys results
    if 'ksys' in parsed and 'error' not in parsed['ksys']:
        ksys = parsed['ksys']
        print(f"\n[KSys]")
        l2 = ksys.get('l2_miss_latency', {})
        l3 = ksys.get('l3_miss_latency', {})
        if l2:
            print(f"  L2 Miss Latency: max={l2.get('cycles_max', 0)} min={l2.get('cycles_min', 0)} avg={l2.get('cycles_avg', 0)} cycles")
        if l3:
            print(f"  L3 Miss Latency: max={l3.get('cycles_max', 0)} min={l3.get('cycles_min', 0)} avg={l3.get('cycles_avg', 0)} cycles")
        if ksys.get('ipc'):
            print(f"  IPC: {ksys.get('ipc', 0):.2f}")
        td = ksys.get('topdown', {})
        if td:
            print(f"  Topdown:")
            if td.get('retiring'):
                print(f"    Retiring:        {td['retiring']:.2f}%")
            if td.get('frontend_bound'):
                print(f"    Frontend Bound:  {td['frontend_bound']:.2f}%")
            if td.get('bad_speculation'):
                print(f"    Bad Speculation: {td['bad_speculation']:.2f}%")
            if td.get('backend_bound'):
                print(f"    Backend Bound:   {td['backend_bound']:.2f}%")

    # ub_watch results
    if 'ub_watch' in parsed and 'error' not in parsed['ub_watch']:
        ub = parsed['ub_watch']
        print(f"\n[UB Watch]")
        lat = ub.get('latency', {})
        if lat:
            print(f"  Latency ({lat.get('path', 'N/A')}):")
            print(f"    Avg: R={lat.get('avg_r', 0)}ns W={lat.get('avg_w', 0)}ns")
            print(f"    Min: R={lat.get('min_r', 0)}ns W={lat.get('min_w', 0)}ns")
            print(f"    Max: R={lat.get('max_r', 0)}ns W={lat.get('max_w', 0)}ns")
        bw_list = ub.get('bandwidth', [])
        if bw_list:
            print(f"  Bandwidth (non-zero only):")
            for bw in bw_list:
                if bw.get('avg_sum', 0) > 0:
                    print(f"    Chip {bw.get('chip', 0)} Ports {bw.get('ports', 'N/A')}: "
                          f"Wr={bw.get('avg_wr', 0):.2f} Rd={bw.get('avg_rd', 0):.2f} Sum={bw.get('avg_sum', 0):.2f} MB/s")

    # getfre results
    if 'getfre' in parsed:
        getfre_data = parsed['getfre']
        print(f"\n[getfre Core Frequency]")
        for numa_id in sorted(getfre_data.keys()):
            gf = getfre_data[numa_id]
            if 'error' not in gf:
                print(f"  NUMA {numa_id}: Avg={gf.get('numa_avg', 0):.0f} MHz "
                      f"(Min={gf.get('numa_min', 0):.0f}, Max={gf.get('numa_max', 0):.0f}), "
                      f"{gf.get('sample_count', 0)} samples")
                # Show top 5 cores by frequency variance if available
                core_stats = gf.get('core_stats', {})
                if core_stats:
                    sorted_cores = sorted(core_stats.items(),
                                          key=lambda x: x[1]['max'] - x[1]['min'],
                                          reverse=True)[:5]
                    if sorted_cores:
                        print(f"    Top 5 variable cores:")
                        for core_id, stats in sorted_cores:
                            print(f"      Core {core_id}: Avg={stats['avg']:.0f} "
                                  f"(Min={stats['min']:.0f}, Max={stats['max']:.0f}) MHz")

    print("=" * 70)