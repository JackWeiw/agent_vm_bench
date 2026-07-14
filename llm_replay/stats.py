#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
stats.py - Async statistics collector for LLM Replay Server.

Provides high-performance statistics collection with batch writing
to avoid blocking the request handling loop.
"""

import asyncio
import json
import os
import time
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class RequestRecord:
    """Record of a single request."""

    timestamp: str  # ISO format
    session_name: str
    turn_index: int
    llm_duration_ms: int
    actual_delay_ms: int
    response_type: str  # "completion" or "stream"
    client_id: Optional[str] = None
    is_out_of_range: bool = False
    error: Optional[str] = None


@dataclass
class SessionStats:
    """Statistics for a single session."""

    name: str
    total_requests: int = 0
    turns_served: int = 0
    out_of_range_requests: int = 0
    llm_time_simulated_ms: int = 0
    errors: int = 0

    # Per-turn distribution
    turn_distribution: Dict[int, int] = field(default_factory=dict)  # turn_index -> count


@dataclass
class GlobalStats:
    """Global statistics across all sessions."""

    start_time: datetime
    sessions: Dict[str, SessionStats] = field(default_factory=dict)

    # Global counters
    total_requests: int = 0
    total_completions: int = 0
    total_streams: int = 0
    total_errors: int = 0
    total_out_of_range: int = 0
    total_llm_time_ms: int = 0

    # Connection tracking
    active_connections: int = 0
    total_connections: int = 0

    def uptime_seconds(self) -> float:
        """Get uptime in seconds."""
        return (datetime.now() - self.start_time).total_seconds()

    def requests_per_second(self) -> float:
        """Get requests per second."""
        uptime = self.uptime_seconds()
        if uptime == 0:
            return 0.0
        return self.total_requests / uptime


class StatsCollector:
    """
    Async statistics collector.

    Collects stats in-memory and writes to disk periodically.
    """

    def __init__(
        self,
        export_dir: str = "./results/llm_replay",
        export_format: str = "excel",
        batch_write_interval: int = 10,
        per_connection_stats: bool = True,
    ):
        self.export_dir = Path(export_dir)
        self.export_format = export_format
        self.batch_write_interval = batch_write_interval
        self.per_connection_stats = per_connection_stats

        # In-memory stats
        self.global_stats = GlobalStats(start_time=datetime.now())
        self.request_records: List[RequestRecord] = []

        # Async state
        self._lock = asyncio.Lock()
        self._write_task: Optional[asyncio.Task] = None
        self._running = False

    async def start(self):
        """Start the collector and periodic write task."""
        self._running = True
        self.export_dir.mkdir(parents=True, exist_ok=True)

        # Start periodic write task
        self._write_task = asyncio.create_task(self._periodic_write())

    async def stop(self):
        """Stop the collector and final write."""
        self._running = False

        if self._write_task:
            self._write_task.cancel()
            try:
                await self._write_task
            except asyncio.CancelledError:
                pass

        # Final write
        await self._write_stats()

    async def record_request(
        self,
        session_name: str,
        turn_index: int,
        llm_duration_ms: int,
        actual_delay_ms: int,
        response_type: str,
        client_id: Optional[str] = None,
        is_out_of_range: bool = False,
        error: Optional[str] = None,
    ):
        """
        Record a request.

        This is called from the request handler.
        Uses async lock for thread safety.
        """
        record = RequestRecord(
            timestamp=datetime.now().isoformat(),
            session_name=session_name,
            turn_index=turn_index,
            llm_duration_ms=llm_duration_ms,
            actual_delay_ms=actual_delay_ms,
            response_type=response_type,
            client_id=client_id,
            is_out_of_range=is_out_of_range,
            error=error,
        )

        async with self._lock:
            # Add to records
            self.request_records.append(record)

            # Update global stats
            self.global_stats.total_requests += 1
            self.global_stats.total_llm_time_ms += llm_duration_ms

            if response_type == "completion":
                self.global_stats.total_completions += 1
            else:
                self.global_stats.total_streams += 1

            if is_out_of_range:
                self.global_stats.total_out_of_range += 1

            if error:
                self.global_stats.total_errors += 1

            # Update session stats
            if session_name not in self.global_stats.sessions:
                self.global_stats.sessions[session_name] = SessionStats(name=session_name)

            session_stats = self.global_stats.sessions[session_name]
            session_stats.total_requests += 1
            session_stats.llm_time_simulated_ms += llm_duration_ms

            if not is_out_of_range and not error:
                session_stats.turns_served += 1
                session_stats.turn_distribution[turn_index] = session_stats.turn_distribution.get(turn_index, 0) + 1

            if is_out_of_range:
                session_stats.out_of_range_requests += 1

            if error:
                session_stats.errors += 1

    async def record_connection_open(self):
        """Record a new connection."""
        async with self._lock:
            self.global_stats.total_connections += 1
            self.global_stats.active_connections += 1

    async def record_connection_close(self):
        """Record a connection close."""
        async with self._lock:
            # Guard against negative counter (race condition safety)
            if self.global_stats.active_connections > 0:
                self.global_stats.active_connections -= 1

    def get_summary(self) -> Dict[str, Any]:
        """Get current summary (no lock needed for read)."""
        return {
            "uptime_s": self.global_stats.uptime_seconds(),
            "requests_per_second": self.global_stats.requests_per_second(),
            "total_requests": self.global_stats.total_requests,
            "total_completions": self.global_stats.total_completions,
            "total_streams": self.global_stats.total_streams,
            "total_errors": self.global_stats.total_errors,
            "total_out_of_range": self.global_stats.total_out_of_range,
            "total_llm_time_s": self.global_stats.total_llm_time_ms / 1000,
            "active_connections": self.global_stats.active_connections,
            "total_connections": self.global_stats.total_connections,
            "sessions": {
                name: {
                    "total_requests": s.total_requests,
                    "turns_served": s.turns_served,
                    "llm_time_s": s.llm_time_simulated_ms / 1000,
                }
                for name, s in self.global_stats.sessions.items()
            },
        }

    async def _periodic_write(self):
        """Periodic write task."""
        while self._running:
            await asyncio.sleep(self.batch_write_interval)
            await self._write_stats()

    async def _write_stats(self):
        """Write stats to disk."""
        async with self._lock:
            if not self.request_records:
                return

            # Generate timestamp for filename
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Write JSON
            if self.export_format in ("json", "both"):
                await self._write_json(ts)

            # Write Excel
            if self.export_format in ("excel", "both") and HAS_OPENPYXL:
                await self._write_excel(ts)

            # Clear records after write
            self.request_records = []

    async def _write_json(self, ts: str):
        """Write stats to JSON file."""
        filename = self.export_dir / f"llm_replay_stats_{ts}.json"

        data = {
            "summary": self.get_summary(),
            "records": [asdict(r) for r in self.request_records],
        }

        # Write async (use asyncio.to_thread for file I/O)
        def _write():
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        await asyncio.to_thread(_write)

    async def _write_excel(self, ts: str):
        """Write stats to Excel file."""
        filename = self.export_dir / f"llm_replay_stats_{ts}.xlsx"

        def _write():
            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            summary = self.get_summary()
            headers = ["Metric", "Value"]
            ws_summary.append(headers)

            for key, value in summary.items():
                if key != "sessions":
                    ws_summary.append([key, value])

            # Sessions sheet
            ws_sessions = wb.create_sheet("Sessions")
            ws_sessions.append(["Session", "Requests", "Turns Served", "LLM Time (s)"])

            for name, s in summary.get("sessions", {}).items():
                ws_sessions.append([name, s["total_requests"], s["turns_served"], s["llm_time_s"]])

            # Requests sheet
            ws_requests = wb.create_sheet("Requests")
            ws_requests.append(
                [
                    "Timestamp",
                    "Session",
                    "Turn",
                    "LLM Duration (ms)",
                    "Actual Delay (ms)",
                    "Type",
                    "Client",
                    "Out of Range",
                    "Error",
                ]
            )

            for r in self.request_records:
                ws_requests.append(
                    [
                        r.timestamp,
                        r.session_name,
                        r.turn_index,
                        r.llm_duration_ms,
                        r.actual_delay_ms,
                        r.response_type,
                        r.client_id or "-",
                        r.is_out_of_range,
                        r.error or "-",
                    ]
                )

            wb.save(filename)

        await asyncio.to_thread(_write)
