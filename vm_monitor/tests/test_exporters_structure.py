"""Structural snapshot test for vm_monitor/exporters.py export_to_excel.

Locks the set of sheets, their column headers, row counts, and chart count
produced from a fully-populated monitor + parsed_logs fixture. Purpose: guard
the mechanical refactor of export_to_excel (extraction into _build_* helpers)
against any structural drift -- a dropped/renamed sheet or column, a lost
chart, or a changed row count will fail this test.
"""

import os
import tempfile
import unittest
from unittest.mock import patch

import pandas as pd

from vm_monitor.base import VMMonitorBase
from vm_monitor.exporters import PANDAS_AVAILABLE, export_to_excel
from vm_monitor.parsers import parse_devkit_mem

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

# Real devkit_mem.log ships in the repo (log_example/) -- drive the real parser.
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_DEVKIT_MEM_LOG = os.path.join(_REPO_ROOT, "log_example", "devkit_mem.log")


class DummyMonitor(VMMonitorBase):
    def get_vms_realtime(self):
        return []

    def get_process_names(self):
        return ("test_process",)

    def extract_vm_id(self, pid, cmdline):
        return "vm0"

    def get_monitor_title(self):
        return "DummyMonitor"

    def get_no_vm_message(self):
        return "No VMs detected"

    def get_csv_filename_prefix(self):
        return "dummy_monitor"


def _full_monitor():
    """A monitor populated across every history the exporter reads."""
    m = DummyMonitor()

    # Raw VM samples (2 VMs, 2 samples each) -> VM_Stats + Raw_VM_Data
    # Carries every field collect_sample() writes (base.py record schema),
    # incl. the per-VM private/heap/swapcache scalars + per-NUMA breakdowns
    # the legacy Raw_VM_Data sheet used to drop.
    m.data = [
        {
            "timestamp": "2026-08-14 10:00:00",
            "vm_name": "vm0",
            "pid": 100,
            "cpu_percent": 10.0,
            "memory_mb": 2048.0,
            "memory_huge_mb": 1024.0,
            "memory_private_mb": 1000.0,
            "memory_heap_mb": 50.0,
            "memory_swapcache_mb": 10.0,
            "memory_per_numa": {0: {"total_mb": 1000.0}, 5: {"total_mb": 1048.0}},
            "memory_swapcache_per_numa": {0: 4.0, 5: 6.0},
            "status": "running",
        },
        {
            "timestamp": "2026-08-14 10:00:05",
            "vm_name": "vm0",
            "pid": 100,
            "cpu_percent": 20.0,
            "memory_mb": 2100.0,
            "memory_huge_mb": 1024.0,
            "memory_private_mb": 1000.0,
            "memory_heap_mb": 50.0,
            "memory_swapcache_mb": 20.0,
            "memory_per_numa": {0: {"total_mb": 1000.0}, 5: {"total_mb": 1100.0}},
            "memory_swapcache_per_numa": {0: 8.0, 5: 12.0},
            "status": "running",
        },
        {
            "timestamp": "2026-08-14 10:00:00",
            "vm_name": "vm1",
            "pid": 101,
            "cpu_percent": 5.0,
            "memory_mb": 4096.0,
            "memory_huge_mb": 2048.0,
            "memory_private_mb": 2000.0,
            "memory_heap_mb": 100.0,
            "memory_swapcache_mb": 30.0,
            "memory_per_numa": {0: {"total_mb": 2000.0}, 5: {"total_mb": 2096.0}},
            "memory_swapcache_per_numa": {0: 10.0, 5: 20.0},
            "status": "running",
        },
    ]
    m.last_vm_count = 2
    m.peak_total_cpu = 30.0
    m.peak_total_memory_mb = 6144.0

    # Host stats -> Summary
    m.host_cpu_history = [40.0, 50.0, 60.0]
    m.host_mem_history = [{"used_mb": 8000.0, "total_mb": 16384.0}, {"used_mb": 9000.0, "total_mb": 16384.0}]
    m.peak_host_cpu = 60.0
    m.peak_host_mem_mb = 9000.0

    # Hugepage totals -> Summary
    m.hugepage_total_mb = 8192.0
    m.hugepage_used_history = [1024.0, 2048.0]
    m.peak_hugepage_used_mb = 2048.0

    # NUMA CPU + memory + hugepage -> NUMA_Overview
    m.numa_cpu_history[0] = [10.0, 20.0]
    m.numa_cpu_peak[0] = 20.0
    m.numa_cpu_history[5] = [30.0, 40.0]
    m.numa_cpu_peak[5] = 40.0

    def _node(nid, used, usage):
        return {
            "node": nid,
            "total": 8192.0,
            "used": used,
            "free": 8192.0 - used,
            "usage": usage,
            "total_mb": 8192.0,
            "used_mb": used,
            "free_mb": 8192.0 - used,
            "available_mb": 7000.0,
            "swap_cached_mb": 100.0,
            "anon_pages_mb": 500.0,
            # Reclaim-pressure signals collected per NUMA node per sample
            # (_NUMA_MEMINFO_FIELDS) but historically dropped from the xlsx.
            "active_mb": round(used * 0.6, 2),
            "inactive_mb": round(used * 0.3, 2),
            "file_pages_mb": round(used * 0.2, 2),
            "usage_pct": usage,
        }

    m.numa_memory_history = [
        {"ts": "2026-08-14 10:00:00", "nodes": [_node(0, 1000.0, 12.0), _node(5, 2000.0, 24.0)]},
        {"ts": "2026-08-14 10:00:05", "nodes": [_node(0, 1200.0, 14.0), _node(5, 2400.0, 29.0)]},
    ]
    m.hugepage_per_numa_history = [
        {
            "ts": "2026-08-14 10:00:00",
            "nodes": {
                0: {"total_mb": 4096.0, "used_mb": 1024.0, "usage_pct": 25.0, "free_mb": 3072.0},
                5: {"total_mb": 4096.0, "used_mb": 2048.0, "usage_pct": 50.0, "free_mb": 2048.0},
            },
        },
        {
            "ts": "2026-08-14 10:00:05",
            "nodes": {
                0: {"total_mb": 4096.0, "used_mb": 1024.0, "usage_pct": 25.0, "free_mb": 3072.0},
                5: {"total_mb": 4096.0, "used_mb": 2048.0, "usage_pct": 50.0, "free_mb": 2048.0},
            },
        },
    ]

    # Swap -> Summary + Swap_Timeline
    m.swap_history = [
        {
            "ts": "2026-08-14 10:00:00",
            "capacity": {"used_mb": 100.0, "free_mb": 1900.0, "total_mb": 2000.0},
            "cache": {"cached_mb": 50.0, "cached_ratio_pct": 2.5},
            "activity": {
                "swap_in_rate": 10.0,
                "swap_out_rate": 5.0,
                "pswpin_cumulative": 100,
                "pswpout_cumulative": 50,
                "pswpin_delta": 0,
                "pswpout_delta": 0,
            },
        },
        {
            "ts": "2026-08-14 10:00:05",
            "capacity": {"used_mb": 200.0, "free_mb": 1800.0, "total_mb": 2000.0},
            "cache": {"cached_mb": 60.0, "cached_ratio_pct": 3.0},
            "activity": {
                "swap_in_rate": 20.0,
                "swap_out_rate": 10.0,
                "pswpin_cumulative": 120,
                "pswpout_cumulative": 60,
                "pswpin_delta": 20,
                "pswpout_delta": 10,
            },
        },
    ]
    m.peak_swap_used_mb = 200.0
    m.peak_swap_cached_mb = 60.0

    # VM total memory -> VM_Total_Memory_Timeline
    # Carries swapcache_mb + swapcache_per_numa that collect_vm_total_memory
    # aggregates each sample but the sheet used to drop.
    m.vm_total_memory_history = [
        {
            "ts": "2026-08-14 10:00:00",
            "total_mb": 6144.0,
            "vm_count": 2,
            "per_numa": {0: 3072.0, 5: 3072.0},
            "swapcache_mb": 40.0,
            "swapcache_per_numa": {0: 14.0, 5: 26.0},
        },
        {
            "ts": "2026-08-14 10:00:05",
            "total_mb": 6200.0,
            "vm_count": 2,
            "per_numa": {0: 3100.0, 5: 3100.0},
            "swapcache_mb": 20.0,
            "swapcache_per_numa": {0: 8.0, 5: 12.0},
        },
    ]

    # Disk I/O (sda/sdb/sdc) + ublk -> Disk_IO_Timeline
    m.target_disks = ["sda", "sdb", "sdc"]

    def dr(r, w, util, ifl, q, ra, wa):
        return {
            "r_mb_s": r,
            "w_mb_s": w,
            "util_pct": util,
            "inflight": ifl,
            "r_mb": r,
            "w_mb": w,
            "avg_queue_depth": q,
            "read_await_ms": ra,
            "write_await_ms": wa,
        }

    m.disk_history = [
        {
            "ts": "2026-08-14 10:00:00",
            "disks": {
                "sda": dr(10.0, 20.0, 5.0, 1, 0.5, 2.0, 3.0),
                "sdb": dr(5.0, 15.0, 3.0, 2, 1.0, 4.0, 5.0),
                "sdc": dr(1.0, 2.0, 1.0, 0, 0.2, 1.0, 1.5),
            },
        },
        {
            "ts": "2026-08-14 10:00:05",
            "disks": {
                "sda": dr(12.0, 25.0, 6.0, 1, 0.8, 2.5, 3.5),
                "sdb": dr(6.0, 18.0, 4.0, 3, 1.2, 4.5, 5.5),
                "sdc": dr(2.0, 4.0, 2.0, 1, 0.3, 1.5, 2.0),
            },
        },
    ]
    m.peak_disk_write_mb_s = 25.0

    # Host mem detail -> Host_Mem_Timeline + Summary peaks
    m.host_mem_detail_history = [
        {"ts": "2026-08-14 10:00:00", "cached_mb": 1000.0, "buffers_mb": 50.0, "dirty_mb": 10.0, "writeback_mb": 5.0},
        {"ts": "2026-08-14 10:00:05", "cached_mb": 1100.0, "buffers_mb": 55.0, "dirty_mb": 20.0, "writeback_mb": 8.0},
    ]
    m.peak_dirty_mb = 20.0
    m.peak_writeback_mb = 8.0

    # ublk device count -> Disk_IO_Timeline column + Summary peak
    m.ublk_history = [
        {"ts": "2026-08-14 10:00:00", "ublk_devices": 3},
        {"ts": "2026-08-14 10:00:05", "ublk_devices": 4},
    ]
    m.peak_ublk_devices = 4

    # Host page-cache pressure -> Host_Pressure_Timeline + Summary peaks
    m.host_pressure_history = [
        {
            "ts": "2026-08-14 10:00:00",
            "page_scan_mib_s": 5.0,
            "page_reclaim_mib_s": 4.0,
            "file_refault_mib_s": 1.0,
            "anon_pages_mb": 2000.0,
            "file_cache_mb": 3000.0,
            "sreclaimable_mb": 100.0,
            "iowait_pct": 2.0,
            "procs_running": 4,
            "procs_blocked": 0,
        },
        {
            "ts": "2026-08-14 10:00:05",
            "page_scan_mib_s": 8.0,
            "page_reclaim_mib_s": 6.0,
            "file_refault_mib_s": 2.0,
            "anon_pages_mb": 2100.0,
            "file_cache_mb": 3100.0,
            "sreclaimable_mb": 110.0,
            "iowait_pct": 5.0,
            "procs_running": 6,
            "procs_blocked": 1,
        },
    ]
    m.peak_page_scan_mib_s = 8.0
    m.peak_page_reclaim_mib_s = 6.0
    m.peak_file_refault_mib_s = 2.0
    m.dirty_limit_mb = 500.0
    m.dirty_background_limit_mb = 250.0
    m._dirty_limits_read = True
    return m


def _full_parsed_logs():
    """parsed_logs fixture covering all collection-tool sheets.

    devkit_mem comes from the real parser against log_example/devkit_mem.log;
    the rest are hand-built to match each parser's return schema.
    """
    parsed = {}

    # --- devkit_mem (real parser) ---
    parsed["devkit_mem"] = parse_devkit_mem(_DEVKIT_MEM_LOG, [0])

    # --- devkit_top_down ---
    # Mirrors parse_devkit_top_down's real return: every metric carries
    # _avg/_max/_min (parser computes all three); the sheet historically
    # exported only _avg (+ ipc_max/min).
    parsed["devkit_top_down"] = {
        "report_count": 2,
        "cycles_avg": 1000000.0,
        "cycles_max": 1200000.0,
        "cycles_min": 800000.0,
        "instructions_avg": 500000.0,
        "instructions_max": 600000.0,
        "instructions_min": 400000.0,
        "ipc_avg": 0.5,
        "ipc_max": 0.6,
        "ipc_min": 0.4,
        "ipc": [0.5, 0.55],
        "bad_speculation_avg": 5.0,
        "bad_speculation_max": 6.0,
        "bad_speculation_min": 4.0,
        "frontend_bound_avg": 10.0,
        "frontend_bound_max": 11.0,
        "frontend_bound_min": 9.0,
        "retiring_avg": 20.0,
        "retiring_max": 22.0,
        "retiring_min": 18.0,
        "backend_bound_avg": 65.0,
        "backend_bound_max": 68.0,
        "backend_bound_min": 62.0,
        "l3_bound_avg": 15.0,
        "l3_bound_max": 17.0,
        "l3_bound_min": 13.0,
        "mem_bound_avg": 50.0,
        "mem_bound_max": 55.0,
        "mem_bound_min": 45.0,
        "mem_latency_bound_avg": 30.0,
        "mem_latency_bound_max": 33.0,
        "mem_latency_bound_min": 27.0,
        "mem_bandwidth_bound_avg": 20.0,
        "mem_bandwidth_bound_max": 23.0,
        "mem_bandwidth_bound_min": 17.0,
        "timestamps": ["2026-08-14 10:00:00", "2026-08-14 10:00:05"],
        "timeline": {
            "timestamp": ["2026-08-14 10:00:00", "2026-08-14 10:00:05"],
            "ipc": [0.5, 0.55],
            "bad_speculation": [5.0, 5.1],
            "frontend_bound": [10.0, 10.1],
            "retiring": [20.0, 20.1],
            "backend_bound": [65.0, 65.1],
            "l3_bound": [15.0, 15.1],
            "mem_bound": [50.0, 50.1],
            "mem_latency_bound": [30.0, 30.1],
            "mem_bandwidth_bound": [20.0, 20.1],
        },
    }

    # --- ksys ---
    parsed["ksys"] = {
        "l2_miss_latency": {"cycles_max": 100, "cycles_min": 10, "cycles_avg": 50},
        "l3_miss_latency": {"cycles_max": 500, "cycles_min": 50, "cycles_avg": 250},
        "ipc": 1.2,
        "topdown": {
            "retiring": 25.0,
            "frontend_bound": 15.0,
            "bad_speculation": 5.0,
            "backend_bound": 55.0,
        },
    }

    # --- ub_watch ---
    parsed["ub_watch"] = {
        "latency": {
            "path": "N0->N2",
            "samples": 1000,
            "avg_r": 100,
            "avg_w": 110,
            "min_r": 80,
            "min_w": 90,
            "max_r": 200,
            "max_w": 210,
        },
        "bandwidth": [
            {
                "chip": 0,
                "ports": "0-1",
                "avg_wr": 1000.0,
                "avg_rd": 2000.0,
                "avg_sum": 3000.0,
                "max_wr": 1500.0,
                "max_rd": 2500.0,
                "max_sum": 4000.0,
            },
            {
                "chip": 1,
                "ports": "2-3",
                "avg_wr": 1100.0,
                "avg_rd": 2100.0,
                "avg_sum": 3200.0,
                "max_wr": 1600.0,
                "max_rd": 2600.0,
                "max_sum": 4200.0,
            },
        ],
    }

    # --- smap_bw ---
    parsed["smap_bw"] = {
        "cycles": [
            {
                "cycle_no": 1,
                "total_pages": 1000,
                "duration": 1.0,
                "bandwidth_gb_s": 10.0,
                "directions": {(0, 5): 600, (5, 0): 400},
            },
            {
                "cycle_no": 2,
                "total_pages": 2000,
                "duration": 2.0,
                "bandwidth_gb_s": 20.0,
                "directions": {(0, 5): 1200, (5, 0): 800},
            },
        ],
        "summary": {
            "total_cycles": 2,
            "total_pages": 3000,
            "avg_bandwidth_gb_s": 15.0,
            "min_bandwidth_gb_s": 10.0,
            "max_bandwidth_gb_s": 20.0,
        },
        "all_directions": {(0, 5), (5, 0)},
    }

    # --- getfre (2 NUMA nodes, each with 2 cores) ---
    # Carries the `timeline` {timestamp: {core_id: freq_mhz}} that parse_getfre
    # returns but no sheet used to surface (no Getfre_Timeline sheet existed).
    parsed["getfre"] = {
        0: {
            "numa_avg": 2400.0,
            "numa_min": 2000.0,
            "numa_max": 2800.0,
            "sample_count": 10,
            "core_stats": {
                10: {"avg": 2400.0, "min": 2000.0, "max": 2800.0, "count": 5},
                11: {"avg": 2500.0, "min": 2100.0, "max": 2900.0, "count": 5},
            },
            "timeline": {
                "2026-08-14 10:00:00": {10: 2400, 11: 2500},
                "2026-08-14 10:00:05": {10: 2600, 11: 2700},
            },
        },
        1: {
            "numa_avg": 2300.0,
            "numa_min": 1900.0,
            "numa_max": 2700.0,
            "sample_count": 10,
            "core_stats": {20: {"avg": 2300.0, "min": 1900.0, "max": 2700.0, "count": 5}},
            "timeline": {
                "2026-08-14 10:00:00": {20: 2300},
                "2026-08-14 10:00:05": {20: 2500},
            },
        },
    }
    return parsed


def _snapshot(output_file):
    """Return (sheet_names, {sheet: [cols]}, {sheet: row_count}, chart_count)."""
    with pd.ExcelFile(output_file) as xl:
        sheet_names = list(xl.sheet_names)
        cols = {}
        rows = {}
        for s in sheet_names:
            df = pd.read_excel(xl, sheet_name=s)
            cols[s] = list(df.columns)
            rows[s] = len(df)
    chart_count = 0
    if load_workbook is not None:
        wb = load_workbook(output_file)
        for ws in wb.worksheets:
            chart_count += len(ws._charts)
        wb.close()
    return sheet_names, cols, rows, chart_count


def _charts_by_sheet(output_file):
    """Return ({sheet_name: chart_count}, {sheet_name: [series_count per chart]}).

    Companion to _snapshot for chart-shape assertions: the snapshot only carries
    a total chart count, but the dropped-metric charts land on specific sheets,
    so per-sheet counts + per-chart series counts pin them precisely.
    """
    by_sheet = {}
    series_by_sheet = {}
    wb = load_workbook(output_file)
    for ws in wb.worksheets:
        by_sheet[ws.title] = len(ws._charts)
        series_by_sheet[ws.title] = [len(c.series) for c in ws._charts]
    wb.close()
    return by_sheet, series_by_sheet


@unittest.skipUnless(PANDAS_AVAILABLE and load_workbook is not None, "pandas/openpyxl required")
class TestExportStructure(unittest.TestCase):
    """Snapshot of export_to_excel output, frozen to guard the refactor."""

    def setUp(self):
        self.monitor = _full_monitor()
        self.parsed_logs = _full_parsed_logs()
        self.log_dir = tempfile.mkdtemp(prefix="vm_monitor_struct_")
        self.output_file = os.path.join(self.log_dir, "analysis_report.xlsx")

    def tearDown(self):
        for f in os.listdir(self.log_dir):
            try:
                os.unlink(os.path.join(self.log_dir, f))
            except PermissionError:
                pass
        os.rmdir(self.log_dir)

    def _export(self):
        with patch("vm_monitor.exporters.parse_all_logs", return_value=self.parsed_logs):
            return export_to_excel(self.monitor, self.log_dir, numa_nodes=[0], output_file=self.output_file)

    def test_sheet_set_and_order(self):
        self._export()
        names, _cols, _rows, _charts = _snapshot(self.output_file)
        self.assertEqual(
            names,
            [
                "Summary",
                "NUMA_Overview",
                "VM_Stats",
                "DevKit_TopDown",
                "TopDown_Timeline",
                "DevKit_Memory",
                "NUMA_Bandwidth",
                "Memory_Timeline",
                "KSys",
                "UBWatch_Latency",
                "UBWatch_Bandwidth",
                "SMAPBW_Summary",
                "SMAPBW_Cycles",
                "Getfre_Summary",
                "Getfre_NUMA0",
                "Getfre_Timeline_NUMA0",
                "Getfre_NUMA1",
                "Getfre_Timeline_NUMA1",
                "Raw_VM_Data",
                "Swap_Timeline",
                "NUMA_Memory_Timeline",
                "VM_Total_Memory_Timeline",
                "Disk_IO_Timeline",
                "Host_Mem_Timeline",
                "Host_Pressure_Timeline",
            ],
        )

    def test_columns_and_row_counts(self):
        self._export()
        _names, cols, rows, _charts = _snapshot(self.output_file)

        expected_cols = {
            "Summary": ["Metric", "Value", "Unit"],
            "NUMA_Overview": [
                "NUMA Node",
                "Avg CPU (%)",
                "Peak CPU (%)",
                "Avg Used (MB)",
                "Peak Used (MB)",
                "Avg Usage (%)",
                "HP Avg Total (MB)",
                "HP Avg Used (MB)",
                "HP Avg Usage (%)",
            ],
            "VM_Stats": [
                "VM Name",
                "PID",
                "Samples",
                "Avg CPU (%)",
                "Max CPU (%)",
                "Avg Memory (MB)",
                "Max Memory (MB)",
                "Min Memory (MB)",
                "Last Memory (MB)",
                "Avg Hugepage (MB)",
                "Max Hugepage (MB)",
                "Avg Private (MB)",
                "Max Private (MB)",
                "Avg Heap (MB)",
                "Max Heap (MB)",
            ],
            "DevKit_TopDown": ["Metric", "Value", "Report Count"],
            "DevKit_Memory": ["Metric", "Value", "Report Count"],
            "NUMA_Bandwidth": ["NUMA Node", "Read (MB/s)", "Write (MB/s)"],
            "KSys": ["Metric", "Value"],
            "UBWatch_Latency": ["Metric", "Value"],
            "UBWatch_Bandwidth": [
                "Chip",
                "Ports",
                "Avg Write (MB/s)",
                "Avg Read (MB/s)",
                "Avg Sum (MB/s)",
                "Max Write (MB/s)",
                "Max Read (MB/s)",
                "Max Sum (MB/s)",
            ],
            "SMAPBW_Summary": ["Metric", "Value"],
            "SMAPBW_Cycles": ["Cycle", "Pages", "Duration (s)", "Bandwidth (GB/s)", "N0->N5_pages", "N5->N0_pages"],
            "Getfre_Summary": [
                "NUMA",
                "Avg Frequency (MHz)",
                "Min Frequency (MHz)",
                "Max Frequency (MHz)",
                "Sample Count",
                "Core Count",
            ],
            "Getfre_NUMA0": [
                "Core ID",
                "Avg Frequency (MHz)",
                "Min Frequency (MHz)",
                "Max Frequency (MHz)",
                "Sample Count",
            ],
            "Getfre_NUMA1": [
                "Core ID",
                "Avg Frequency (MHz)",
                "Min Frequency (MHz)",
                "Max Frequency (MHz)",
                "Sample Count",
            ],
            "Getfre_Timeline_NUMA0": ["Timestamp", "Core 10 (MHz)", "Core 11 (MHz)"],
            "Getfre_Timeline_NUMA1": ["Timestamp", "Core 20 (MHz)"],
            "Raw_VM_Data": [
                "Timestamp",
                "VM Name",
                "PID",
                "CPU (%)",
                "Memory (MB)",
                "Hugepage (MB)",
                "Private (MB)",
                "Heap (MB)",
                "SwapCache (MB)",
                "Status",
                "NUMA0 Memory (MB)",
                "NUMA5 Memory (MB)",
                "NUMA0 SwapCache (MB)",
                "NUMA5 SwapCache (MB)",
            ],
            "Disk_IO_Timeline": [
                "Timestamp",
                "sda Read (MB/s)",
                "sda Write (MB/s)",
                "sda Util (%)",
                "sda Inflight",
                "sda Queue Depth",
                "sda Read Await (ms)",
                "sda Write Await (ms)",
                "sdb Read (MB/s)",
                "sdb Write (MB/s)",
                "sdb Util (%)",
                "sdb Inflight",
                "sdb Queue Depth",
                "sdb Read Await (ms)",
                "sdb Write Await (ms)",
                "sdc Read (MB/s)",
                "sdc Write (MB/s)",
                "sdc Util (%)",
                "sdc Inflight",
                "sdc Queue Depth",
                "sdc Read Await (ms)",
                "sdc Write Await (ms)",
                "ublk Devices",
            ],
            "Host_Mem_Timeline": ["Timestamp", "Cached (MB)", "Buffers (MB)", "Dirty (MB)", "Writeback (MB)"],
            "Host_Pressure_Timeline": [
                "Timestamp",
                "Page Scan (MiB/s)",
                "Page Reclaim (MiB/s)",
                "File Refault (MiB/s)",
                "Anon Pages (MB)",
                "File Cache (MB)",
                "SReclaimable (MB)",
                "IOWait (%)",
                "Procs Running",
                "Procs Blocked",
            ],
            "VM_Total_Memory_Timeline": [
                "Timestamp",
                "VM Total Memory (MB)",
                "VM Count",
                "SwapCache (MB)",
                "NUMA0 VM Memory (MB)",
                "NUMA5 VM Memory (MB)",
                "NUMA0 SwapCache (MB)",
                "NUMA5 SwapCache (MB)",
            ],
            "NUMA_Memory_Timeline": [
                "Timestamp",
                "NUMA0 Total (MB)",
                "NUMA0 Used (MB)",
                "NUMA0 Free (MB)",
                "NUMA0 Available (MB)",
                "NUMA0 SwapCache (MB)",
                "NUMA0 AnonPages (MB)",
                "NUMA0 Active (MB)",
                "NUMA0 Inactive (MB)",
                "NUMA0 File Pages (MB)",
                "NUMA0 Usage (%)",
                # Remote borrowing node (remote_numa_id=5) is a focus node too,
                # so its reclaim-pressure fields land here as well.
                "NUMA5 Total (MB)",
                "NUMA5 Used (MB)",
                "NUMA5 Free (MB)",
                "NUMA5 Available (MB)",
                "NUMA5 SwapCache (MB)",
                "NUMA5 AnonPages (MB)",
                "NUMA5 Active (MB)",
                "NUMA5 Inactive (MB)",
                "NUMA5 File Pages (MB)",
                "NUMA5 Usage (%)",
            ],
        }
        for sheet, expected in expected_cols.items():
            self.assertEqual(cols[sheet], expected, f"columns mismatch for {sheet}")

        expected_rows = {
            "Summary": 39,  # meta(4)+host(4)+hugepage(4)+swap-cap(4)+swap-cache(3)+swap-act(6)+vm(6)+disk/mem(5)+pressure(3)
            "NUMA_Overview": 2,
            "VM_Stats": 2,
            "DevKit_TopDown": 33,  # 13 base (avg + ipc max/min) + 20 max/min for cycles/instructions + 8 topdown metrics
            "NUMA_Bandwidth": 1,  # devkit_mem.log NUMA bandwidth filtered to numa_nodes=[0]
            "KSys": 11,  # 3+3+1+4
            "UBWatch_Latency": 8,
            "UBWatch_Bandwidth": 2,
            "SMAPBW_Summary": 5,
            "SMAPBW_Cycles": 2,
            "Getfre_Summary": 2,
            "Getfre_NUMA0": 2,
            "Getfre_Timeline_NUMA0": 2,
            "Getfre_NUMA1": 1,
            "Getfre_Timeline_NUMA1": 2,
            "Raw_VM_Data": 3,
            "Swap_Timeline": 2,
            "NUMA_Memory_Timeline": 2,
            "VM_Total_Memory_Timeline": 2,
            "Disk_IO_Timeline": 2,
            "Host_Mem_Timeline": 2,
            "Host_Pressure_Timeline": 2,
        }
        for sheet, expected in expected_rows.items():
            self.assertEqual(rows[sheet], expected, f"row count mismatch for {sheet}")

    def test_chart_count(self):
        self._export()
        _names, _cols, _rows, charts = _snapshot(self.output_file)
        # 1 DevKit pie + 1 IPC line + 1 MemBound bar + 1 DDR line + 1 CacheMiss bar
        # + 2 Swap (in/out + SwapCache) + 3 NUMA_Memory_Timeline (8A + 8B + 8C reclaim)
        # + 1 VM_Total + 1 Disk Write line + 1 Dirty+Writeback line + 1 Host Pressure line
        # + 1 VM_Stats composition bar + 2 Getfre_Timeline (NUMA0 + NUMA1 freq lines)
        self.assertEqual(charts, 17)


@unittest.skipUnless(PANDAS_AVAILABLE and load_workbook is not None, "pandas/openpyxl required")
class TestDroppedMetricsExported(unittest.TestCase):
    """Value checks for metrics that were always collected (monitor histories
    or parsed logs) but never reached the xlsx. The exporter now surfaces them;
    these tests pin the values so a future refactor cannot silently drop them
    again. No new collection -- every value below originates in _full_monitor()
    / _full_parsed_logs(), which mirror the real base.py / parsers.py schemas.
    """

    def setUp(self):
        self.monitor = _full_monitor()
        self.parsed_logs = _full_parsed_logs()
        self.log_dir = tempfile.mkdtemp(prefix="vm_monitor_dropped_")
        self.output_file = os.path.join(self.log_dir, "analysis_report.xlsx")

    def tearDown(self):
        for f in os.listdir(self.log_dir):
            try:
                os.unlink(os.path.join(self.log_dir, f))
            except PermissionError:
                pass
        os.rmdir(self.log_dir)

    def _export(self):
        with patch("vm_monitor.exporters.parse_all_logs", return_value=self.parsed_logs):
            return export_to_excel(self.monitor, self.log_dir, numa_nodes=[0], output_file=self.output_file)

    def test_raw_vm_data_carries_dropped_scalars_and_per_numa(self):
        self._export()
        df = pd.read_excel(self.output_file, sheet_name="Raw_VM_Data")
        row0 = df.iloc[0]  # vm0 @ 10:00:00
        # Per-VM scalars the legacy 6-column sheet dropped
        self.assertAlmostEqual(row0["Private (MB)"], 1000.0)
        self.assertAlmostEqual(row0["Heap (MB)"], 50.0)
        self.assertAlmostEqual(row0["SwapCache (MB)"], 10.0)
        self.assertEqual(row0["Status"], "running")
        # Per-NUMA breakdown spread (vm0 sample1: NUMA0=1000, NUMA5=1048)
        self.assertAlmostEqual(row0["NUMA0 Memory (MB)"], 1000.0)
        self.assertAlmostEqual(row0["NUMA5 Memory (MB)"], 1048.0)
        self.assertAlmostEqual(row0["NUMA0 SwapCache (MB)"], 4.0)
        self.assertAlmostEqual(row0["NUMA5 SwapCache (MB)"], 6.0)

    def test_vm_stats_carries_min_last_huge_private_heap(self):
        self._export()
        df = pd.read_excel(self.output_file, sheet_name="VM_Stats").sort_values("VM Name").reset_index(drop=True)
        vm0 = df.iloc[0]  # memory [2048, 2100]
        self.assertAlmostEqual(vm0["Min Memory (MB)"], 2048.0)
        self.assertAlmostEqual(vm0["Last Memory (MB)"], 2100.0)
        self.assertAlmostEqual(vm0["Max Hugepage (MB)"], 1024.0)
        self.assertAlmostEqual(vm0["Avg Private (MB)"], 1000.0)
        self.assertAlmostEqual(vm0["Max Private (MB)"], 1000.0)
        self.assertAlmostEqual(vm0["Avg Heap (MB)"], 50.0)
        self.assertAlmostEqual(vm0["Max Heap (MB)"], 50.0)
        vm1 = df.iloc[1]  # memory [4096]
        self.assertAlmostEqual(vm1["Min Memory (MB)"], 4096.0)
        self.assertAlmostEqual(vm1["Last Memory (MB)"], 4096.0)
        self.assertAlmostEqual(vm1["Max Hugepage (MB)"], 2048.0)
        self.assertAlmostEqual(vm1["Max Private (MB)"], 2000.0)
        self.assertAlmostEqual(vm1["Max Heap (MB)"], 100.0)

    def test_vm_total_memory_timeline_carries_swapcache(self):
        self._export()
        df = pd.read_excel(self.output_file, sheet_name="VM_Total_Memory_Timeline")
        row0 = df.iloc[0]
        self.assertAlmostEqual(row0["SwapCache (MB)"], 40.0)
        self.assertAlmostEqual(row0["NUMA0 SwapCache (MB)"], 14.0)
        self.assertAlmostEqual(row0["NUMA5 SwapCache (MB)"], 26.0)

    def test_numa_memory_timeline_carries_active_inactive_filepages(self):
        self._export()
        df = pd.read_excel(self.output_file, sheet_name="NUMA_Memory_Timeline")
        row0 = df.iloc[0]  # node0 sample1 used=1000 -> active=600/inactive=300/file=200
        self.assertAlmostEqual(row0["NUMA0 Active (MB)"], 600.0)
        self.assertAlmostEqual(row0["NUMA0 Inactive (MB)"], 300.0)
        self.assertAlmostEqual(row0["NUMA0 File Pages (MB)"], 200.0)

    def test_summary_carries_vm_peak_total_memory(self):
        self._export()
        df = pd.read_excel(self.output_file, sheet_name="Summary")
        peaks = dict(zip(df["Metric"], df["Value"]))
        self.assertIn("VM Peak Total Memory", peaks)
        self.assertAlmostEqual(peaks["VM Peak Total Memory"], 6144.0)

    def test_devkit_topdown_carries_max_min_rows(self):
        self._export()
        df = pd.read_excel(self.output_file, sheet_name="DevKit_TopDown")
        m = dict(zip(df["Metric"], df["Value"]))
        self.assertAlmostEqual(m["Cycles Max"], 1200000.0)
        self.assertAlmostEqual(m["Cycles Min"], 800000.0)
        self.assertAlmostEqual(m["Instructions Max"], 600000.0)
        self.assertAlmostEqual(m["Instructions Min"], 400000.0)
        self.assertAlmostEqual(m["Bad Speculation Max"], 6.0)
        self.assertAlmostEqual(m["Bad Speculation Min"], 4.0)
        self.assertAlmostEqual(m["Bandwidth Bound Max"], 23.0)
        self.assertAlmostEqual(m["Bandwidth Bound Min"], 17.0)

    def test_getfre_timeline_numa0_carries_per_core_freq(self):
        self._export()
        df = pd.read_excel(self.output_file, sheet_name="Getfre_Timeline_NUMA0")
        self.assertEqual(list(df.columns), ["Timestamp", "Core 10 (MHz)", "Core 11 (MHz)"])
        self.assertEqual(len(df), 2)
        row0 = df.iloc[0]  # 10:00:00
        self.assertAlmostEqual(row0["Core 10 (MHz)"], 2400)
        self.assertAlmostEqual(row0["Core 11 (MHz)"], 2500)
        row1 = df.iloc[1]  # 10:00:05
        self.assertAlmostEqual(row1["Core 10 (MHz)"], 2600)
        self.assertAlmostEqual(row1["Core 11 (MHz)"], 2700)

    def test_getfre_timeline_skipped_when_no_timeline(self):
        """A getfre NUMA without a `timeline` (legacy/parse-fail) still emits
        Getfre_NUMA{n} but omits the timeline sheet."""
        parsed = _full_parsed_logs()
        del parsed["getfre"][0]["timeline"]  # NUMA0 lacks timeline
        with patch("vm_monitor.exporters.parse_all_logs", return_value=parsed):
            export_to_excel(self.monitor, self.log_dir, numa_nodes=[0], output_file=self.output_file)
        with pd.ExcelFile(self.output_file) as xl:
            names = set(xl.sheet_names)
        self.assertIn("Getfre_NUMA0", names)
        self.assertNotIn("Getfre_Timeline_NUMA0", names)
        # NUMA1 still has its timeline -> its sheet still appears
        self.assertIn("Getfre_Timeline_NUMA1", names)


@unittest.skipUnless(PANDAS_AVAILABLE and load_workbook is not None, "pandas/openpyxl required")
class TestDroppedMetricCharts(unittest.TestCase):
    """Charts for the metrics surfaced in the dropped-metrics PR. The data
    already lands in the xlsx (TestDroppedMetricsExported pins the values);
    these tests pin that each high-value dataset also gets a chart so it is
    visually consumable, not just tabular.
    """

    def setUp(self):
        self.monitor = _full_monitor()
        self.parsed_logs = _full_parsed_logs()
        self.log_dir = tempfile.mkdtemp(prefix="vm_monitor_charts_")
        self.output_file = os.path.join(self.log_dir, "analysis_report.xlsx")

    def tearDown(self):
        for f in os.listdir(self.log_dir):
            try:
                os.unlink(os.path.join(self.log_dir, f))
            except PermissionError:
                pass
        os.rmdir(self.log_dir)

    def _export(self):
        with patch("vm_monitor.exporters.parse_all_logs", return_value=self.parsed_logs):
            return export_to_excel(self.monitor, self.log_dir, numa_nodes=[0], output_file=self.output_file)

    def test_getfre_timeline_sheets_have_frequency_chart(self):
        """Each Getfre_Timeline_NUMA{n} sheet gets a per-core frequency line;
        series count tracks the number of cores in that NUMA's timeline."""
        self._export()
        by_sheet, series = _charts_by_sheet(self.output_file)
        self.assertEqual(by_sheet.get("Getfre_Timeline_NUMA0"), 1)
        self.assertEqual(by_sheet.get("Getfre_Timeline_NUMA1"), 1)
        # NUMA0 timeline carries cores 10 + 11; NUMA1 carries core 20 only
        self.assertEqual(series["Getfre_Timeline_NUMA0"], [2])
        self.assertEqual(series["Getfre_Timeline_NUMA1"], [1])

    def test_numa_memory_timeline_has_reclaim_pressure_chart(self):
        """NUMA_Memory_Timeline gains a third chart (8C) for the
        Active/Inactive/File Pages reclaim-pressure signals."""
        self._export()
        by_sheet, _series = _charts_by_sheet(self.output_file)
        # 8A free/used + 8B swapcache/usage + 8C reclaim pressure
        self.assertEqual(by_sheet.get("NUMA_Memory_Timeline"), 3)

    def test_vm_stats_has_memory_composition_bar(self):
        """VM_Stats gains a grouped bar of the avg memory composition
        (Memory / Private / Heap / Hugepage) per VM."""
        self._export()
        by_sheet, series = _charts_by_sheet(self.output_file)
        self.assertEqual(by_sheet.get("VM_Stats"), 1)
        self.assertEqual(series["VM_Stats"], [4])

    def test_getfre_timeline_without_timeline_has_no_chart(self):
        """A NUMA whose timeline was deleted must not get a stray chart on a
        sheet that does not exist -- the chart helper skips absent sheets."""
        parsed = _full_parsed_logs()
        del parsed["getfre"][0]["timeline"]  # NUMA0 lacks timeline -> no sheet
        with patch("vm_monitor.exporters.parse_all_logs", return_value=parsed):
            export_to_excel(self.monitor, self.log_dir, numa_nodes=[0], output_file=self.output_file)
        by_sheet, _series = _charts_by_sheet(self.output_file)
        self.assertNotIn("Getfre_Timeline_NUMA0", by_sheet)
        # NUMA1 still charted
        self.assertEqual(by_sheet.get("Getfre_Timeline_NUMA1"), 1)


if __name__ == "__main__":
    unittest.main()
