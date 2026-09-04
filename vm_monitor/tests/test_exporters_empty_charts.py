"""Regression: empty charts (zero series) must not be embedded.

When a timeline sheet has timestamps but no metric columns (e.g. focus NUMA
nodes filter to none), the chart helpers used to call ``ws.add_chart()``
unconditionally. openpyxl then writes a ``chartN.xml`` with an empty
``<c:plotArea>`` (no ``<c:ser>`` children), which Excel rejects on open with a
"unreadable content" repair prompt. These tests pin the guard: an empty-data
sheet produces zero embedded charts, while a populated sheet still does.
"""
from __future__ import annotations

import os
import tempfile
import unittest
from zipfile import ZipFile

from openpyxl import Workbook

from vm_monitor.exporters import (
    _add_numa_memory_charts,
    _add_swap_inout_line,
    _add_swapcache_line,
)


def _chart_series_counts(xlsx_path: str) -> list[int]:
    """Per-chart count of ``<ser>`` series elements in the saved workbook's
    chart XMLs. openpyxl uses the default drawingml/chart namespace (no ``c:``
    prefix), so a series is a ``<ser>``/``<ser ...`` opening tag. A chart with
    zero series is the corruption signature Excel repairs away.
    """
    import re

    counts: list[int] = []
    with ZipFile(xlsx_path) as zf:
        for name in zf.namelist():
            if not name.startswith("xl/charts/chart") or not name.endswith(".xml"):
                continue
            xml = zf.read(name).decode("utf-8", "ignore")
            counts.append(len(re.findall(r"<ser[\s/>]", xml)))
    return counts


class TestEmptyChartsNotEmbedded(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp(prefix="vm_empty_charts_")
        self.out = os.path.join(self.tmp, "out.xlsx")

    def tearDown(self):
        for f in os.listdir(self.tmp):
            os.unlink(os.path.join(self.tmp, f))
        os.rmdir(self.tmp)

    def _save_and_check_series(self, wb: Workbook) -> list[int]:
        wb.save(self.out)
        return _chart_series_counts(self.out)

    def test_numa_memory_charts_skipped_when_only_timestamps(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "NUMA_Memory_Timeline"
        ws.append(["Timestamp"])
        for i in range(5):
            ws.append([1700000000 + i])  # timestamps but no Free/Used/... columns

        _add_numa_memory_charts(wb)

        # No series-bearing columns -> both 8A and 8B must be skipped.
        self.assertEqual(len(ws._charts), 0)
        self.assertEqual(self._save_and_check_series(wb), [])

    def test_numa_memory_charts_present_when_data_columns(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "NUMA_Memory_Timeline"
        ws.append(["Timestamp", "Node0 Free (MB)", "Node0 Usage (%)"])
        for i in range(5):
            ws.append([1700000000 + i, 1024 - i * 10, 40 + i])

        _add_numa_memory_charts(wb)

        # 8A (Free) + 8B (Usage%) both embed -> 2 charts, each with >=1 series.
        self.assertEqual(len(ws._charts), 2)
        counts = self._save_and_check_series(wb)
        self.assertEqual(len(counts), 2)
        self.assertTrue(all(c >= 1 for c in counts), f"empty series chart: {counts}")

    def test_swap_inout_chart_skipped_when_no_rate_columns(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Swap_Timeline"
        ws.append(["Timestamp", "Swap Cached (MB)"])  # no In/Out Rate columns
        for i in range(5):
            ws.append([1700000000 + i, 100 + i])

        _add_swap_inout_line(wb)

        self.assertEqual(len(ws._charts), 0)
        self.assertEqual(self._save_and_check_series(wb), [])

    def test_swapcache_chart_skipped_when_no_swapcache_columns(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Swap_Timeline"
        ws.append(["Timestamp", "Swap In Rate", "Swap Out Rate"])  # no Cached col
        for i in range(5):
            ws.append([1700000000 + i, i, i * 2])

        _add_swapcache_line(wb)

        self.assertEqual(len(ws._charts), 0)
        self.assertEqual(self._save_and_check_series(wb), [])

    def test_swapcache_chart_uses_only_swapcache_columns(self):
        """Non-contiguous SwapCache columns must not pull in intervening columns."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Swap_Timeline"
        # col2=SwapCached, col3=intervening (In Rate), col4=NUMA SwapCache
        ws.append(["Timestamp", "SwapCached (MB)", "Swap In Rate", "Node0 SwapCache (MB)"])
        for i in range(5):
            ws.append([1700000000 + i, 100, i, 200])

        _add_swapcache_line(wb)

        self.assertEqual(len(ws._charts), 1)
        # Two SwapCache columns -> two series, NOT three (In Rate must be excluded).
        self.assertEqual(len(ws._charts[0].series), 2)
        wb.save(self.out)
        # Round-trips without Excel-repairable empty charts.
        self.assertTrue(all(c >= 1 for c in _chart_series_counts(self.out)))


if __name__ == "__main__":
    unittest.main()
