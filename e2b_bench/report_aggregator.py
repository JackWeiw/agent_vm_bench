"""
Report Aggregator Module

Aggregates metrics from multiple batch test tasks into a single Excel report.
Supports styled output with data source grouping and merged cells.
"""

import os
from datetime import datetime
from typing import Any, Dict, List


class ReportAggregator:
    """Aggregate batch test results into Excel report"""

    COLUMN_GROUPS = {
        "Basic": ["task_id", "total_count", "ratio", "benchmark_percent"],
        "Browser": ["Browser_Success_Rate", "Browser_Avg_Latency_ms", "Browser_P99_Latency_ms", "Browser_Total_Tasks"],
        "VM_CPU": ["VM_CPU_Mean", "VM_CPU_Max"],
        "DevKit_TopDown": [
            "DevKit_TopDown_Cycles_Avg",
            "DevKit_TopDown_Instructions_Avg",
            "DevKit_TopDown_IPC_Avg",
            "DevKit_TopDown_IPC_Max",
            "DevKit_TopDown_IPC_Min",
            "DevKit_TopDown_Bad_Speculation",
            "DevKit_TopDown_Frontend_Bound",
            "DevKit_TopDown_Retiring",
            "DevKit_TopDown_Backend_Bound",
            "DevKit_TopDown_L3_Bound",
            "DevKit_TopDown_Mem_Bound",
            "DevKit_TopDown_Latency_Bound",
            "DevKit_TopDown_Bandwidth_Bound",
        ],
        "DevKit_Memory": [
            "DevKit_Memory_L1D_Miss",
            "DevKit_Memory_L1I_Miss",
            "DevKit_Memory_L2D_Miss",
            "DevKit_Memory_L2I_Miss",
            "DevKit_Memory_DDR_Read",
            "DevKit_Memory_DDR_Write",
        ],
    }

    # Source colors (hex without # for openpyxl)
    SOURCE_COLORS = {
        "Basic": "4472C4",  # Blue
        "Browser": "70AD47",  # Green
        "VM_CPU": "FFC000",  # Orange
        "DevKit_TopDown": "ED7D31",  # Dark Orange
        "DevKit_Memory": "A5A5A5",  # Gray
        "NUMA_Bandwidth": "5B9BD5",  # Light Blue
        "KSys": "7030A0",  # Purple
        "UBWatch_Latency": "C55A11",  # Brown
        "UBWatch_Bandwidth": "00B050",  # Dark Green
        "SMAPBW": "FF6B6B",  # Red/Pink
        "Getfre": "00B0F0",  # Cyan
    }

    def __init__(self, output_dir: str = "results/e2b/batch"):
        self.output_dir = output_dir

    def aggregate(self, metrics_data: List[Dict[str, Any]], output_filename: str = None) -> str:
        """
        Aggregate all test metrics into Excel report

        Args:
            metrics_data: List of dicts, each containing task_id and metrics
            output_filename: Optional custom filename

        Returns:
            Path to generated Excel file
        """
        if not metrics_data:
            print("[ReportAggregator] No metrics data to aggregate")
            return ""

        df = self._build_dataframe(metrics_data)
        df = df.sort_values(by=["total_count", "ratio", "benchmark_percent"])

        os.makedirs(self.output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = output_filename or f"e2b_batch_summary_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)

        self._export_excel(df, output_path)

        print(f"[ReportAggregator] Report saved to: {output_path}")
        return output_path

    def _build_dataframe(self, metrics_data: List[Dict[str, Any]]) -> "pd.DataFrame":
        """Build DataFrame from metrics list"""
        import pandas as pd

        rows = []
        for m in metrics_data:
            row = {
                "task_id": m.get("task_id", ""),
                "total_count": m.get("total_count", 0),
                "ratio": m.get("ratio", 0),
                "benchmark_percent": m.get("benchmark_percent", 0),
            }

            for key, value in m.items():
                if key not in row and key not in ["success", "error_msg", "result_dir"]:
                    row[key] = value

            rows.append(row)

        df = pd.DataFrame(rows)

        # Remove columns that are entirely empty/NaN (tools not enabled)
        # Keep Basic columns (task_id, total_count, ratio, benchmark_percent) always
        basic_cols = ["task_id", "total_count", "ratio", "benchmark_percent"]

        # Identify columns to drop: all values are NaN, None, 0, or empty string
        cols_to_drop = []
        for col in df.columns:
            if col in basic_cols:
                continue  # Always keep basic columns

            # Check if column is entirely empty
            if df[col].isna().all():
                cols_to_drop.append(col)
            elif (df[col].fillna(0) == 0).all() and df[col].dtype in ["float64", "int64"]:
                # All numeric values are 0 (tool not enabled or no data)
                cols_to_drop.append(col)
            elif (df[col].fillna("") == "").all() and df[col].dtype == "object":
                # All string values are empty
                cols_to_drop.append(col)

        if cols_to_drop:
            print(f"[ReportAggregator] Dropping empty columns (tools not enabled): {cols_to_drop}")
            df = df.drop(columns=cols_to_drop)

        return df

    def _export_excel(self, df: "pd.DataFrame", output_path: str) -> None:
        """Export DataFrame to styled Excel with merged source headers"""
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        # Save basic Excel first with pandas
        df.to_excel(output_path, index=False, sheet_name="Summary")

        # Load with openpyxl to add styling
        wb = load_workbook(output_path)
        ws = wb["Summary"]

        # Build column source mapping
        column_sources = self._build_column_sources(df)

        # Insert source header row at row 1
        ws.insert_rows(1)

        # Style definitions
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        header_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Create merged cells for source headers
        for source_name, start_col, end_col in column_sources:
            if start_col <= end_col:
                # Merge cells (openpyxl uses 1-indexed columns)
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

                # Set value and style
                cell = ws.cell(row=1, column=start_col)
                cell.value = source_name
                cell.alignment = center_align
                cell.font = header_font_white

                # Get color for this source
                source_key = self._get_source_key_from_name(source_name)
                color = self.SOURCE_COLORS.get(source_key, "4472C4")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                # Add border to merged range
                for c in range(start_col, end_col + 1):
                    ws.cell(row=1, column=c).border = thin_border

        # Style the column header row (row 2 now, after insert)
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=2, column=c)
            cell.font = header_font
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.alignment = center_align
            cell.border = thin_border

        # Set row height for header rows
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20

        # Freeze panes (freeze first two rows)
        ws.freeze_panes = "A3"

        # Save the workbook
        wb.save(output_path)

        print("  - Row 1: Data source headers (merged cells)")
        print("  - Row 2: Column names")
        print("  - Row 3+: Test data")

    def _build_column_sources(self, df: "pd.DataFrame") -> List[tuple]:
        """Build column source groups for merging

        Returns: List of (source_name, start_col, end_col) tuples (1-indexed)
        """
        column_sources = []
        current_source = None
        source_start = 1

        for col_idx, col_name in enumerate(df.columns, start=1):
            source = self._find_column_group(col_name)

            if source != current_source:
                # Close previous group
                if current_source is not None:
                    source_display = self._get_source_display_name(current_source)
                    column_sources.append((source_display, source_start, col_idx - 1))

                # Start new group
                current_source = source
                source_start = col_idx

        # Close last group
        if current_source is not None:
            source_display = self._get_source_display_name(current_source)
            column_sources.append((source_display, source_start, len(df.columns)))

        return column_sources

    def _get_source_key_from_name(self, display_name: str) -> str:
        """Get source key from display name"""
        # Reverse mapping
        name_to_key = {
            "Basic": "Basic",
            "Browser": "Browser",
            "VM CPU": "VM_CPU",
            "DevKit TopDown": "DevKit_TopDown",
            "DevKit Memory": "DevKit_Memory",
            "NUMA Bandwidth": "NUMA_Bandwidth",
            "KSys": "KSys",
            "UBWatch Latency": "UBWatch_Latency",
            "UBWatch Bandwidth": "UBWatch_Bandwidth",
            "SMAPBW": "SMAPBW",
            "Getfre": "Getfre",
        }
        return name_to_key.get(display_name, "Basic")

    def _get_source_display_name(self, source: str) -> str:
        """Get human-readable display name for data source"""
        display_names = {
            "Basic": "Basic",
            "Browser": "Browser",
            "VM_CPU": "VM CPU",
            "DevKit_TopDown": "DevKit TopDown",
            "DevKit_Memory": "DevKit Memory",
            "NUMA_Bandwidth": "NUMA Bandwidth",
            "KSys": "KSys",
            "UBWatch_Latency": "UBWatch Latency",
            "UBWatch_Bandwidth": "UBWatch Bandwidth",
            "SMAPBW": "SMAPBW",
            "Getfre": "Getfre",
        }
        return display_names.get(source, source.replace("_", " "))

    def _find_column_group(self, col_name: str) -> str:
        """Find which group a column belongs to"""
        # Check defined groups first
        for group, columns in self.COLUMN_GROUPS.items():
            if col_name in columns:
                return group

        # Infer group from column name prefix
        if col_name.startswith("Browser_"):
            return "Browser"
        elif col_name.startswith("VM_CPU"):
            return "VM_CPU"
        elif col_name.startswith("DevKit_TopDown"):
            return "DevKit_TopDown"
        elif col_name.startswith("DevKit_Memory"):
            return "DevKit_Memory"
        elif col_name.startswith("NUMA_Bandwidth"):
            return "NUMA_Bandwidth"
        elif col_name.startswith("KSys"):
            return "KSys"
        elif col_name.startswith("UBWatch_Latency"):
            return "UBWatch_Latency"
        elif col_name.startswith("UBWatch_Bandwidth"):
            return "UBWatch_Bandwidth"
        elif col_name.startswith("SMAPBW"):
            return "SMAPBW"
        elif col_name.startswith("Getfre"):
            return "Getfre"

        return "Basic"
