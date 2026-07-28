#!/usr/bin/env python3
"""
Shared utilities for batch snapshot scripts.

Provides .env loading, E2B config.json credential loading,
statistics computation, and Excel report generation.
"""

import json
import math
import os
from pathlib import Path
from typing import Any, Dict, List, Optional


def load_env(env_path: str, config_path: str, cli_api_key: str = None, cli_access_token: str = None) -> dict:
    """Load E2B credentials and config from config.json, .env file, and CLI args.

    Priority: CLI args > .env > config.json.

    config.json provides: accessToken, teamApiKey.
    .env provides: E2B_API_URL (required), optional E2B_HTTP_SSL, E2B_DOMAIN,
                   and optional E2B_API_KEY / E2B_ACCESS_TOKEN overrides.

    Args:
        env_path: Path to .env file.
        config_path: Path to ~/.e2b/config.json.
        cli_api_key: CLI --api-key override (highest priority).
        cli_access_token: CLI --access-token override (highest priority).

    Returns:
        Dict with: e2b_api_url, e2b_api_key, e2b_access_token,
        e2b_http_ssl, e2b_domain, base_url, headers.
    """
    e2b_api_key = ""
    e2b_access_token = ""
    e2b_api_url = ""
    e2b_http_ssl = "false"
    e2b_domain = "e2b.app"

    # 1. Load from config.json (baseline)
    config_file = Path(config_path)
    if config_file.exists():
        print(f"  Loading credentials from {config_path}")
        with open(config_file, encoding="utf-8") as f:
            data = json.load(f)
        e2b_api_key = data.get("teamApiKey", "")
        e2b_access_token = data.get("accessToken", "")
    else:
        print(f"  WARNING: Config file not found: {config_path}")

    # 2. Load from .env file (overrides config.json values)
    env_file = Path(env_path)
    if env_file.exists():
        print(f"  Loading config from {env_path}")
        with open(env_file, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    key, _, value = line.partition("=")
                    key = key.strip()
                    value = value.strip()
                    # Remove surrounding quotes
                    if len(value) >= 2 and (
                        (value.startswith('"') and value.endswith('"'))
                        or (value.startswith("'") and value.endswith("'"))
                    ):
                        value = value[1:-1]

                    if key == "E2B_API_URL":
                        e2b_api_url = value
                    elif key == "E2B_API_KEY":
                        e2b_api_key = value
                    elif key == "E2B_ACCESS_TOKEN":
                        e2b_access_token = value
                    elif key == "E2B_HTTP_SSL":
                        e2b_http_ssl = value
                    elif key == "E2B_DOMAIN":
                        e2b_domain = value
    else:
        print(f"  WARNING: .env file not found: {env_path}")

    # 3. CLI overrides (highest priority)
    if cli_api_key:
        e2b_api_key = cli_api_key
    if cli_access_token:
        e2b_access_token = cli_access_token

    # 4. Validate required fields
    if not e2b_api_url:
        raise ValueError("E2B_API_URL must be set in .env file or environment")
    if not e2b_api_key:
        raise ValueError("E2B_API_KEY not found in config.json, .env, or CLI args")
    if not e2b_access_token:
        raise ValueError("E2B_ACCESS_TOKEN not found in config.json, .env, or CLI args")

    # 5. Auto-detect SSL from URL
    if e2b_api_url.startswith("http://"):
        e2b_http_ssl = "false"
    elif e2b_api_url.startswith("https://"):
        e2b_http_ssl = "true"

    # 6. Set environment variables for E2B SDK
    os.environ["E2B_API_URL"] = e2b_api_url
    os.environ["E2B_API_KEY"] = e2b_api_key
    os.environ["E2B_ACCESS_TOKEN"] = e2b_access_token
    os.environ["E2B_HTTP_SSL"] = e2b_http_ssl
    os.environ["E2B_DOMAIN"] = e2b_domain

    base_url = e2b_api_url.rstrip("/")
    headers = {
        "X-API-Key": e2b_api_key,
        "Authorization": f"Bearer {e2b_access_token}",
        "Content-Type": "application/json",
    }

    return {
        "e2b_api_url": e2b_api_url,
        "e2b_api_key": e2b_api_key,
        "e2b_access_token": e2b_access_token,
        "e2b_http_ssl": e2b_http_ssl,
        "e2b_domain": e2b_domain,
        "base_url": base_url,
        "headers": headers,
    }


def compute_stats(values: List[float]) -> Dict[str, Any]:
    """Compute summary statistics for a list of timing values.

    Uses percentile calculation without numpy dependency.

    Args:
        values: List of timing values (only successful operations).

    Returns:
        Dict with: count, avg, min, max, p50, p90, p99, std.
        Returns zeroed dict if values is empty.
    """
    if not values:
        return {
            "count": 0,
            "avg": 0.0,
            "min": 0.0,
            "max": 0.0,
            "p50": 0.0,
            "p90": 0.0,
            "p99": 0.0,
            "std": 0.0,
        }

    sorted_vals = sorted(values)
    n = len(sorted_vals)

    def percentile(p: float) -> float:
        """Compute percentile using linear interpolation."""
        k = (n - 1) * p / 100.0
        f = math.floor(k)
        c = math.ceil(k)
        if f == c:
            return sorted_vals[int(k)]
        return sorted_vals[int(f)] * (c - k) + sorted_vals[int(c)] * (k - f)

    mean = sum(sorted_vals) / n
    variance = sum((x - mean) ** 2 for x in sorted_vals) / n if n > 0 else 0.0

    return {
        "count": n,
        "avg": round(mean, 4),
        "min": round(sorted_vals[0], 4),
        "max": round(sorted_vals[-1], 4),
        "p50": round(percentile(50), 4),
        "p90": round(percentile(90), 4),
        "p99": round(percentile(99), 4),
        "std": round(math.sqrt(variance), 4),
    }


def write_excel_report(
    raw_data: List[Dict[str, Any]],
    summary_data: Dict[str, Dict[str, Any]],
    snapshots_data: List[Dict[str, Any]],
    output_path: str,
    report_title: str,
) -> None:
    """Write 3-sheet Excel report with raw data, summary stats, and snapshot registry.

    Args:
        raw_data: Per-instance timing records (list of dicts).
        summary_data: Statistics keyed by column name, each value is a dict
                      from compute_stats() plus 'success_rate'.
        snapshots_data: Snapshot ID registry records.
        output_path: Path to write the .xlsx file.
        report_title: Title displayed in the Summary sheet header.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    # --- Sheet 1: Raw ---
    ws_raw = wb.active
    ws_raw.title = "Raw"

    if raw_data:
        # Write headers from first record's keys
        headers = list(raw_data[0].keys())
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, 1):
            cell = ws_raw.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Write data rows
        for row_idx, record in enumerate(raw_data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws_raw.cell(row=row_idx, column=col_idx, value=record.get(header, ""))
                cell.alignment = center_align

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, len(raw_data) + 2):
                cell_val = str(raw_data[row_idx - 2].get(header, ""))
                max_len = max(max_len, len(cell_val))
            ws_raw.column_dimensions[ws_raw.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    # --- Sheet 2: Summary ---
    ws_summary = wb.create_sheet("Summary")

    # Title row
    title_font = Font(bold=True, size=14)
    ws_summary.cell(row=1, column=1, value=report_title).font = title_font

    # Build summary table
    if summary_data:
        # Column headers: metric name + one column per data series
        series_names = list(summary_data.keys())
        stat_keys = ["count", "success_rate", "avg", "min", "max", "p50", "p90", "p99", "std"]

        # Header row
        ws_summary.cell(row=3, column=1, value="metric").font = Font(bold=True)
        ws_summary.cell(row=3, column=1).fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        for col_idx, name in enumerate(series_names, 2):
            cell = ws_summary.cell(row=3, column=col_idx, value=name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.alignment = center_align

        # Data rows
        for row_idx, stat_key in enumerate(stat_keys, 4):
            ws_summary.cell(row=row_idx, column=1, value=stat_key)
            for col_idx, name in enumerate(series_names, 2):
                stats = summary_data[name]
                value = stats.get(stat_key, "")
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center_align

        # Auto-width
        for col_idx in range(1, len(series_names) + 2):
            ws_summary.column_dimensions[ws_summary.cell(row=3, column=col_idx).column_letter].width = 20

    # --- Sheet 3: Snapshots ---
    ws_snapshots = wb.create_sheet("Snapshots")

    if snapshots_data:
        snap_headers = list(snapshots_data[0].keys())
        for col_idx, header in enumerate(snap_headers, 1):
            cell = ws_snapshots.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row_idx, record in enumerate(snapshots_data, 2):
            for col_idx, header in enumerate(snap_headers, 1):
                cell = ws_snapshots.cell(row=row_idx, column=col_idx, value=record.get(header, ""))
                cell.alignment = center_align

        for col_idx, header in enumerate(snap_headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, len(snapshots_data) + 2):
                cell_val = str(snapshots_data[row_idx - 2].get(header, ""))
                max_len = max(max_len, len(cell_val))
            ws_snapshots.column_dimensions[ws_snapshots.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f"  Excel report saved to {output_path}")
